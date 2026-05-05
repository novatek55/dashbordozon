from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("REPORT_ROW_SOURCES.xlsx")


ROWS = [
    {
        "group": "План и продажи",
        "row_name": "Озон",
        "formula": "Секция, без расчета",
        "tables": "-",
        "fields": "-",
        "status": "Справочная строка",
        "notes": "",
    },
    {
        "group": "План и продажи",
        "row_name": "Выручка план",
        "formula": "Накопительный план по месяцу",
        "tables": "finance_month_plan",
        "fields": "revenue_plan; если записи нет, используется PLAN_BASE_VALUES['revenue_mp']",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "План и продажи",
        "row_name": "заказано",
        "formula": "Сумма quantity по доставленным и не возвращенным отправлениям из 'Доставка покупателю'",
        "tables": "fact_order_items; fallback transactions.raw_data.items; fallback posting_transaction_snapshots",
        "fields": "fact_order_items.quantity, posting_number; raw_data.items[*].quantity; response_json.items[*]",
        "status": "Используется",
        "notes": "Фильтруется через fact_orders.status='Доставлен' и исключение returns/returns_fbo",
    },
    {
        "group": "План и продажи",
        "row_name": "Выручка накопительно",
        "formula": "Накопительная сумма строки 'выручка / продажи'",
        "tables": "Производная",
        "fields": "revenue_sales",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "План и продажи",
        "row_name": "выручка / продажи",
        "formula": "выручка - Возврат выручки",
        "tables": "Производная",
        "fields": "revenue - returns_revenue",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "План и продажи",
        "row_name": "выручка",
        "formula": "По 'Доставка покупателю' добавляется max(raw_data.accruals_for_sale, 0)",
        "tables": "transactions",
        "fields": "description, raw_data.accruals_for_sale, posting_number, operation_date",
        "status": "Используется",
        "notes": "Сейчас не использует данные из Баланса магазина",
    },
    {
        "group": "Возвраты и комиссия",
        "row_name": "Возвраты",
        "formula": "Равна строке 'Возврат выручки'",
        "tables": "Производная",
        "fields": "returns_revenue",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Возвраты и комиссия",
        "row_name": "Возврат выручки",
        "formula": "По возврату добавляется abs(min(raw_data.accruals_for_sale, 0))",
        "tables": "transactions",
        "fields": "description, raw_data.accruals_for_sale, operation_date",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Возвраты и комиссия",
        "row_name": "Вознаграждение Ozon",
        "formula": "Вознаграждение за продажу - Возврат вознаграждения",
        "tables": "Производная",
        "fields": "sale_commission - return_commission",
        "status": "Производная",
        "notes": "Сейчас не равна fee из Баланса магазина",
    },
    {
        "group": "Возвраты и комиссия",
        "row_name": "Вознаграждение за продажу",
        "formula": "По 'Доставка покупателю' добавляется abs(min(raw_data.sale_commission, 0))",
        "tables": "transactions",
        "fields": "description, raw_data.sale_commission, posting_number, operation_date",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Возвраты и комиссия",
        "row_name": "Возврат вознаграждения",
        "formula": "По возврату добавляется max(raw_data.sale_commission, 0)",
        "tables": "transactions",
        "fields": "description, raw_data.sale_commission, operation_date",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Доставка",
        "row_name": "Услуги доставки",
        "formula": "pickup_processing + courier_departure + dropoff_processing + logistics + reverse_logistics + pickup_courier_delivery",
        "tables": "Производная",
        "fields": "Сумма 6 строк доставки",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Доставка",
        "row_name": "Обработка отправления Pick-up",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Доставка",
        "row_name": "Организация выезда курьера",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Доставка",
        "row_name": "Обработка отправления Drop-off",
        "formula": "Сервисы MarketplaceServiceItemDropoffPVZ и MarketplaceServiceItemDropoffSC",
        "tables": "transactions.raw_data.services",
        "fields": "services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Доставка",
        "row_name": "Логистика",
        "formula": "Сервис MarketplaceServiceItemDirectFlowLogistic",
        "tables": "transactions.raw_data.services",
        "fields": "services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Доставка",
        "row_name": "Обратная логистика",
        "formula": "По возвратной доставке: сумма сервисов + остаток abs(amount) - matched_total",
        "tables": "transactions; transactions.raw_data.services",
        "fields": "amount, services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Доставка",
        "row_name": "Доставка курьером Pick-up",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Услуги агентов",
        "formula": "partner_returns_processing + star_products + temporary_partner_storage + partner_dropoff_processing + delivery_to_pickup + acquiring",
        "tables": "Производная",
        "fields": "Сумма 6 строк агентских услуг",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Обработка возвратов, отмен и невыкупов партнёрами",
        "formula": "Сервис MarketplaceServiceItemRedistributionReturnsPVZ",
        "tables": "transactions.raw_data.services",
        "fields": "services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Звёздные товары",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Временное размещение товара партнерами",
        "formula": "Сервис MarketplaceServiceItemTemporaryStorageRedistribution или description == 'Временное размещение товара партнерами'",
        "tables": "transactions; transactions.raw_data.services",
        "fields": "description, amount, services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Обработка отправления Drop-off партнёрами (АПВЗ)",
        "formula": "Сервис MarketplaceServiceItemRedistributionDropOffApvz",
        "tables": "transactions.raw_data.services",
        "fields": "services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Доставка до места выдачи",
        "formula": "Сервис MarketplaceServiceItemRedistributionLastMileCourier",
        "tables": "transactions.raw_data.services",
        "fields": "services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Агентские услуги",
        "row_name": "Эквайринг",
        "formula": "description == 'Оплата эквайринга' или сервис MarketplaceRedistributionOfAcquiringOperation",
        "tables": "transactions; transactions.raw_data.services",
        "fields": "amount, services[*].name, services[*].price",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "FBO услуги",
        "row_name": "Услуги FBO",
        "formula": "cross_docking + valid_preparation + ozon_delivery_to_pvz + warehouse_placement + piece_acceptance + zone_sorting + excess_processing",
        "tables": "Производная",
        "fields": "Сумма 7 FBO строк",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "FBO услуги",
        "row_name": "Кросс-докинг",
        "formula": "description == 'Кросс-докинг'",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "FBO услуги",
        "row_name": "Подготовка товара к вывозу: Валид",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "FBO услуги",
        "row_name": "Вывоз товара со склада силами Ozon: Доставка до ПВЗ",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "FBO услуги",
        "row_name": "Размещение товаров на складах Ozon",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "FBO услуги",
        "row_name": "Обработка товара в составе грузоместа: Поштучная приёмка",
        "formula": "description.startswith('Обработка товара в составе грузоместа на FBO')",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "FBO услуги",
        "row_name": "Обработка товара в составе грузоместа: Сортировка по зонам размещения",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "FBO услуги",
        "row_name": "Обработка опознанных излишков в составе грузоместа",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Продвижение и реклама",
        "formula": "seller_bonus_mailing + seller_bonus + premium_subscription + premium_plus_subscription + pay_per_click + review_pin + top_search + internet_ads + review_points",
        "tables": "Производная",
        "fields": "Сумма 9 строк продвижения",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Продвижение",
        "row_name": "Бонусы продавца - рассылка",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Бонусы продавца",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Подписка Premium",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Подписка Premium Plus",
        "formula": "description == 'Подписка Premium Plus'",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Продвижение",
        "row_name": "Оплата за клик",
        "formula": "description == 'Оплата за клик'",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Продвижение",
        "row_name": "Закрепление отзыва",
        "formula": "description == 'Закрепление отзыва'",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Продвижение",
        "row_name": "Вывод в топ",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Реклама в сети интернет на сайте",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Продвижение",
        "row_name": "Баллы за отзыв",
        "formula": "description == 'Баллы за отзывы'",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Прочее",
        "row_name": "Другие услуги",
        "formula": "operational_errors + temporary_sc_storage + utilization + loans_factoring + other_accrual_adjustments + shortage_retention + compensations",
        "tables": "Производная",
        "fields": "Сумма блока прочих строк",
        "status": "Производная",
        "notes": "Хвост из transactions сначала накапливается, но затем строка перезаписывается формулой",
    },
    {
        "group": "Прочее",
        "row_name": "Расходы маркетплейса (все)",
        "formula": "returns_revenue + ozon_fee_total + delivery_services_total + agent_services_total + fbo_services_total + promotion_total + other_services",
        "tables": "Производная",
        "fields": "Сумма строк расходов с возвратами",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Прочее",
        "row_name": "Обработка операционных ошибок продавца",
        "formula": "description.startswith('Обработка операционных ошибок продавца') или description.startswith('Жалобы покупателей')",
        "tables": "transactions",
        "fields": "description, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Прочее",
        "row_name": "Временное размещение товара в СЦ/ПВЗ",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Прочее",
        "row_name": "Утилизация товаров",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Прочее",
        "row_name": "Займы и факторинг",
        "formula": "Сейчас не заполняется",
        "tables": "-",
        "fields": "-",
        "status": "Всегда 0",
        "notes": "Нет маппинга",
    },
    {
        "group": "Прочее",
        "row_name": "Прочие начисления - Корректировка стоимости услуг",
        "formula": "Берется из report_compensation_items, если article_name содержит 'корректировк' и 'услуг'",
        "tables": "report_compensation_items",
        "fields": "effective_date, article_name, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Прочее",
        "row_name": "Удержание за недовложение товара",
        "formula": "Берется из report_compensation_items, если article_name содержит 'недовлож'",
        "tables": "report_compensation_items",
        "fields": "effective_date, article_name, amount",
        "status": "Используется",
        "notes": "",
    },
    {
        "group": "Прочее",
        "row_name": "Компенсации и декомпенсации",
        "formula": "Все прочие строки report_compensation_items",
        "tables": "report_compensation_items",
        "fields": "effective_date, article_name, amount",
        "status": "Используется",
        "notes": "Транзакции с 'компенсац' и 'потеря по вине ozon' специально пропускаются",
    },
    {
        "group": "Итоги",
        "row_name": "Расходы МП",
        "formula": "ozon_fee_total + delivery_services_total + agent_services_total + fbo_services_total + promotion_total + other_services",
        "tables": "Производная",
        "fields": "Сумма строк расходов МП",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Расходы МП, %",
        "formula": "marketplace_expenses / revenue_sales",
        "tables": "Производная",
        "fields": "marketplace_expenses, revenue_sales",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Маркетинг, %",
        "formula": "(pay_per_click + review_pin + top_search + internet_ads) / revenue_sales",
        "tables": "Производная",
        "fields": "pay_per_click, review_pin, top_search, internet_ads, revenue_sales",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Начислено",
        "formula": "revenue_sales - returns_total - marketplace_expenses",
        "tables": "Производная",
        "fields": "revenue_sales, returns_total, marketplace_expenses",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Себестоимость",
        "formula": "По продажам quantity * unit_cost; по возвратам вычитается с минусом",
        "tables": "fact_order_items; transactions.raw_data.items; posting_transaction_snapshots; returns; returns_fbo; finance_article_costs",
        "fields": "quantity, offer_id, sku, returned_at, unit_cost",
        "status": "Используется",
        "notes": "Сначала поиск себестоимости по sku, потом по article/offer_id",
    },
    {
        "group": "Итоги",
        "row_name": "Валовая прибыль",
        "formula": "accrued - material_cost",
        "tables": "Производная",
        "fields": "accrued, material_cost",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Валовая прибыль, % к OZ",
        "formula": "gross_profit / revenue_sales",
        "tables": "Производная",
        "fields": "gross_profit, revenue_sales",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Валовая прибыль, % к РС",
        "formula": "gross_profit / accrued",
        "tables": "Производная",
        "fields": "gross_profit, accrued",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Валовая накопительно",
        "formula": "Накопительная сумма строки 'Валовая прибыль'",
        "tables": "Производная",
        "fields": "gross_profit",
        "status": "Производная",
        "notes": "",
    },
    {
        "group": "Итоги",
        "row_name": "Валовая план",
        "formula": "Накопительный план от PLAN_BASE_VALUES['gross_profit'], масштабированный к месячному плану выручки",
        "tables": "finance_month_plan + код",
        "fields": "revenue_plan, PLAN_BASE_VALUES['gross_profit']",
        "status": "Используется",
        "notes": "",
    },
]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Источники строк"

    headers = ["Группа", "Строка отчета", "Формула сейчас", "Таблица/источник", "Поля/столбцы", "Статус", "Примечание"]
    ws.append(headers)

    for row in ROWS:
        ws.append(
            [
                row["group"],
                row["row_name"],
                row["formula"],
                row["tables"],
                row["fields"],
                row["status"],
                row["notes"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="8C5C2A")
    header_font = Font(color="FFFFFF", bold=True)
    group_fills = {
        "План и продажи": PatternFill("solid", fgColor="F8E7D1"),
        "Возвраты и комиссия": PatternFill("solid", fgColor="FCE4D6"),
        "Доставка": PatternFill("solid", fgColor="E2F0D9"),
        "Агентские услуги": PatternFill("solid", fgColor="DDEBF7"),
        "FBO услуги": PatternFill("solid", fgColor="E4DFEC"),
        "Продвижение": PatternFill("solid", fgColor="FFF2CC"),
        "Прочее": PatternFill("solid", fgColor="F2F2F2"),
        "Итоги": PatternFill("solid", fgColor="D9EAD3"),
    }
    zero_fill = PatternFill("solid", fgColor="F4CCCC")
    derived_fill = PatternFill("solid", fgColor="D9EAD3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        group_name = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=6).value
        fill = group_fills.get(group_name)

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

        if status == "Всегда 0":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = zero_fill
        elif status == "Производная":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = derived_fill

    widths = {
        1: 20,
        2: 38,
        3: 58,
        4: 42,
        5: 55,
        6: 16,
        7: 42,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28

    legend = wb.create_sheet("Легенда")
    legend.append(["Статус", "Смысл"])
    legend.append(["Используется", "Строка заполняется прямыми данными из БД/API"])
    legend.append(["Производная", "Строка вычисляется из других строк отчета"])
    legend.append(["Всегда 0", "Сейчас источник или маппинг для строки отсутствует"])
    legend.append(["Справочная строка", "Техническая или визуальная строка без расчета"])
    for cell in legend[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, width in {1: 18, 2: 70}.items():
        legend.column_dimensions[get_column_letter(col_idx)].width = width
    for row in range(2, 6):
        for col in range(1, 3):
            legend.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()

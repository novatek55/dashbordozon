"""Импорт прайса логистики Ozon FBO/FBS из xlsx в таблицу logistics_tariffs.

Формат входа: файлы `logistika-fbo-fbs-*.xlsx`.
Листы:
  - «Тарифы на логистику» — (Объём, Кластер отправки, Кластер назначения, до 300₽, свыше 300₽)
  - «Универсальные тарифы» — (Объём, до 300₽, свыше 300₽) — fallback, пишутся с cluster_from='*'/cluster_to='*'

Политика перезагрузки: TRUNCATE + bulk INSERT (идемпотентно относительно файла).

Запуск: python -m src.import_logistics_tariffs <path/to/logistika-*.xlsx>
"""
from __future__ import annotations

import asyncio
import logging
import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import openpyxl
from sqlalchemy import delete

from src.database import db_manager
from src.models import LogisticsTariff

logger = logging.getLogger(__name__)


_VOLUME_RE = re.compile(r"([\d.,]+)\s*[-–—]\s*([\d.,]+)")
_VOLUME_FROM_RE = re.compile(r"[Оо]т\s+([\d.,]+)")
_OPEN_BUCKET_MAX = Decimal("99999")


def _parse_volume_bucket(text: Optional[str]) -> Optional[Tuple[Decimal, Decimal]]:
    """'0-0,200 л' → (0, 0.200); 'От 800,001 л' → (800.001, 99999)."""
    if not text:
        return None
    s = str(text)
    m = _VOLUME_RE.search(s)
    if m:
        lo = Decimal(m.group(1).replace(",", "."))
        hi = Decimal(m.group(2).replace(",", "."))
        return lo, hi
    m = _VOLUME_FROM_RE.search(s)
    if m:
        lo = Decimal(m.group(1).replace(",", "."))
        return lo, _OPEN_BUCKET_MAX
    return None


def _iter_pair_rows(ws) -> Iterable[Tuple[str, str, str, float, float]]:
    """Лист парных тарифов. Шапка в строке 3 (индекс 2), данные — с 4-й."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx < 3:
            continue
        # row: (None, volume, cluster_from, cluster_to, under300, over300, ...)
        if not row or len(row) < 6:
            continue
        volume, c_from, c_to, under, over = row[1], row[2], row[3], row[4], row[5]
        if volume is None or c_from is None or c_to is None:
            continue
        if under is None or over is None:
            continue
        yield volume, c_from, c_to, under, over


def _iter_universal_rows(ws) -> Iterable[Tuple[str, float, float]]:
    """Лист универсальных тарифов. Шапка в строке 3 (индекс 2)."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx < 3:
            continue
        if not row or len(row) < 4:
            continue
        volume, under, over = row[1], row[2], row[3]
        if volume is None or under is None or over is None:
            continue
        yield volume, under, over


def build_records(xlsx_path: Path) -> List[dict]:
    """Читает xlsx и возвращает список dict'ов для bulk insert."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    records: List[dict] = []
    source_file = xlsx_path.name

    # Парные тарифы
    pair_sheet = None
    for name in wb.sheetnames:
        if "Тарифы на логистику" in name:
            pair_sheet = name
            break
    if pair_sheet:
        skipped = 0
        for volume, c_from, c_to, under, over in _iter_pair_rows(wb[pair_sheet]):
            bucket = _parse_volume_bucket(volume)
            if bucket is None:
                skipped += 1
                continue
            records.append({
                "cluster_from": str(c_from).strip(),
                "cluster_to": str(c_to).strip(),
                "volume_min_l": bucket[0],
                "volume_max_l": bucket[1],
                "price_under_300": Decimal(str(under)),
                "price_over_300": Decimal(str(over)),
                "source_file": source_file,
            })
        if skipped:
            logger.warning("Pair sheet: skipped %d rows (unparseable volume)", skipped)

    # Универсальные
    uni_sheet = None
    for name in wb.sheetnames:
        if "ниверсальные" in name:
            uni_sheet = name
            break
    if uni_sheet:
        skipped = 0
        for volume, under, over in _iter_universal_rows(wb[uni_sheet]):
            bucket = _parse_volume_bucket(volume)
            if bucket is None:
                skipped += 1
                continue
            records.append({
                "cluster_from": "*",
                "cluster_to": "*",
                "volume_min_l": bucket[0],
                "volume_max_l": bucket[1],
                "price_under_300": Decimal(str(under)),
                "price_over_300": Decimal(str(over)),
                "source_file": source_file,
            })
        if skipped:
            logger.warning("Universal sheet: skipped %d rows", skipped)

    wb.close()
    return records


async def import_tariffs(xlsx_path: Path) -> int:
    """TRUNCATE + bulk INSERT. Возвращает число вставленных строк."""
    records = build_records(xlsx_path)
    if not records:
        raise ValueError(f"Не распознано ни одной строки тарифа в {xlsx_path}")

    await db_manager.initialize()
    try:
        async with db_manager.session() as session:
            await session.execute(delete(LogisticsTariff))
            session.add_all([LogisticsTariff(**r) for r in records])
        logger.info("Imported %d tariff rows from %s", len(records), xlsx_path.name)
        return len(records)
    finally:
        await db_manager.close()


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    if len(sys.argv) < 2:
        print("Usage: python -m src.import_logistics_tariffs <path/to/logistika-*.xlsx>")
        sys.exit(2)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)
    n = asyncio.run(import_tariffs(path))
    print(f"OK: imported {n} rows")


if __name__ == "__main__":
    main()

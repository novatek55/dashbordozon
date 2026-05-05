"""Menedzher sinhronizacii dannyh iz Ozon API."""
from calendar import monthrange
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any, Tuple
import logging
import csv
import io
import hashlib
import os
from io import BytesIO
from sqlalchemy import select, insert, update, delete, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
import asyncio
import openpyxl
from tenacity import RetryError

from src.ozon_client import OzonClient, OzonAPIError, RateLimitError
from src.database import db_manager
from src.models import (
    Product, Posting, PostingItem, Transaction,
    Campaign, CampaignStatistic, Return, StockHistory,
    SyncLog, AnalyticsData, CampaignDetail, CampaignObject,
    PostingFBO, ReturnFBO, ReturnRFBS, FinanceTransactionTotal,
    CashFlowStatement, MutualSettlement, B2BSale, AnalyticsTurnover,
    AnalyticsAverageDeliveryTime, AnalyticsStockManagement, AnalyticsStock, Warehouse, Cluster, DeliveryMethod,
    SellerRating, SellerRatingHistory, Review, ReviewComment, ReviewRatingSnapshot,
    Question, QuestionAnswer, ChatThread, ChatMessage,
    PromoAction, PromoProduct, AsyncReport, RealizationReport,
    RealizationReportDetail, FactOrder, FactOrderItem, ReportProductItem,
    ReportReturnItem, ReportWarehouseStockItem, ReportCompensationItem, ReportDownloadRetry,
    TransactionItem, TransactionService, FBSWarehouseStock,
    AnalyticsProductQuerySummary, AnalyticsProductQueryDetail,
    StockDailySnapshot, DeliveryTimeDailySnapshot,
    ProductDimension,
)
from sqlalchemy import text as _sql_text
from src.config import settings

logger = logging.getLogger(__name__)
MSK = timezone(timedelta(hours=3))


class SyncManager:
    """Menedzher sinhronizacii dannyh."""
    
    def __init__(self, client: OzonClient):
        self.client = client

    def _iter_month_windows(self, from_date: datetime, to_date: datetime):
        """Razbit' period na okna ne bolee odnogo kalendarnogo mesjaca."""
        cursor = from_date
        while cursor < to_date:
            if cursor.month == 12:
                next_month = datetime(cursor.year + 1, 1, 1, tzinfo=cursor.tzinfo)
            else:
                next_month = datetime(cursor.year, cursor.month + 1, 1, tzinfo=cursor.tzinfo)
            window_end = min(next_month, to_date)
            yield cursor, window_end
            cursor = window_end

    def _iter_report_months(self, from_date: datetime, to_date: datetime):
        """Iteracija po nachalam mesjacev dlja async-otchetov вида YYYY-MM."""
        tz = from_date.tzinfo or timezone.utc
        cursor = datetime(from_date.year, from_date.month, 1, tzinfo=tz)
        end_month = datetime(to_date.year, to_date.month, 1, tzinfo=to_date.tzinfo or tz)
        while cursor <= end_month:
            yield cursor
            if cursor.month == 12:
                cursor = datetime(cursor.year + 1, 1, 1, tzinfo=tz)
            else:
                cursor = datetime(cursor.year, cursor.month + 1, 1, tzinfo=tz)

    def _ensure_aware_utc(self, value: Optional[datetime]) -> Optional[datetime]:
        """Privezenija datetime k aware UTC dlja stabil'noj agregacii po dnju."""
        if value is None:
            return None
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)

    def _iter_day_windows(self, from_date: datetime, to_date: datetime):
        """Итерация по полным суткам [00:00:00..23:59:59] в UTC."""
        cursor = datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc)
        end_day = datetime(to_date.year, to_date.month, to_date.day, tzinfo=timezone.utc)
        while cursor <= end_day:
            window_end = cursor + timedelta(days=1) - timedelta(seconds=1)
            yield cursor, min(window_end, to_date)
            cursor += timedelta(days=1)

    def _iter_week_windows(self, from_date: datetime, to_date: datetime):
        """Итерация по 7-дневным окнам для старшей истории."""
        cursor = datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc)
        while cursor <= to_date:
            window_end = min(cursor + timedelta(days=7) - timedelta(seconds=1), to_date)
            yield cursor, window_end
            cursor = window_end + timedelta(seconds=1)

    async def _ensure_analytics_product_queries_schema(self) -> None:
        """Минимальная миграция таблиц product queries для PostgreSQL."""
        if "postgresql" not in settings.database_url.lower():
            return

        create_summary_sql = """
            CREATE TABLE IF NOT EXISTS analytics_product_query_summary (
                id BIGSERIAL PRIMARY KEY,
                period_start TIMESTAMPTZ NOT NULL,
                period_end TIMESTAMPTZ NOT NULL,
                granularity VARCHAR(16) NOT NULL DEFAULT 'day',
                sku BIGINT NOT NULL,
                offer_id VARCHAR(255),
                product_name VARCHAR(1000),
                searches INTEGER,
                views INTEGER,
                avg_position DOUBLE PRECISION,
                conversion NUMERIC(10, 4),
                gmv NUMERIC(15, 2),
                raw_data JSONB,
                last_synced_at TIMESTAMPTZ DEFAULT now()
            );
        """
        create_details_sql = """
            CREATE TABLE IF NOT EXISTS analytics_product_query_details (
                id BIGSERIAL PRIMARY KEY,
                period_start TIMESTAMPTZ NOT NULL,
                period_end TIMESTAMPTZ NOT NULL,
                granularity VARCHAR(16) NOT NULL DEFAULT 'day',
                sku BIGINT NOT NULL,
                offer_id VARCHAR(255),
                product_name VARCHAR(1000),
                query_text VARCHAR(1000) NOT NULL,
                searches INTEGER,
                views INTEGER,
                avg_position DOUBLE PRECISION,
                conversion NUMERIC(10, 4),
                gmv NUMERIC(15, 2),
                raw_data JSONB,
                last_synced_at TIMESTAMPTZ DEFAULT now()
            );
        """
        alter_sql = [
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_apq_summary_period_granularity_sku_idx ON analytics_product_query_summary(period_start, period_end, granularity, sku);",
            "CREATE INDEX IF NOT EXISTS idx_apq_summary_sku_period ON analytics_product_query_summary(sku, period_start);",
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_apq_details_period_granularity_sku_query_idx ON analytics_product_query_details(period_start, period_end, granularity, sku, query_text);",
            "CREATE INDEX IF NOT EXISTS idx_apq_details_sku_period ON analytics_product_query_details(sku, period_start);",
            "CREATE INDEX IF NOT EXISTS idx_apq_details_query ON analytics_product_query_details(query_text);",
        ]

        async with db_manager.session() as session:
            await session.execute(text(create_summary_sql))
            await session.execute(text(create_details_sql))
            for stmt in alter_sql:
                await session.execute(text(stmt))

    async def _load_product_query_sku_reference(self) -> Dict[int, Dict[str, Any]]:
        """Загружает последнее сопоставление SKU -> offer_id/product_name из отчета по товарам."""
        async with db_manager.session() as session:
            rows = await session.execute(
                text(
                    """
                    WITH src AS (
                        SELECT fbo_sku_id::bigint AS sku, trim(offer_id) AS offer_id, product_name, last_synced_at
                        FROM report_products_items
                        WHERE fbo_sku_id IS NOT NULL
                        UNION ALL
                        SELECT fbs_sku_id::bigint AS sku, trim(offer_id) AS offer_id, product_name, last_synced_at
                        FROM report_products_items
                        WHERE fbs_sku_id IS NOT NULL
                        UNION ALL
                        SELECT sku::bigint AS sku, trim(offer_id) AS offer_id, name AS product_name, last_synced_at
                        FROM analytics_stocks
                        WHERE sku IS NOT NULL
                        UNION ALL
                        SELECT sku::bigint AS sku, trim(offer_id) AS offer_id, product_name, last_synced_at
                        FROM report_warehouse_stock_items
                        WHERE sku IS NOT NULL
                        UNION ALL
                        SELECT sku::bigint AS sku, trim(offer_id) AS offer_id, NULL::varchar AS product_name, last_synced_at
                        FROM fbs_warehouse_stocks
                        WHERE sku IS NOT NULL
                    )
                    SELECT DISTINCT ON (sku)
                        sku,
                        offer_id,
                        product_name
                    FROM src
                    WHERE sku IS NOT NULL AND sku > 0
                    ORDER BY sku, last_synced_at DESC NULLS LAST
                    """
                )
            )
            result: Dict[int, Dict[str, Any]] = {}
            for row in rows.mappings():
                sku = self._parse_int_flexible(row.get("sku"))
                if not sku:
                    continue
                result[int(sku)] = {
                    "offer_id": str(row.get("offer_id") or "").strip() or None,
                    "product_name": str(row.get("product_name") or "").strip() or None,
                }
            return result

    async def _fetch_analytics_stocks_chunk_resilient(
        self,
        skus: List[int],
    ) -> Tuple[List[Dict[str, Any]], List[int]]:
        """Fetch analytics stocks with batch splitting on transient/server failures."""
        if not skus:
            return [], []

        try:
            result = await self.client.get_analytics_stocks(skus)
            items = result.get("items", []) if isinstance(result, dict) else []
            return items, []
        except OzonAPIError as exc:
            is_retryable_server_error = exc.status_code in {500, 502, 503, 504}
            if not is_retryable_server_error:
                raise
            if len(skus) == 1:
                logger.error(
                    "analytics_stocks: skipping SKU %s after server error: %s",
                    skus[0],
                    exc,
                )
                return [], [skus[0]]

            middle = len(skus) // 2
            left_skus = skus[:middle]
            right_skus = skus[middle:]
            logger.warning(
                "analytics_stocks: server error for chunk of %s SKU(s), splitting into %s and %s",
                len(skus),
                len(left_skus),
                len(right_skus),
            )
            left_items, left_skipped = await self._fetch_analytics_stocks_chunk_resilient(left_skus)
            right_items, right_skipped = await self._fetch_analytics_stocks_chunk_resilient(right_skus)
            return left_items + right_items, left_skipped + right_skipped

    async def _wait_report_info(
        self,
        report_code: str,
        attempts: int = 45,
        delay_seconds: float = 2.0,
    ) -> Dict[str, Any]:
        """Dozhdat'sja, poka async-otchet stanet gotovym ili zavershitsja s oshibkoj."""
        last_result: Dict[str, Any] = {}
        for _ in range(attempts):
            info = await self.client.get_report_info(report_code)
            result = info.get("result", {}) if isinstance(info, dict) else {}
            last_result = result
            status = str(result.get("status", "")).lower()
            if status in {"success", "failed", "error"}:
                return result
            await asyncio.sleep(delay_seconds)
        return last_result

    def _build_unique_headers(self, raw_headers: List[Any]) -> List[str]:
        """Postroit' stabil'nye unikal'nye imena kolonok dlja xlsx/csv."""
        headers: List[str] = []
        seen: Dict[str, int] = {}
        for idx, value in enumerate(raw_headers):
            base = str(value).strip() if value not in (None, "") else f"column_{idx + 1}"
            count = seen.get(base, 0)
            seen[base] = count + 1
            headers.append(base if count == 0 else f"{base}_{count + 1}")
        return headers

    def _detect_header_row_index(self, sheet_rows: List[List[Any]], header_tokens: List[str]) -> int:
        """Vybrat' naibolee pohozhuju na zagolovok stroku."""
        best_idx = 0
        best_score = float("-inf")
        for idx, values in enumerate(sheet_rows[:25]):
            non_empty = [v for v in values if v not in (None, "")]
            if len(non_empty) < 2:
                continue
            row_text = " | ".join(str(v).strip().lower() for v in non_empty)
            score = float(len(non_empty))
            for token in header_tokens:
                if token in row_text:
                    score += 5.0
            if score > best_score:
                best_idx = idx
                best_score = score
        return best_idx

    def _extract_tabular_rows(self, sheet_rows: List[List[Any]], header_tokens: List[str]) -> List[Dict[str, Any]]:
        """Preobrazovat' tablichnye dannye v spisok slovarej."""
        if not sheet_rows:
            return []
        header_idx = self._detect_header_row_index(sheet_rows, header_tokens)
        headers = self._build_unique_headers(sheet_rows[header_idx])
        rows: List[Dict[str, Any]] = []
        for values in sheet_rows[header_idx + 1:]:
            if not values or all(v in (None, "") for v in values):
                continue
            row_map: Dict[str, Any] = {}
            for idx, value in enumerate(values):
                key = headers[idx] if idx < len(headers) else f"column_{idx + 1}"
                row_map[key] = value
            rows.append(row_map)
        return rows

    def _parse_tabular_report_file(self, file_bytes: bytes, header_tokens: List[str]) -> List[Dict[str, Any]]:
        """Raspakovat' xlsx/csv async-otchet v tablichnyj vid."""
        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            try:
                ws = wb.active
                sheet_rows = [list(row) for row in ws.iter_rows(values_only=True)]
            finally:
                wb.close()
            rows = self._extract_tabular_rows(sheet_rows, header_tokens)
            if rows:
                return rows
        except Exception:
            pass

        text = ""
        for encoding in ("utf-8-sig", "cp1251"):
            try:
                text = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            text = file_bytes.decode("utf-8", errors="replace")

        for delimiter in (";", ",", "\t"):
            try:
                reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=delimiter)
                rows = list(reader)
            except csv.Error:
                continue
            if rows:
                return rows
        return []

    def _pick_amount_by_header(self, row: Dict[str, Any]) -> Optional[float]:
        """Naiti naibolee verojatnuju summu v stroke otcheta."""
        for key, value in row.items():
            if value in (None, ""):
                continue
            key_l = str(key).strip().lower()
            if not any(token in key_l for token in ("сум", "amount", "начисл", "итог", "компенсац", "декомпенсац")):
                continue
            if any(token in key_l for token in ("ндс", "vat", "процент", "%", "колич", "quantity", "курс")):
                continue
            amount = self._parse_decimal_flexible(value)
            if amount is not None:
                return amount
        return None

    def _normalize_compensation_report_row(
        self,
        row: Dict[str, Any],
        report_kind: str,
        report_month: datetime,
    ) -> Optional[Dict[str, Any]]:
        """Svesti syruiu stroku otcheta kompensacij k edinoj strukture."""
        article_name = (
            self._pick_value(
                row,
                [
                    "Статья",
                    "Article",
                    "Основание",
                    "Причина",
                    "Описание",
                    "Операция",
                    "Тип компенсации",
                    "Тип декомпенсации",
                    "Вид компенсации",
                    "Вид декомпенсации",
                    "Наименование услуги",
                ],
            )
            or self._pick_by_contains(
                row,
                [
                    "статья",
                    "article",
                    "основан",
                    "причин",
                    "описан",
                    "операц",
                    "тип компенсац",
                    "тип декомпенсац",
                    "вид компенсац",
                    "вид декомпенсац",
                    "услуг",
                ],
            )
        )
        product_name = (
            self._pick_value(row, ["Название товара", "Наименование товара", "Product name", "Name"])
            or self._pick_by_contains(row, ["название товара", "наименование товара", "product name"])
        )
        raw_amount = self._pick_amount_by_header(row)
        if raw_amount is None:
            return None

        article_text = str(article_name or "").strip()
        product_text = str(product_name or "").strip()
        row_text_values = [
            str(value).strip().lower()
            for value in row.values()
            if isinstance(value, str) and str(value).strip()
        ]
        posting_number = self._pick_value(row, ["posting_number", "Номер отправления", "Posting number"]) or self._pick_by_contains(
            row, ["номер отправления", "posting"]
        )
        order_id = self._pick_value(row, ["order_id", "Номер заказа", "Order ID"]) or self._pick_by_contains(
            row, ["номер заказа", "order id"]
        )
        offer_id = self._pick_value(row, ["offer_id", "Артикул", "Offer ID"]) or self._pick_by_contains(
            row, ["артикул", "offer id"]
        )
        sku = self._parse_int_flexible(
            self._pick_value(row, ["sku", "SKU", "Ozon SKU ID"])
            or self._pick_by_contains(row, ["sku"])
        )
        if article_text.lower().startswith(("итого", "всего")):
            return None
        if not article_text and product_text.lower().startswith(("итого", "всего")):
            return None
        if any("всего к начислению" in value or "итого к начислению" in value for value in row_text_values):
            return None
        if not article_text and not product_text and not any([posting_number, order_id, offer_id, sku]):
            return None

        effective_date = self._parse_datetime_flexible(
            self._pick_value(
                row,
                ["Дата", "Дата операции", "Дата начисления", "Operation date", "Date", "Период"],
            )
            or self._pick_by_contains(
                row,
                ["дата операции", "дата начисления", "date", "период"],
            )
        )
        if effective_date is None:
            last_day = monthrange(report_month.year, report_month.month)[1]
            effective_date = report_month.replace(day=last_day)
        effective_date = self._ensure_aware_utc(effective_date)

        signed_amount = abs(raw_amount) if report_kind == "compensation" else -abs(raw_amount)
        raw_row = {
            key: (value.isoformat() if isinstance(value, datetime) else value)
            for key, value in row.items()
        }

        return {
            "report_month": self._ensure_aware_utc(report_month),
            "effective_date": effective_date,
            "article_name": article_text or product_text,
            "raw_amount": raw_amount,
            "amount": signed_amount,
            "posting_number": posting_number,
            "order_id": order_id,
            "offer_id": offer_id,
            "sku": sku,
            "product_name": product_text or None,
            "raw_data": raw_row,
        }

    async def _enqueue_report_download_retry(
        self,
        report_code: str,
        report_type: str,
        file_url: Optional[str],
        error_message: str,
        raw_data: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Sohranit' problemnyj async-otchet v ochered' povtornogo skachivanija."""
        async with db_manager.session() as session:
            existing = await session.execute(
                select(ReportDownloadRetry).where(ReportDownloadRetry.report_code == report_code)
            )
            retry_row = existing.scalar_one_or_none()
            if retry_row:
                retry_row.report_type = report_type
                retry_row.file_url = file_url
                retry_row.status = "pending"
                retry_row.attempts = (retry_row.attempts or 0) + 1
                retry_row.last_error = error_message
                retry_row.raw_data = raw_data
                retry_row.updated_at = datetime.now()
            else:
                session.add(
                    ReportDownloadRetry(
                        report_code=report_code,
                        report_type=report_type,
                        file_url=file_url,
                        status="pending",
                        attempts=1,
                        last_error=error_message,
                        raw_data=raw_data,
                    )
                )
    
    async def _create_sync_log(self, entity_type: str) -> SyncLog:
        """Sozdanie zapisi o sinhronizacii."""
        async with db_manager.session() as session:
            sync_log = SyncLog(
                entity_type=entity_type,
                started_at=datetime.now(),
                status="running"
            )
            session.add(sync_log)
            await session.flush()
            return sync_log
    
    async def _update_sync_log(
        self, 
        sync_log: SyncLog, 
        status: str,
        records_processed: int = 0,
        records_inserted: int = 0,
        records_updated: int = 0,
        error_message: Optional[str] = None
    ):
        """Obnovlenie zapisi o sinhronizacii."""
        async with db_manager.session() as session:
            sync_log.status = status
            sync_log.completed_at = datetime.now()
            sync_log.records_processed = records_processed
            sync_log.records_inserted = records_inserted
            sync_log.records_updated = records_updated
            sync_log.error_message = error_message
            await session.merge(sync_log)

    async def _get_last_successful_sync(self, entity_type: str) -> Optional[SyncLog]:
        """Vozvrat poslednej uspeshnoj sinhronizacii po tipu sushhnosti."""
        async with db_manager.session() as session:
            result = await session.execute(
                select(SyncLog)
                .where(
                    SyncLog.entity_type == entity_type,
                    SyncLog.status == "success",
                    SyncLog.completed_at.is_not(None),
                )
                .order_by(SyncLog.completed_at.desc())
                .limit(1)
            )
            return result.scalar_one_or_none()

    async def _skip_recent_async_report_sync_if_fresh(
        self,
        sync_log: SyncLog,
        entity_type: str,
    ) -> Optional[Dict[str, int]]:
        """Propuskaet povtornuju zagruzku async-otcheta, esli on uzhe nedavno sinhronizirovalsja."""
        refresh_hours = max(int(settings.async_report_refresh_hours or 0), 0)
        if refresh_hours <= 0:
            return None

        last_sync = await self._get_last_successful_sync(entity_type)
        if not last_sync or not last_sync.completed_at:
            return None

        completed_at = self._ensure_aware_utc(last_sync.completed_at)
        if completed_at is None:
            return None

        now_utc = datetime.now(timezone.utc)
        age = now_utc - completed_at
        freshness_window = timedelta(hours=refresh_hours)
        if age > freshness_window:
            return None

        message = (
            f"Skipped: last successful sync at {completed_at.isoformat()} "
            f"is within {refresh_hours}h refresh window"
        )
        logger.info("%s for %s", message, entity_type)
        await self._update_sync_log(sync_log, "success", error_message=message)
        return {
            "skipped": 1,
            "reason": "fresh_recent_sync",
            "last_synced_at": completed_at.isoformat(),
            "refresh_hours": refresh_hours,
        }

    async def _skip_recent_campaigns_sync_if_fresh(
        self,
        sync_log: SyncLog,
    ) -> Optional[Dict[str, Any]]:
        """Propuskaet campaigns sync, esli on uzhe uspeshno vypolnjalsja nedavno."""
        refresh_hours = max(int(settings.campaigns_refresh_hours or 0), 0)
        if refresh_hours <= 0:
            return None

        last_sync = await self._get_last_successful_sync("campaigns")
        if not last_sync or not last_sync.completed_at:
            return None

        completed_at = self._ensure_aware_utc(last_sync.completed_at)
        if completed_at is None:
            return None

        now_utc = datetime.now(timezone.utc)
        if now_utc - completed_at > timedelta(hours=refresh_hours):
            return None

        message = (
            f"Skipped campaigns sync: last successful sync at {completed_at.isoformat()} "
            f"is within {refresh_hours}h refresh window"
        )
        logger.info(message)
        await self._update_sync_log(sync_log, "success", error_message=message)
        return {
            "skipped": 1,
            "reason": "fresh_recent_sync",
            "last_synced_at": completed_at.isoformat(),
            "refresh_hours": refresh_hours,
        }
    
    def _parse_datetime(self, value) -> Optional[datetime]:
        """Parser daty iz raznyh formatov."""
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        try:
            # ISO format
            return datetime.fromisoformat(value.replace('Z', '+00:00'))
        except:
            try:
                # Ozon format
                return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S.%fZ")
            except:
                try:
                    # Date only
                    return datetime.strptime(value, "%Y-%m-%d")
                except:
                    return None
    
    def _parse_decimal(self, value) -> Optional[float]:
        """Parser decimal'nyh chisel."""
        if value is None:
            return None
        try:
            return float(value)
        except:
            return None

    def _normalize_return_payload(self, return_data: Dict[str, Any]) -> Dict[str, Any]:
        """Normalizacija struktury vozvrata iz raznyh versij API."""
        product = return_data.get("product", {}) if isinstance(return_data.get("product"), dict) else {}
        visual = return_data.get("visual", {}) if isinstance(return_data.get("visual"), dict) else {}
        status_info = visual.get("status", {}) if isinstance(visual.get("status"), dict) else {}
        logistic = return_data.get("logistic", {}) if isinstance(return_data.get("logistic"), dict) else {}
        price_obj = product.get("price", {}) if isinstance(product.get("price"), dict) else {}

        return {
            "return_id": return_data.get("id"),
            "posting_number": return_data.get("posting_number") or return_data.get("order_number"),
            "sku": product.get("sku") or return_data.get("sku"),
            "offer_id": product.get("offer_id") or return_data.get("offer_id"),
            "product_name": product.get("name") or return_data.get("product_name"),
            "quantity": product.get("quantity") or return_data.get("quantity"),
            "return_reason": return_data.get("return_reason_name") or return_data.get("return_reason"),
            "status": (
                status_info.get("sys_name")
                or status_info.get("display_name")
                or return_data.get("status")
                or return_data.get("status_name")
            ),
            "returned_at": self._parse_datetime(
                logistic.get("return_date")
                or logistic.get("final_moment")
                or return_data.get("returned_at")
                or return_data.get("moment")
            ),
            "refund_amount": self._parse_decimal(
                price_obj.get("price")
                or return_data.get("refund_amount")
            ),
            "schema": (return_data.get("schema") or "").lower(),
        }

    def _report_code_to_id(self, code: str) -> int:
        """Preobrazovanie string-koda otcheta v stabil'nyj int id."""
        digest = hashlib.sha1(code.encode("utf-8")).hexdigest()[:15]
        return int(digest, 16)

    def _parse_datetime_flexible(self, value: Any) -> Optional[datetime]:
        """Parser daty iz raznyh tekstovyh formatov (CSV/ISO)."""
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        if not text:
            return None

        for fmt in (
            "%Y-%m-%dT%H:%M:%S.%fZ",
            "%Y-%m-%dT%H:%M:%SZ",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def _parse_operation_datetime(self, value: Any) -> Optional[datetime]:
        """Парсинг operation_date с нормализацией в UTC.

        В выгрузках Ozon дата операции может приходить:
        - с timezone (ISO8601) — сохраняем корректный момент времени в UTC;
        - без timezone (только дата/локальное время) — трактуем как Europe/Moscow.
        """
        if not value:
            return None
        if isinstance(value, datetime):
            dt = value
        else:
            text = str(value).strip()
            if not text:
                return None
            normalized = text.replace("Z", "+00:00")
            try:
                dt = datetime.fromisoformat(normalized)
            except ValueError:
                dt = self._parse_datetime_flexible(text)
                if dt is None:
                    return None

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=MSK)
        return dt.astimezone(timezone.utc)

    def _parse_decimal_flexible(self, value: Any) -> Optional[float]:
        """Parser decimal s uchetom zapjatoj i probelov."""
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(" ", "").replace("\xa0", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    def _pick_value(self, row: Dict[str, Any], keys: List[str]) -> Any:
        for key in keys:
            if key in row and row.get(key) not in (None, ""):
                return row.get(key)
        return None

    def _pick_by_contains(self, row: Dict[str, Any], patterns: List[str]) -> Any:
        lowered = {str(k).lower(): v for k, v in row.items()}
        for pattern in patterns:
            p = pattern.lower()
            for key_l, value in lowered.items():
                if p in key_l and value not in (None, ""):
                    return value
        return None

    def _parse_int_flexible(self, value: Any) -> Optional[int]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(" ", "").replace("\xa0", "").replace(",", ".")
        try:
            return int(float(text))
        except ValueError:
            return None

    async def _with_rate_limit_retry(self, coro_factory, attempts: int = 5, base_delay: float = 1.0):
        """Retry-wrapper dlja endpointov, kotorye chasto vozvrashhajut 429."""
        last_exc: Optional[Exception] = None
        for attempt in range(1, attempts + 1):
            try:
                return await coro_factory()
            except RateLimitError as e:
                last_exc = e
                if attempt == attempts:
                    break
                delay = base_delay * (2 ** (attempt - 1))
                logger.warning(
                    "Rate limit from Ozon, retrying in %.1f sec (attempt %s/%s)",
                    delay,
                    attempt,
                    attempts,
                )
                await asyncio.sleep(delay)
            except OzonAPIError as e:
                if e.status_code == 429:
                    last_exc = e
                    if attempt == attempts:
                        break
                    delay = base_delay * (2 ** (attempt - 1))
                    logger.warning(
                        "HTTP 429 from Ozon, retrying in %.1f sec (attempt %s/%s)",
                        delay,
                        attempt,
                        attempts,
                    )
                    await asyncio.sleep(delay)
                else:
                    raise
            except RetryError as e:
                root_exc = e.last_attempt.exception() if e.last_attempt else e
                if isinstance(root_exc, RateLimitError):
                    last_exc = root_exc
                    if attempt == attempts:
                        break
                    delay = base_delay * (2 ** (attempt - 1))
                    logger.warning(
                        "Rate limit retry exhausted inside client, retrying wrapper in %.1f sec (attempt %s/%s)",
                        delay,
                        attempt,
                        attempts,
                    )
                    await asyncio.sleep(delay)
                elif isinstance(root_exc, OzonAPIError) and root_exc.status_code == 429:
                    last_exc = root_exc
                    if attempt == attempts:
                        break
                    delay = base_delay * (2 ** (attempt - 1))
                    logger.warning(
                        "HTTP 429 retry exhausted inside client, retrying wrapper in %.1f sec (attempt %s/%s)",
                        delay,
                        attempt,
                        attempts,
                    )
                    await asyncio.sleep(delay)
                else:
                    raise
        raise last_exc if last_exc else RuntimeError("Unknown retry error")

    async def _ensure_analytics_data_schema(self) -> None:
        """Mininal'naja migracija tablicy analytics_data dlja PostgreSQL."""
        if "postgresql" not in settings.database_url.lower():
            return

        create_sql = """
            CREATE TABLE IF NOT EXISTS analytics_data (
                id BIGSERIAL PRIMARY KEY,
                date TIMESTAMPTZ NOT NULL,
                sku BIGINT NOT NULL,
                ordered_units INTEGER,
                delivered_units INTEGER,
                returned_units INTEGER,
                revenue NUMERIC(15, 2),
                position DOUBLE PRECISION,
                raw_data JSONB
            );
        """
        alter_sql = [
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS impressions INTEGER;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS clicks INTEGER;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS ctr NUMERIC(8, 4);",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS position_category DOUBLE PRECISION;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS position_promo DOUBLE PRECISION;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS metric_values JSONB;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS dimensions JSONB;",
            "ALTER TABLE analytics_data ADD COLUMN IF NOT EXISTS last_synced_at TIMESTAMPTZ DEFAULT now();",
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_analytics_data_date_sku_idx ON analytics_data(date, sku);",
            "CREATE INDEX IF NOT EXISTS idx_analytics_date_sku ON analytics_data(date, sku);",
        ]

        async with db_manager.session() as session:
            await session.execute(text(create_sql))
            for stmt in alter_sql:
                await session.execute(text(stmt))

    async def _ensure_fact_orders_cluster_schema(self) -> None:
        """Dobavljaet kolонки klasterov dostavki v fact_orders, esli ih net."""
        if "postgresql" not in settings.database_url.lower():
            return
        alter_sql = [
            "ALTER TABLE fact_orders ADD COLUMN IF NOT EXISTS delivery_cluster_from VARCHAR(255);",
            "ALTER TABLE fact_orders ADD COLUMN IF NOT EXISTS delivery_cluster_to VARCHAR(255);",
            "ALTER TABLE fact_orders ADD COLUMN IF NOT EXISTS shipping_warehouse_name VARCHAR(255);",
            "CREATE INDEX IF NOT EXISTS idx_fact_order_cluster_to ON fact_orders(delivery_cluster_to);",
            "CREATE INDEX IF NOT EXISTS idx_fact_order_cluster_from ON fact_orders(delivery_cluster_from);",
        ]
        async with db_manager.session() as session:
            for stmt in alter_sql:
                await session.execute(text(stmt))

    async def sync_order_delivery_clusters(self, days_back: int = 35) -> Dict[str, int]:
        """
        Obogashhaet fact_orders klasterami otgruzki/dostavki iz posting API.
        Istochnik: financial_data.cluster_from / financial_data.cluster_to.
        """
        logger.info("Starting order delivery clusters sync...")
        sync_log = await self._create_sync_log("order_delivery_clusters")
        await self._ensure_fact_orders_cluster_schema()

        now_utc = datetime.now(timezone.utc)
        since_utc = now_utc - timedelta(days=max(1, int(days_back)))

        updates_by_posting: Dict[str, Dict[str, Any]] = {}

        async def collect_fbo() -> int:
            processed = 0
            offset = 0
            while True:
                resp = await self._with_rate_limit_retry(
                    lambda o=offset: self.client._make_request(
                        "POST",
                        "/v2/posting/fbo/list",
                        {
                            "dir": "asc",
                            "filter": {
                                "since": since_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                                "to": now_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                            },
                            "limit": 1000,
                            "offset": o,
                            "with": {"analytics_data": True, "financial_data": True},
                        },
                    )
                )
                batch = resp.get("result", []) if isinstance(resp, dict) else []
                if not batch:
                    break
                for posting in batch:
                    if not isinstance(posting, dict):
                        continue
                    posting_number = str(posting.get("posting_number") or "").strip()
                    if not posting_number:
                        continue
                    fd = posting.get("financial_data") or {}
                    ad = posting.get("analytics_data") or {}
                    updates_by_posting[posting_number] = {
                        "posting_number": posting_number,
                        "order_id": str(posting.get("order_id") or posting_number),
                        "delivery_cluster_from": (fd.get("cluster_from") or None),
                        "delivery_cluster_to": (fd.get("cluster_to") or None),
                        "shipping_warehouse_name": ad.get("warehouse_name") or ad.get("warehouse") or None,
                    }
                    processed += 1
                if len(batch) < 1000:
                    break
                offset += 1000
            return processed

        async def collect_fbs() -> int:
            processed = 0
            offset = 0
            while True:
                resp = await self._with_rate_limit_retry(
                    lambda o=offset: self.client._make_request(
                        "POST",
                        "/v3/posting/fbs/list",
                        {
                            "dir": "asc",
                            "filter": {
                                "since": since_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                                "to": now_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                            },
                            "limit": 1000,
                            "offset": o,
                            "with": {
                                "analytics_data": True,
                                "financial_data": True,
                                "barcodes": False,
                                "translit": False,
                            },
                        },
                    )
                )
                result = resp.get("result", {}) if isinstance(resp, dict) else {}
                batch = result.get("postings", []) if isinstance(result, dict) else []
                if not batch:
                    break
                for posting in batch:
                    if not isinstance(posting, dict):
                        continue
                    posting_number = str(posting.get("posting_number") or "").strip()
                    if not posting_number:
                        continue
                    fd = posting.get("financial_data") or {}
                    ad = posting.get("analytics_data") or {}
                    delivery = posting.get("delivery") or {}
                    cluster_to = fd.get("cluster_to") or delivery.get("cluster")
                    updates_by_posting[posting_number] = {
                        "posting_number": posting_number,
                        "order_id": str(posting.get("order_id") or posting_number),
                        "delivery_cluster_from": (fd.get("cluster_from") or None),
                        "delivery_cluster_to": (cluster_to or None),
                        "shipping_warehouse_name": ad.get("warehouse_name") or ad.get("warehouse") or None,
                    }
                    processed += 1
                if len(batch) < 1000:
                    break
                offset += 1000
            return processed

        try:
            fbo_processed = await collect_fbo()
            fbs_processed = await collect_fbs()
            updates_applied = 0
            async with db_manager.session() as session:
                update_sql = text(
                    """
                    UPDATE fact_orders
                    SET
                        delivery_cluster_from = COALESCE(:delivery_cluster_from, delivery_cluster_from),
                        delivery_cluster_to = COALESCE(:delivery_cluster_to, delivery_cluster_to),
                        shipping_warehouse_name = COALESCE(:shipping_warehouse_name, shipping_warehouse_name),
                        last_synced_at = now()
                    WHERE posting_number = :posting_number
                       OR order_id = :order_id
                    """
                )
                for payload in updates_by_posting.values():
                    result = await session.execute(update_sql, payload)
                    updates_applied += int(result.rowcount or 0)

            total_processed = fbo_processed + fbs_processed
            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=updates_applied,
                records_inserted=0,
                records_updated=updates_applied,
            )
            logger.info(
                "order delivery clusters sync completed: source_rows=%s updates_applied=%s",
                total_processed,
                updates_applied,
            )
            return {
                "source_rows": total_processed,
                "updates_applied": updates_applied,
                "fbo_source_rows": fbo_processed,
                "fbs_source_rows": fbs_processed,
            }
        except Exception as e:
            logger.error(f"order delivery clusters sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def _replace_transaction_children(
        self,
        session,
        transaction_id: int,
        posting_number: Optional[str],
        operation_date: Optional[datetime],
        raw_data: Optional[Dict[str, Any]],
    ) -> None:
        payload = raw_data if isinstance(raw_data, dict) else {}
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        services = payload.get("services") if isinstance(payload.get("services"), list) else []

        await session.execute(delete(TransactionItem).where(TransactionItem.transaction_id == transaction_id))
        await session.execute(delete(TransactionService).where(TransactionService.transaction_id == transaction_id))

        for line_no, item in enumerate(items, start=1):
            row_data = {
                "transaction_id": transaction_id,
                "line_no": line_no,
                "posting_number": posting_number,
                "operation_date": operation_date,
                "sku": self._parse_int_flexible(item.get("sku")) if isinstance(item, dict) else None,
                "name": item.get("name") if isinstance(item, dict) else None,
                "quantity": self._parse_int_flexible(item.get("quantity")) if isinstance(item, dict) else 1,
                "raw_data": item,
                "last_synced_at": datetime.now(),
            }
            await session.execute(insert(TransactionItem).values(**row_data))

        for line_no, service in enumerate(services, start=1):
            row_data = {
                "transaction_id": transaction_id,
                "line_no": line_no,
                "posting_number": posting_number,
                "operation_date": operation_date,
                "service_name": service.get("name") if isinstance(service, dict) else None,
                "price": self._parse_decimal_flexible(service.get("price")) if isinstance(service, dict) else None,
                "raw_data": service,
                "last_synced_at": datetime.now(),
            }
            await session.execute(insert(TransactionService).values(**row_data))

    async def _replace_fact_order_items(
        self,
        session,
        order_id: str,
        posting_number: Optional[str],
        items: Optional[List[Dict[str, Any]]],
    ) -> None:
        if posting_number:
            await session.execute(
                delete(FactOrderItem).where(FactOrderItem.posting_number == posting_number)
            )
            await session.execute(
                delete(FactOrder).where(
                    FactOrder.posting_number == posting_number,
                    FactOrder.order_id != order_id,
                )
            )
        await session.execute(delete(FactOrderItem).where(FactOrderItem.order_id == order_id))
        for line_no, item in enumerate(items or [], start=1):
            row_data = {
                "order_id": order_id,
                "posting_number": posting_number,
                "line_no": line_no,
                "offer_id": item.get("offer_id") if isinstance(item, dict) else None,
                "sku": self._parse_int_flexible(item.get("sku")) if isinstance(item, dict) else None,
                "product_name": item.get("name") if isinstance(item, dict) else None,
                "quantity": self._parse_decimal_flexible(item.get("quantity")) if isinstance(item, dict) else None,
                "price": self._parse_decimal_flexible(item.get("price")) if isinstance(item, dict) else None,
                "buyer_paid": self._parse_decimal_flexible(item.get("buyer_paid")) if isinstance(item, dict) else None,
                "raw_data": item,
                "last_synced_at": datetime.now(),
            }
            await session.execute(insert(FactOrderItem).values(**row_data))

    async def _upsert_realization_detail(
        self,
        session,
        realization_report_id: int,
        raw_data: Optional[Dict[str, Any]],
    ) -> None:
        payload = raw_data if isinstance(raw_data, dict) else {}
        header = payload.get("header", {}) if isinstance(payload.get("header"), dict) else {}
        row = payload.get("row", {}) if isinstance(payload.get("row"), dict) else {}
        item = row.get("item", {}) if isinstance(row.get("item"), dict) else {}
        delivery = row.get("delivery_commission", {}) if isinstance(row.get("delivery_commission"), dict) else {}
        return_commission = payload.get("return_commission", {}) if isinstance(payload.get("return_commission"), dict) else {}

        row_data = {
            "realization_report_id": realization_report_id,
            "document_number": header.get("number"),
            "document_date": self._parse_datetime_flexible(header.get("doc_date")),
            "period_start": self._parse_datetime_flexible(header.get("start_date")),
            "period_end": self._parse_datetime_flexible(header.get("stop_date")),
            "contract_number": header.get("contract_number"),
            "row_number": self._parse_int_flexible(row.get("rowNumber")),
            "barcode": item.get("barcode"),
            "delivery_quantity": self._parse_int_flexible(delivery.get("quantity")),
            "delivery_amount": self._parse_decimal_flexible(delivery.get("amount")),
            "delivery_bonus": self._parse_decimal_flexible(delivery.get("bonus")),
            "delivery_standard_fee": self._parse_decimal_flexible(delivery.get("standard_fee")),
            "delivery_total": self._parse_decimal_flexible(delivery.get("total")),
            "bank_coinvestment": self._parse_decimal_flexible(delivery.get("bank_coinvestment")),
            "pick_up_point_coinvestment": self._parse_decimal_flexible(delivery.get("pick_up_point_coinvestment")),
            "stars": self._parse_decimal_flexible(delivery.get("stars")),
            "return_commission_amount": self._parse_decimal_flexible(return_commission.get("amount")),
            "raw_data": payload,
            "last_synced_at": datetime.now(),
        }
        stmt = pg_insert(RealizationReportDetail).values(**row_data)
        stmt = stmt.on_conflict_do_update(
            index_elements=["realization_report_id"],
            set_={k: v for k, v in row_data.items() if k != "realization_report_id"},
        )
        await session.execute(stmt)

    async def sync_transactions(self, days_back: Optional[int] = None) -> Dict[str, int]:
        """Sinhronizacija /v3/finance/transaction/list v tablicu transactions."""
        if days_back is None:
            days_back = max(settings.sync_days_back, 30)

        logger.info(f"Starting transactions sync for last {days_back} days...")
        sync_log = await self._create_sync_log("transactions")

        records_processed = 0
        records_inserted = 0

        to_date = datetime.now(timezone.utc)
        from_date = to_date - timedelta(days=days_back)

        try:
            async with db_manager.session() as session:
                for window_from, window_to in self._iter_month_windows(from_date, to_date):
                    logger.info(
                        "transactions window: %s -> %s",
                        window_from.strftime("%Y-%m-%d"),
                        window_to.strftime("%Y-%m-%d"),
                    )
                    async for transactions in self.client.get_all_transactions(
                        from_date=window_from,
                        to_date=window_to,
                        transaction_type="all",
                        page_size=1000,
                    ):
                        for transaction_data in transactions:
                            operation_id = transaction_data.get("operation_id")
                            if not operation_id:
                                continue

                            posting = transaction_data.get("posting", {}) if isinstance(transaction_data.get("posting"), dict) else {}
                            transaction_dict = {
                                "transaction_id": int(operation_id),
                                "operation_id": int(operation_id),
                                "operation_type": transaction_data.get("operation_type"),
                                "operation_date": self._parse_operation_datetime(transaction_data.get("operation_date")),
                                "posting_number": posting.get("posting_number"),
                                "amount": self._parse_decimal_flexible(transaction_data.get("amount")),
                                "currency": transaction_data.get("currency_code") or transaction_data.get("currency"),
                                "type": transaction_data.get("type"),
                                "description": transaction_data.get("operation_type_name") or transaction_data.get("description"),
                                "raw_data": transaction_data,
                                "last_synced_at": datetime.now(),
                            }

                            stmt = pg_insert(Transaction).values(**transaction_dict)
                            stmt = stmt.on_conflict_do_update(
                                index_elements=["transaction_id"],
                                set_={k: v for k, v in transaction_dict.items() if k != "transaction_id"},
                            )
                            await session.execute(stmt)
                            await self._replace_transaction_children(
                                session=session,
                                transaction_id=int(operation_id),
                                posting_number=posting.get("posting_number"),
                                operation_date=transaction_dict["operation_date"],
                                raw_data=transaction_data,
                            )
                            records_processed += 1
                            records_inserted += 1

            records_updated = max(records_processed - records_inserted, 0)
            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=records_processed,
                records_inserted=records_inserted,
                records_updated=records_updated,
            )
            logger.info(
                "Transactions sync completed: processed=%s inserted=%s updated=%s",
                records_processed,
                records_inserted,
                records_updated,
            )
            return {
                "processed": records_processed,
                "inserted": records_inserted,
                "updated": records_updated,
            }
        except Exception as e:
            logger.error(f"Transactions sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def backfill_normalized_finance_data(self) -> Dict[str, int]:
        """Razlozhit' uzhe zagruzhennye JSON-polja po strukturirovannym tablicam."""
        logger.info("Starting normalized finance data backfill...")
        sync_log = await self._create_sync_log("normalized_finance_backfill")

        transactions_processed = 0
        fact_orders_processed = 0
        realization_processed = 0

        try:
            async with db_manager.session() as session:
                transactions = (await session.execute(select(Transaction))).scalars().all()
                for transaction in transactions:
                    await self._replace_transaction_children(
                        session=session,
                        transaction_id=transaction.transaction_id,
                        posting_number=transaction.posting_number,
                        operation_date=transaction.operation_date,
                        raw_data=transaction.raw_data,
                    )
                    transactions_processed += 1

                fact_orders = (await session.execute(select(FactOrder))).scalars().all()
                for fact_order in fact_orders:
                    await self._replace_fact_order_items(
                        session=session,
                        order_id=fact_order.order_id,
                        posting_number=fact_order.posting_number,
                        items=fact_order.items if isinstance(fact_order.items, list) else [],
                    )
                    fact_orders_processed += 1

                realization_rows = (await session.execute(select(RealizationReport))).scalars().all()
                for realization_row in realization_rows:
                    await self._upsert_realization_detail(
                        session=session,
                        realization_report_id=realization_row.id,
                        raw_data=realization_row.raw_data,
                    )
                    realization_processed += 1

            total = transactions_processed + fact_orders_processed + realization_processed
            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=total,
                records_inserted=total,
                records_updated=0,
            )
            logger.info(
                "normalized finance backfill completed: transactions=%s fact_orders=%s realization=%s",
                transactions_processed,
                fact_orders_processed,
                realization_processed,
            )
            return {
                "transactions_processed": transactions_processed,
                "fact_orders_processed": fact_orders_processed,
                "realization_processed": realization_processed,
            }
        except Exception as e:
            logger.error(f"normalized finance backfill failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== PRODUCTS SYNC ====================
    
    async def sync_products(self) -> Dict[str, int]:
        """Sinhronizacija tovarov."""
        logger.info("Starting products sync...")
        sync_log = await self._create_sync_log("products")
        
        records_processed = 0
        records_inserted = 0
        records_updated = 0
        
        try:
            all_items: List[Dict[str, Any]] = []
            async for items in self.client.get_all_products():
                if items:
                    all_items.extend(items)

            if not all_items:
                await self._update_sync_log(sync_log, "success", 0, 0, 0)
                return {"processed": 0, "inserted": 0, "updated": 0}

            def _collect_sku_candidates(item: Dict[str, Any]) -> List[int]:
                result: List[int] = []
                sku_direct = self._parse_int_flexible(item.get("sku"))
                if sku_direct:
                    result.append(sku_direct)
                sources = item.get("sources")
                if isinstance(sources, list):
                    for src in sources:
                        if not isinstance(src, dict):
                            continue
                        sku_val = self._parse_int_flexible(src.get("sku"))
                        if sku_val:
                            result.append(sku_val)
                return result

            sku_set = set()
            for item in all_items:
                for sku in _collect_sku_candidates(item):
                    if sku > 0:
                        sku_set.add(int(sku))

            offer_ids = [str(it.get("offer_id") or "").strip() for it in all_items if str(it.get("offer_id") or "").strip()]
            offer_to_sku: Dict[str, int] = {}
            if offer_ids:
                try:
                    async with db_manager.session() as session:
                        lookup_rows = await session.execute(
                            text(
                                """
                                WITH src AS (
                                    SELECT trim(offer_id) AS offer_id, fbo_sku_id::bigint AS sku, last_synced_at AS synced_at
                                    FROM report_products_items
                                    WHERE fbo_sku_id IS NOT NULL AND coalesce(trim(offer_id), '') <> ''
                                    UNION ALL
                                    SELECT trim(offer_id) AS offer_id, fbs_sku_id::bigint AS sku, last_synced_at AS synced_at
                                    FROM report_products_items
                                    WHERE fbs_sku_id IS NOT NULL AND coalesce(trim(offer_id), '') <> ''
                                    UNION ALL
                                    SELECT trim(offer_id) AS offer_id, sku::bigint AS sku, updated_at AS synced_at
                                    FROM article_characteristics
                                    WHERE sku IS NOT NULL AND coalesce(trim(offer_id), '') <> ''
                                )
                                SELECT DISTINCT ON (lower(offer_id))
                                    lower(offer_id) AS offer_key,
                                    sku
                                FROM src
                                WHERE lower(offer_id) = ANY(:offer_keys)
                                ORDER BY lower(offer_id), synced_at DESC NULLS LAST
                                """
                            ),
                            {"offer_keys": [o.lower() for o in offer_ids]},
                        )
                        for row in lookup_rows.fetchall():
                            key = str(row[0] or "").strip().lower()
                            sku_val = self._parse_int_flexible(row[1])
                            if key and sku_val:
                                offer_to_sku[key] = int(sku_val)
                                sku_set.add(int(sku_val))
                except Exception as e:
                    logger.warning(f"Failed to build offer->sku fallback map: {e}")
            all_skus = sorted(sku_set)

            v3_by_sku: Dict[int, Dict[str, Any]] = {}
            v4_by_sku: Dict[int, Dict[str, Any]] = {}
            chunk_size = 100
            for i in range(0, len(all_skus), chunk_size):
                chunk = all_skus[i:i + chunk_size]
                if not chunk:
                    continue
                try:
                    v3_resp = await self.client.get_product_info_list_v3(chunk)
                    for info_item in (v3_resp.get("items") or []):
                        sku_val = self._parse_int_flexible(info_item.get("sku"))
                        if sku_val:
                            v3_by_sku[int(sku_val)] = info_item
                except Exception as e:
                    logger.warning(f"Failed to get /v3/product/info/list chunk {i // chunk_size + 1}: {e}")

                try:
                    v4_resp = await self.client.get_product_info_attributes(chunk)
                    v4_items = (v4_resp.get("result") or [])
                    for attr_item in v4_items:
                        sku_val = self._parse_int_flexible(attr_item.get("sku"))
                        if sku_val:
                            v4_by_sku[int(sku_val)] = attr_item
                except Exception as e:
                    logger.warning(f"Failed to get /v4/product/info/attributes chunk {i // chunk_size + 1}: {e}")

            for item in all_items:
                product_id = item.get("product_id")
                offer_id = item.get("offer_id")
                if not product_id:
                    continue

                sku_candidates = _collect_sku_candidates(item)
                if not sku_candidates:
                    key = str(item.get("offer_id") or "").strip().lower()
                    fallback_sku = offer_to_sku.get(key)
                    if fallback_sku:
                        sku_candidates = [fallback_sku]
                sku_main = int(sku_candidates[0]) if sku_candidates else None
                v3_info = v3_by_sku.get(sku_main or 0, {})
                v4_attrs = v4_by_sku.get(sku_main or 0, {})

                # /v2/product/info often returns 404 on current account setup,
                # so we use v3/v4 as primary canonical source.
                product_info_v2 = {}

                resolved_offer_id = (
                    item.get("offer_id")
                    or v3_info.get("offer_id")
                    or v4_attrs.get("offer_id")
                    or offer_id
                )
                resolved_name = (
                    product_info_v2.get("name")
                    or v3_info.get("name")
                    or v4_attrs.get("name")
                    or item.get("name")
                )
                resolved_barcode = (
                    product_info_v2.get("barcode")
                    or ((v3_info.get("barcodes") or [None])[0] if isinstance(v3_info.get("barcodes"), list) else None)
                    or ((v4_attrs.get("barcodes") or [None])[0] if isinstance(v4_attrs.get("barcodes"), list) else None)
                )

                stock_fbo = 0
                stock_fbs = 0
                stocks_raw = item.get("stocks")
                if isinstance(stocks_raw, dict):
                    stock_fbo = int(stocks_raw.get("fbo") or 0)
                    stock_fbs = int(stocks_raw.get("fbs") or 0)
                elif isinstance(stocks_raw, list):
                    for s in stocks_raw:
                        if not isinstance(s, dict):
                            continue
                        stock_type = str(s.get("type") or "").lower()
                        if stock_type == "fbo":
                            stock_fbo = int(s.get("present") or s.get("stock") or 0)
                        elif stock_type == "fbs":
                            stock_fbs = int(s.get("present") or s.get("stock") or 0)

                category_id = (
                    product_info_v2.get("category_id")
                    or v3_info.get("category_id")
                    or v4_attrs.get("category_id")
                )
                type_id = (
                    product_info_v2.get("type_id")
                    or v3_info.get("type_id")
                    or v4_attrs.get("type_id")
                )

                raw_payload = {
                    "product_list_item": item,
                    "product_info_v2": product_info_v2,
                    "product_info_v3": v3_info,
                    "product_attributes_v4": v4_attrs,
                    "resolved_ids": {
                        "product_id": product_id,
                        "offer_id": resolved_offer_id,
                        "sku": sku_main,
                    },
                }

                product_data = {
                    "product_id": product_id,
                    "offer_id": resolved_offer_id,
                    "name": resolved_name,
                    "barcode": resolved_barcode,
                    "category_id": self._parse_int_flexible(category_id),
                    "type_id": self._parse_int_flexible(type_id),
                    "created_at": self._parse_datetime(
                        product_info_v2.get("created_at")
                        or v3_info.get("created_at")
                        or v4_attrs.get("created_at")
                    ),
                    "updated_at": self._parse_datetime(
                        product_info_v2.get("updated_at")
                        or v3_info.get("updated_at")
                        or v4_attrs.get("updated_at")
                    ),
                    "price": self._parse_decimal(product_info_v2.get("price")),
                    "old_price": self._parse_decimal(product_info_v2.get("old_price")),
                    "retail_price": self._parse_decimal(product_info_v2.get("retail_price")),
                    "min_ozon_price": self._parse_decimal(product_info_v2.get("min_ozon_price")),
                    "is_visible": bool(product_info_v2.get("visible", True)),
                    "status": product_info_v2.get("status") or v3_info.get("status"),
                    "stock_fbo": stock_fbo,
                    "stock_fbs": stock_fbs,
                    "raw_data": raw_payload,
                    "last_synced_at": datetime.now(),
                }

                async with db_manager.session() as session:
                    exists_row = await session.execute(
                        select(Product.id).where(Product.product_id == product_id)
                    )
                    existed_before = exists_row.scalar_one_or_none() is not None

                    stmt = pg_insert(Product).values(**product_data)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["product_id"],
                        set_={k: v for k, v in product_data.items() if k != "product_id"},
                    )
                    await session.execute(stmt)

                    records_processed += 1
                    if existed_before:
                        records_updated += 1
                    else:
                        records_inserted += 1
            
            await self._update_sync_log(
                sync_log, "success",
                records_processed, records_inserted, records_updated
            )
            
            logger.info(f"Products sync completed: {records_processed} processed")
            return {
                "processed": records_processed,
                "inserted": records_inserted,
                "updated": records_updated
            }
            
        except Exception as e:
            logger.error(f"Products sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== PRICES SYNC ====================
    
    async def sync_prices(self) -> Dict[str, int]:
        """Sinhronizacija cen."""
        logger.info("Starting prices sync...")
        sync_log = await self._create_sync_log("prices")
        
        records_processed = 0
        
        try:
            async for items in self.client.get_all_product_prices():
                for item in items:
                    offer_id = item.get("offer_id")
                    
                    async with db_manager.session() as session:
                        result = await session.execute(
                            select(Product).where(Product.offer_id == offer_id)
                        )
                        product = result.scalar_one_or_none()
                        
                        if product:
                            product.price = self._parse_decimal(item.get("price"))
                            product.old_price = self._parse_decimal(item.get("old_price"))
                            product.retail_price = self._parse_decimal(item.get("retail_price"))
                            product.min_ozon_price = self._parse_decimal(item.get("min_ozon_price"))
                            product.last_synced_at = datetime.now()
                            records_processed += 1
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Prices sync completed: {records_processed} updated")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"Prices sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== STOCKS SYNC ====================
    
    async def sync_stocks(self) -> Dict[str, int]:
        """Sinhronizacija ostatkov."""
        logger.info("Starting stocks sync...")
        sync_log = await self._create_sync_log("stocks")
        
        records_processed = 0
        
        try:
            async for items in self.client.get_all_product_stocks():
                for item in items:
                    offer_id = item.get("offer_id")
                    stocks = item.get("stocks", [])
                    
                    stock_fbo = 0
                    stock_fbs = 0
                    
                    for stock in stocks:
                        if stock.get("type") == "fbo":
                            stock_fbo = stock.get("present", 0)
                        elif stock.get("type") == "fbs":
                            stock_fbs = stock.get("present", 0)
                    
                    async with db_manager.session() as session:
                        result = await session.execute(
                            select(Product).where(Product.offer_id == offer_id)
                        )
                        product = result.scalar_one_or_none()
                        
                        if product:
                            product.stock_fbo = stock_fbo
                            product.stock_fbs = stock_fbs
                            product.last_synced_at = datetime.now()
                            records_processed += 1
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Stocks sync completed: {records_processed} updated")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"Stocks sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_analytics_stocks(self) -> Dict[str, int]:
        """Sinhronizacija /v1/analytics/stocks."""
        logger.info("Starting analytics_stocks sync...")
        sync_log = await self._create_sync_log("analytics_stocks")

        try:
            async with db_manager.session() as session:
                sku_rows = await session.execute(
                    select(ReportProductItem.fbo_sku_id)
                    .where(ReportProductItem.fbo_sku_id.isnot(None))
                    .distinct()
                )
                skus = [int(s) for s in sku_rows.scalars().all() if s is not None and int(s) > 0]

                if not skus:
                    sku_rows = await session.execute(
                        select(ReportProductItem.fbs_sku_id)
                        .where(ReportProductItem.fbs_sku_id.isnot(None))
                        .distinct()
                    )
                    skus = [int(s) for s in sku_rows.scalars().all() if s is not None and int(s) > 0]

            if not skus:
                raise ValueError("No SKUs found in DB for /v1/analytics/stocks request")

            skus = sorted(set(skus))
            logger.info(f"analytics_stocks: requesting {len(skus)} SKU(s)")

            rows: List[Dict[str, Any]] = []
            skipped_skus: List[int] = []
            chunk_size = 100
            for i in range(0, len(skus), chunk_size):
                chunk = skus[i:i + chunk_size]
                chunk_rows, chunk_skipped = await self._fetch_analytics_stocks_chunk_resilient(chunk)
                rows.extend(chunk_rows)
                skipped_skus.extend(chunk_skipped)

            rows_upserted = 0
            async with db_manager.session() as session:
                await session.execute(delete(AnalyticsStock))

                for row in rows:
                    data = {
                        "sku": self._parse_int_flexible(row.get("sku")),
                        "offer_id": row.get("offer_id"),
                        "name": row.get("name"),
                        "warehouse_id": self._parse_int_flexible(row.get("warehouse_id")) or 0,
                        "warehouse_name": row.get("warehouse_name"),
                        "cluster_id": self._parse_int_flexible(row.get("cluster_id")),
                        "cluster_name": row.get("cluster_name"),
                        "macrolocal_cluster_id": self._parse_int_flexible(row.get("macrolocal_cluster_id")),
                        "ads": self._parse_decimal_flexible(row.get("ads")),
                        "idc": self._parse_decimal_flexible(row.get("idc")),
                        "days_without_sales": self._parse_int_flexible(row.get("days_without_sales")),
                        "turnover_grade": row.get("turnover_grade"),
                        "available_stock_count": self._parse_int_flexible(row.get("available_stock_count")),
                        "valid_stock_count": self._parse_int_flexible(row.get("valid_stock_count")),
                        "waiting_docs_stock_count": self._parse_int_flexible(row.get("waiting_docs_stock_count")),
                        "expiring_stock_count": self._parse_int_flexible(row.get("expiring_stock_count")),
                        "transit_defect_stock_count": self._parse_int_flexible(row.get("transit_defect_stock_count")),
                        "stock_defect_stock_count": self._parse_int_flexible(row.get("stock_defect_stock_count")),
                        "excess_stock_count": self._parse_int_flexible(row.get("excess_stock_count")),
                        "other_stock_count": self._parse_int_flexible(row.get("other_stock_count")),
                        "requested_stock_count": self._parse_int_flexible(row.get("requested_stock_count")),
                        "transit_stock_count": self._parse_int_flexible(row.get("transit_stock_count")),
                        "return_from_customer_stock_count": self._parse_int_flexible(row.get("return_from_customer_stock_count")),
                        "return_to_seller_stock_count": self._parse_int_flexible(row.get("return_to_seller_stock_count")),
                        "ads_cluster": self._parse_decimal_flexible(row.get("ads_cluster")),
                        "idc_cluster": self._parse_decimal_flexible(row.get("idc_cluster")),
                        "days_without_sales_cluster": self._parse_int_flexible(row.get("days_without_sales_cluster")),
                        "turnover_grade_cluster": row.get("turnover_grade_cluster"),
                        "item_tags": row.get("item_tags"),
                        "raw_data": row,
                        "last_synced_at": datetime.now(),
                    }
                    if not data["sku"]:
                        continue
                    stmt = pg_insert(AnalyticsStock).values(**data)
                    stmt = stmt.on_conflict_do_update(
                        constraint="uq_analytics_stocks_sku_wh_cluster",
                        set_={k: v for k, v in data.items() if k not in {"sku", "warehouse_id", "cluster_id"}},
                    )
                    await session.execute(stmt)
                    rows_upserted += 1

            # Снапшот в историю после полной загрузки актуальных остатков
            try:
                await self._capture_stock_daily_snapshot(session)
            except Exception as snap_err:
                logger.error(f"stock_daily_snapshot failed (non-fatal): {snap_err}")

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
                error_message=(
                    f"Skipped {len(skipped_skus)} SKU(s): {', '.join(map(str, skipped_skus[:20]))}"
                    if skipped_skus else None
                ),
            )
            if skipped_skus:
                logger.warning(
                    "analytics_stocks sync completed with skipped SKU(s): %s",
                    ", ".join(map(str, skipped_skus[:20])) + ("..." if len(skipped_skus) > 20 else ""),
                )
            logger.info(f"analytics_stocks sync completed: rows_upserted={rows_upserted}")
            return {
                "rows_upserted": rows_upserted,
                "skus_requested": len(skus),
                "skus_skipped": len(skipped_skus),
            }
        except Exception as e:
            logger.error(f"analytics_stocks sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def _capture_stock_daily_snapshot(self, session) -> int:
        """Сохраняет в stock_daily_snapshots снимок остатков (FBO + FBS) на текущую дату.
        Идемпотентно: повторный вызов в тот же день перезаписывает строки за этот день."""
        # FBO: из analytics_stocks (детализация по складу/кластеру)
        fbo_sql = _sql_text("""
            INSERT INTO stock_daily_snapshots (
                snapshot_date, sku, offer_id, stock_type,
                warehouse_id, warehouse_name, cluster_id, cluster_name,
                stock_total, stock_available, stock_supply, stock_transit, stock_acceptance, stock_reserved,
                source_table
            )
            SELECT
                CURRENT_DATE,
                sku,
                offer_id,
                'FBO',
                warehouse_id,
                warehouse_name,
                COALESCE(cluster_id, 0),
                cluster_name,
                COALESCE(available_stock_count, 0)
                  + COALESCE(requested_stock_count, 0)
                  + COALESCE(transit_stock_count, 0)
                  + COALESCE(waiting_docs_stock_count, 0),
                COALESCE(available_stock_count, 0),
                COALESCE(requested_stock_count, 0),
                COALESCE(transit_stock_count, 0),
                COALESCE(waiting_docs_stock_count, 0),
                0,
                'analytics_stocks'
            FROM analytics_stocks
            WHERE sku IS NOT NULL AND warehouse_id IS NOT NULL
            ON CONFLICT ON CONSTRAINT uq_stock_daily_snapshots_key DO UPDATE SET
                offer_id = EXCLUDED.offer_id,
                warehouse_name = EXCLUDED.warehouse_name,
                cluster_name = EXCLUDED.cluster_name,
                stock_total = EXCLUDED.stock_total,
                stock_available = EXCLUDED.stock_available,
                stock_supply = EXCLUDED.stock_supply,
                stock_transit = EXCLUDED.stock_transit,
                stock_acceptance = EXCLUDED.stock_acceptance,
                snapshot_at = now()
        """)
        # FBS: из fbs_warehouse_stocks
        fbs_sql = _sql_text("""
            INSERT INTO stock_daily_snapshots (
                snapshot_date, sku, offer_id, stock_type,
                warehouse_id, warehouse_name, cluster_id, cluster_name,
                stock_total, stock_available, stock_supply, stock_transit, stock_acceptance, stock_reserved,
                source_table
            )
            SELECT
                CURRENT_DATE,
                sku,
                offer_id,
                'FBS',
                warehouse_id,
                warehouse_name,
                0,
                NULL,
                COALESCE(present, 0),
                COALESCE(present, 0),
                0, 0, 0,
                COALESCE(reserved, 0),
                'fbs_warehouse_stocks'
            FROM fbs_warehouse_stocks
            WHERE sku IS NOT NULL AND warehouse_id IS NOT NULL
            ON CONFLICT ON CONSTRAINT uq_stock_daily_snapshots_key DO UPDATE SET
                offer_id = EXCLUDED.offer_id,
                warehouse_name = EXCLUDED.warehouse_name,
                stock_total = EXCLUDED.stock_total,
                stock_available = EXCLUDED.stock_available,
                stock_reserved = EXCLUDED.stock_reserved,
                snapshot_at = now()
        """)
        fbo_res = await session.execute(fbo_sql)
        fbs_res = await session.execute(fbs_sql)
        fbo_n = fbo_res.rowcount or 0
        fbs_n = fbs_res.rowcount or 0
        logger.info(f"stock_daily_snapshot captured: FBO={fbo_n} FBS={fbs_n}")
        return fbo_n + fbs_n

    async def _capture_delivery_time_daily_snapshot(self, session) -> int:
        """Снимок AoT кластеров на текущую дату из analytics_average_delivery_time."""
        sql = _sql_text("""
            INSERT INTO delivery_time_daily_snapshots (
                snapshot_date, delivery_cluster_id, average_delivery_time, average_delivery_time_status,
                orders_total, orders_fast, orders_medium, orders_long
            )
            SELECT
                CURRENT_DATE,
                delivery_cluster_id,
                average_delivery_time,
                average_delivery_time_status,
                orders_total,
                orders_fast,
                orders_medium,
                orders_long
            FROM analytics_average_delivery_time
            WHERE delivery_cluster_id IS NOT NULL
            ON CONFLICT ON CONSTRAINT uq_delivery_time_daily_snapshots_key DO UPDATE SET
                average_delivery_time = EXCLUDED.average_delivery_time,
                average_delivery_time_status = EXCLUDED.average_delivery_time_status,
                orders_total = EXCLUDED.orders_total,
                orders_fast = EXCLUDED.orders_fast,
                orders_medium = EXCLUDED.orders_medium,
                orders_long = EXCLUDED.orders_long,
                snapshot_at = now()
        """)
        res = await session.execute(sql)
        n = res.rowcount or 0
        logger.info(f"delivery_time_daily_snapshot captured: {n} clusters")
        return n

    async def sync_analytics_data(
        self,
        days_back: int = 7,
        metrics: Optional[List[str]] = None,
        dimensions: Optional[List[str]] = None,
    ) -> Dict[str, int]:
        """Sinhronizacija /v1/analytics/data (po SKU i dnju)."""
        logger.info("Starting analytics_data sync...")
        sync_log = await self._create_sync_log("analytics_data")

        from_date = datetime.now(timezone.utc) - timedelta(days=days_back)
        to_date = datetime.now(timezone.utc)
        limit = 1000
        offset = 0

        requested_dimensions = [str(v).strip() for v in (dimensions or ["sku", "day"]) if str(v).strip()]
        if "sku" not in requested_dimensions:
            requested_dimensions.append("sku")
        if "day" not in requested_dimensions and "date" not in requested_dimensions:
            requested_dimensions.append("day")

        requested_metrics = [str(v).strip() for v in (metrics or []) if str(v).strip()]
        if not requested_metrics:
            requested_metrics = [
                "revenue",
                "ordered_units",
                "hits_view_search",
                "hits_view_pdp",
                "hits_view",
                "hits_tocart_search",
                "hits_tocart_pdp",
                "hits_tocart",
                "session_view_search",
                "session_view_pdp",
                "session_view",
                "conv_tocart_search",
                "conv_tocart_pdp",
                "conv_tocart",
                "returns",
                "cancellations",
                "delivered_units",
                "position_category",
            ]

        rows_upserted = 0
        api_rows = 0

        def _extract_items(payload: Any) -> List[Dict[str, Any]]:
            if isinstance(payload, list):
                return [item for item in payload if isinstance(item, dict)]
            if not isinstance(payload, dict):
                return []
            for key in ("data", "rows", "items", "result"):
                value = payload.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
            nested = payload.get("result")
            if isinstance(nested, dict):
                for key in ("data", "rows", "items"):
                    value = nested.get(key)
                    if isinstance(value, list):
                        return [item for item in value if isinstance(item, dict)]
            return []

        def _extract_metric_value(metrics_map: Dict[str, Any], names: List[str]) -> Any:
            for name in names:
                if name in metrics_map and metrics_map.get(name) not in (None, ""):
                    return metrics_map.get(name)
            return None

        def _normalize_item(item: Dict[str, Any], single_metric: Optional[str] = None) -> Optional[Dict[str, Any]]:
            dimensions_map: Dict[str, Any] = {}
            raw_dimensions = item.get("dimensions")
            if isinstance(raw_dimensions, list):
                for idx, dim_name in enumerate(requested_dimensions):
                    if idx < len(raw_dimensions):
                        dim_value = raw_dimensions[idx]
                        if isinstance(dim_value, dict):
                            dimensions_map[dim_name] = dim_value.get("id") or dim_value.get("name")
                            if dim_value.get("name") not in (None, ""):
                                dimensions_map[f"{dim_name}_name"] = dim_value.get("name")
                        else:
                            dimensions_map[dim_name] = dim_value
            elif isinstance(raw_dimensions, dict):
                dimensions_map.update(raw_dimensions)

            for dim_name in requested_dimensions:
                if dim_name in item and dim_name not in dimensions_map:
                    dimensions_map[dim_name] = item.get(dim_name)

            sku = self._parse_int_flexible(
                dimensions_map.get("sku")
                or item.get("sku")
                or item.get("id")
            )
            row_date = self._parse_datetime_flexible(
                dimensions_map.get("day")
                or dimensions_map.get("date")
                or item.get("day")
                or item.get("date")
            )
            if not sku or row_date is None:
                return None
            row_date = self._ensure_aware_utc(row_date)

            metrics_map: Dict[str, Any] = {}
            raw_metrics = item.get("metrics")
            if isinstance(raw_metrics, list):
                if single_metric:
                    metrics_map[single_metric] = raw_metrics[0] if raw_metrics else None
                else:
                    for idx, metric_name in enumerate(requested_metrics):
                        if idx < len(raw_metrics):
                            metrics_map[metric_name] = raw_metrics[idx]
            elif isinstance(raw_metrics, dict):
                metrics_map.update(raw_metrics)

            for metric_name in requested_metrics:
                if metric_name in item and metric_name not in metrics_map:
                    metrics_map[metric_name] = item.get(metric_name)

            for key, value in item.items():
                if key in {"dimensions", "metrics"}:
                    continue
                if key in dimensions_map:
                    continue
                if key not in metrics_map:
                    metrics_map[key] = value

            impressions = self._parse_int_flexible(
                _extract_metric_value(
                    metrics_map,
                    ["impressions", "views", "session_view", "session_view_search", "session_view_pdp"],
                )
            )
            clicks = self._parse_int_flexible(
                _extract_metric_value(
                    metrics_map,
                    ["clicks", "hits_tocart_search", "hits_tocart_pdp", "hits_tocart"],
                )
            )
            ctr = self._parse_decimal_flexible(_extract_metric_value(metrics_map, ["ctr"]))

            position_category = self._parse_decimal_flexible(
                _extract_metric_value(metrics_map, ["position_category", "avg_position_category"])
            )
            position_promo = self._parse_decimal_flexible(
                _extract_metric_value(metrics_map, ["position_promo", "avg_position_promo"])
            )
            position = self._parse_decimal_flexible(
                _extract_metric_value(metrics_map, ["position", "avg_position"])
            )
            if position is None and position_category is not None:
                position = position_category

            return {
                "date": row_date,
                "sku": int(sku),
                "ordered_units": self._parse_int_flexible(
                    _extract_metric_value(metrics_map, ["ordered_units", "orders"])
                ),
                "delivered_units": self._parse_int_flexible(
                    _extract_metric_value(metrics_map, ["delivered_units"])
                ),
                "returned_units": self._parse_int_flexible(
                    _extract_metric_value(metrics_map, ["returned_units"])
                ),
                "revenue": self._parse_decimal_flexible(
                    _extract_metric_value(metrics_map, ["revenue"])
                ),
                "impressions": impressions,
                "clicks": clicks,
                "ctr": ctr,
                "position": position,
                "position_category": position_category,
                "position_promo": position_promo,
                "metric_values": metrics_map,
                "dimensions": dimensions_map,
                "raw_data": item,
                "last_synced_at": datetime.now(),
            }

        try:
            await self._ensure_analytics_data_schema()
            merged_rows: Dict[tuple, Dict[str, Any]] = {}

            for metric_name in requested_metrics:
                offset = 0
                try:
                    while True:
                        result = await self._with_rate_limit_retry(
                            lambda m=metric_name, o=offset: self.client.get_analytics_data(
                                date_from=from_date,
                                date_to=to_date,
                                metrics=[m],
                                dimension=requested_dimensions,
                                limit=limit,
                                offset=o,
                            ),
                            attempts=3,
                            base_delay=5.0,
                        )
                        items = _extract_items(result)
                        if not items:
                            break
                        api_rows += len(items)

                        for item in items:
                            row = _normalize_item(item, single_metric=metric_name)
                            if not row:
                                continue
                            key = (row["date"], row["sku"])
                            existing = merged_rows.get(key)
                            if not existing:
                                existing = {
                                    "date": row["date"],
                                    "sku": row["sku"],
                                    "ordered_units": None,
                                    "delivered_units": None,
                                    "returned_units": None,
                                    "revenue": None,
                                    "impressions": None,
                                    "clicks": None,
                                    "ctr": None,
                                    "position": None,
                                    "position_category": None,
                                    "position_promo": None,
                                    "metric_values": {},
                                    "dimensions": row.get("dimensions") or {},
                                    "raw_data": row.get("raw_data"),
                                    "last_synced_at": datetime.now(),
                                }
                                merged_rows[key] = existing

                            metric_val = (row.get("metric_values") or {}).get(metric_name)
                            existing["metric_values"][metric_name] = metric_val
                            existing["dimensions"] = row.get("dimensions") or existing.get("dimensions")
                            existing["raw_data"] = row.get("raw_data") or existing.get("raw_data")
                            existing["last_synced_at"] = datetime.now()

                            if metric_name == "ordered_units":
                                existing["ordered_units"] = self._parse_int_flexible(metric_val)
                            elif metric_name == "delivered_units":
                                existing["delivered_units"] = self._parse_int_flexible(metric_val)
                            elif metric_name == "returned_units":
                                existing["returned_units"] = self._parse_int_flexible(metric_val)
                            elif metric_name == "revenue":
                                existing["revenue"] = self._parse_decimal_flexible(metric_val)
                            elif metric_name in {"hits_view", "hits_view_search", "hits_view_pdp", "session_view", "session_view_search", "session_view_pdp"}:
                                existing["impressions"] = self._parse_int_flexible(metric_val)
                            elif metric_name in {"hits_tocart_search", "hits_tocart_pdp", "hits_tocart"}:
                                existing["clicks"] = self._parse_int_flexible(metric_val)
                            elif metric_name == "ctr":
                                existing["ctr"] = self._parse_decimal_flexible(metric_val)
                            elif metric_name == "position":
                                existing["position"] = self._parse_decimal_flexible(metric_val)
                            elif metric_name == "position_category":
                                existing["position_category"] = self._parse_decimal_flexible(metric_val)
                                if existing.get("position") is None:
                                    existing["position"] = existing["position_category"]
                            elif metric_name == "position_promo":
                                existing["position_promo"] = self._parse_decimal_flexible(metric_val)

                        total = None
                        if isinstance(result, dict):
                            total = self._parse_int_flexible(result.get("total"))
                            if total is None and isinstance(result.get("result"), dict):
                                total = self._parse_int_flexible(result.get("result", {}).get("total"))

                        if len(items) < limit:
                            break
                        offset += limit
                        if total is not None and offset >= total:
                            break
                except OzonAPIError as metric_error:
                    logger.warning("analytics_data: metric '%s' skipped (%s)", metric_name, metric_error)
                    continue

            async with db_manager.session() as session:
                for row in merged_rows.values():
                    stmt = pg_insert(AnalyticsData).values(**row)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["date", "sku"],
                        set_={k: v for k, v in row.items() if k not in {"date", "sku"}},
                    )
                    await session.execute(stmt)
                    rows_upserted += 1

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
            )
            logger.info(
                "analytics_data sync completed: rows_upserted=%s api_rows=%s",
                rows_upserted,
                api_rows,
            )
            return {
                "rows_upserted": rows_upserted,
                "api_rows": api_rows,
                "days_back": days_back,
            }
        except Exception as e:
            logger.error(f"analytics_data sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_analytics_product_queries(
        self,
        days_back: int = 30,
        limit_by_sku: int = 15,
        sku_batch_size: int = 1000,
        max_availability_probe_days: int = 7,
    ) -> Dict[str, int]:
        """Синхронизация аналитики поисковых запросов по SKU."""
        logger.info("Starting analytics_product_queries sync...")
        sync_log = await self._create_sync_log("analytics_product_queries")

        now_utc = datetime.now(timezone.utc)
        last_complete_day = datetime(now_utc.year, now_utc.month, now_utc.day, tzinfo=timezone.utc) - timedelta(days=1)
        if days_back <= 0:
            days_back = 30
        granularity = "day" if days_back <= 31 else "week"
        from_date: Optional[datetime] = None
        to_date: Optional[datetime] = None
        window_iter = iter(())

        summary_rows_upserted = 0
        detail_rows_upserted = 0
        windows_processed = 0
        api_rows = 0

        def _extract_result_container(payload: Any) -> Dict[str, Any]:
            if isinstance(payload, dict) and isinstance(payload.get("result"), dict):
                return payload.get("result") or {}
            return payload if isinstance(payload, dict) else {}

        def _extract_list(container: Any, keys: List[str]) -> List[Dict[str, Any]]:
            if isinstance(container, list):
                return [item for item in container if isinstance(item, dict)]
            if not isinstance(container, dict):
                return []
            for key in keys:
                value = container.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
            return []

        def _extract_page_count(container: Dict[str, Any]) -> Optional[int]:
            return self._parse_int_flexible(container.get("page_count") or container.get("pages") or container.get("pageCount"))

        def _extract_total(container: Dict[str, Any]) -> Optional[int]:
            return self._parse_int_flexible(container.get("total") or container.get("count"))

        def _is_no_data_error(exc: Exception) -> bool:
            if not isinstance(exc, OzonAPIError):
                return False
            if exc.status_code != 400:
                return False
            message = str(exc).lower()
            return "there is no data for the specified period" in message or "no data for the specified period" in message

        def _metric_int(source: Dict[str, Any], names: List[str]) -> Optional[int]:
            for name in names:
                if name in source and source.get(name) not in (None, ""):
                    return self._parse_int_flexible(source.get(name))
            return None

        def _metric_decimal(source: Dict[str, Any], names: List[str]) -> Optional[float]:
            for name in names:
                if name in source and source.get(name) not in (None, ""):
                    value = self._parse_decimal_flexible(source.get(name))
                    return float(value) if value is not None else None
            return None

        def _normalize_summary_row(
            item: Dict[str, Any],
            period_start: datetime,
            period_end: datetime,
            sku_ref: Dict[int, Dict[str, Any]],
        ) -> Optional[Dict[str, Any]]:
            sku = self._parse_int_flexible(item.get("sku") or item.get("id") or item.get("product_id"))
            if not sku:
                product = item.get("product") if isinstance(item.get("product"), dict) else {}
                sku = self._parse_int_flexible(product.get("sku") or product.get("id"))
            if not sku:
                return None
            ref = sku_ref.get(int(sku), {})
            return {
                "period_start": period_start,
                "period_end": period_end,
                "granularity": granularity,
                "sku": int(sku),
                "offer_id": str(item.get("offer_id") or ref.get("offer_id") or "").strip() or None,
                "product_name": str(
                    item.get("product_name")
                    or item.get("name")
                    or ref.get("product_name")
                    or ""
                ).strip() or None,
                "searches": _metric_int(item, ["searches", "searches_count", "queries_count", "unique_search_users"]),
                "views": _metric_int(item, ["views", "views_count", "unique_view_users"]),
                "avg_position": _metric_decimal(item, ["avg_position", "position", "position_avg"]),
                "conversion": _metric_decimal(item, ["conversion", "conversion_rate", "avg_conversion", "view_conversion"]),
                "gmv": _metric_decimal(item, ["gmv", "revenue", "sales"]),
                "raw_data": item,
                "last_synced_at": datetime.now(),
            }

        def _normalize_detail_rows(
            parent_item: Dict[str, Any],
            period_start: datetime,
            period_end: datetime,
            sku_ref: Dict[int, Dict[str, Any]],
        ) -> List[Dict[str, Any]]:
            rows: List[Dict[str, Any]] = []
            parent_sku = self._parse_int_flexible(parent_item.get("sku") or parent_item.get("id") or parent_item.get("product_id"))
            query_items = _extract_list(parent_item, ["queries", "items", "rows", "data"])
            if query_items and parent_sku:
                base_offer_id = str(parent_item.get("offer_id") or "").strip() or None
                base_name = str(parent_item.get("product_name") or parent_item.get("name") or "").strip() or None
                for query_item in query_items:
                    ref = sku_ref.get(int(parent_sku), {})
                    query_text = str(
                        query_item.get("query")
                        or query_item.get("search_query")
                        or query_item.get("text")
                        or query_item.get("phrase")
                        or ""
                    ).strip()
                    if not query_text:
                        continue
                    rows.append(
                        {
                            "period_start": period_start,
                            "period_end": period_end,
                            "granularity": granularity,
                            "sku": int(parent_sku),
                            "offer_id": base_offer_id or ref.get("offer_id"),
                            "product_name": base_name or ref.get("product_name"),
                            "query_text": query_text[:1000],
                            "searches": _metric_int(query_item, ["searches", "searches_count", "queries_count", "unique_search_users"]),
                            "views": _metric_int(query_item, ["views", "views_count", "unique_view_users"]),
                            "avg_position": _metric_decimal(query_item, ["avg_position", "position", "position_avg"]),
                            "conversion": _metric_decimal(query_item, ["conversion", "conversion_rate", "avg_conversion", "view_conversion"]),
                            "gmv": _metric_decimal(query_item, ["gmv", "revenue", "sales"]),
                            "raw_data": {"parent": parent_item, "query": query_item},
                            "last_synced_at": datetime.now(),
                        }
                    )
                return rows

            query_text = str(
                parent_item.get("query")
                or parent_item.get("search_query")
                or parent_item.get("text")
                or parent_item.get("phrase")
                or ""
            ).strip()
            if not query_text:
                return rows
            sku = self._parse_int_flexible(parent_item.get("sku") or parent_item.get("product_id"))
            if not sku:
                return rows
            ref = sku_ref.get(int(sku), {})
            rows.append(
                {
                    "period_start": period_start,
                    "period_end": period_end,
                    "granularity": granularity,
                    "sku": int(sku),
                    "offer_id": str(parent_item.get("offer_id") or ref.get("offer_id") or "").strip() or None,
                    "product_name": str(parent_item.get("product_name") or parent_item.get("name") or ref.get("product_name") or "").strip() or None,
                    "query_text": query_text[:1000],
                    "searches": _metric_int(parent_item, ["searches", "searches_count", "queries_count", "unique_search_users"]),
                    "views": _metric_int(parent_item, ["views", "views_count", "unique_view_users"]),
                    "avg_position": _metric_decimal(parent_item, ["avg_position", "position", "position_avg"]),
                    "conversion": _metric_decimal(parent_item, ["conversion", "conversion_rate", "avg_conversion", "view_conversion"]),
                    "gmv": _metric_decimal(parent_item, ["gmv", "revenue", "sales"]),
                    "raw_data": parent_item,
                    "last_synced_at": datetime.now(),
                }
            )
            return rows

        try:
            await self._ensure_analytics_product_queries_schema()
            sku_ref = await self._load_product_query_sku_reference()
            skus = sorted(sku_ref.keys())
            if not skus:
                raise ValueError("No SKUs found in local reference tables for /v1/analytics/product-queries")

            latest_available_day: Optional[datetime] = None
            probe_days = max(1, min(int(max_availability_probe_days), 14))
            probe_skus = skus[: min(50, len(skus))]
            for lag in range(probe_days):
                probe_day = last_complete_day - timedelta(days=lag)
                probe_start = probe_day.replace(hour=0, minute=0, second=0)
                probe_end = probe_day.replace(hour=23, minute=59, second=59)
                try:
                    probe_payload = await self._with_rate_limit_retry(
                        lambda start=probe_start, end=probe_end, chunk=probe_skus: self.client.get_analytics_product_queries(
                            date_from=start,
                            date_to=end,
                            skus=chunk,
                            page=0,
                            page_size=1,
                        ),
                        attempts=4,
                        base_delay=3.0,
                    )
                    probe_container = _extract_result_container(probe_payload)
                    probe_items = _extract_list(probe_container, ["items", "rows", "data", "products"])
                    if probe_items:
                        latest_available_day = probe_day
                        break
                except OzonAPIError as exc:
                    if _is_no_data_error(exc):
                        continue
                    raise

            if latest_available_day is None:
                logger.warning(
                    "analytics_product_queries: no available data detected for last %s day(s), skipping update",
                    probe_days,
                )
                await self._update_sync_log(
                    sync_log,
                    "success",
                    records_processed=0,
                    records_inserted=0,
                    records_updated=0,
                )
                return {
                    "summary_rows_upserted": 0,
                    "detail_rows_upserted": 0,
                    "windows_processed": 0,
                    "api_rows": 0,
                    "days_back": days_back,
                    "granularity": granularity,
                    "latest_available_date": None,
                    "requested_last_complete_date": last_complete_day.date().isoformat(),
                }

            if latest_available_day.date() != last_complete_day.date():
                logger.info(
                    "analytics_product_queries: latest available date is %s (requested latest complete date %s)",
                    latest_available_day.date().isoformat(),
                    last_complete_day.date().isoformat(),
                )

            from_date = latest_available_day - timedelta(days=days_back - 1)
            to_date = latest_available_day.replace(hour=23, minute=59, second=59)
            window_iter = self._iter_day_windows(from_date, to_date) if granularity == "day" else self._iter_week_windows(from_date, to_date)

            async with db_manager.session() as session:
                for period_start, period_end in window_iter:
                    windows_processed += 1
                    logger.info(
                        "analytics_product_queries: processing %s window %s .. %s",
                        granularity,
                        period_start.isoformat(),
                        period_end.isoformat(),
                    )
                    for chunk_start in range(0, len(skus), max(1, int(sku_batch_size))):
                        sku_chunk = skus[chunk_start:chunk_start + max(1, int(sku_batch_size))]

                        summary_page = 0
                        while True:
                            try:
                                summary_payload = await self._with_rate_limit_retry(
                                    lambda p=summary_page, chunk=sku_chunk, start=period_start, end=period_end: self.client.get_analytics_product_queries(
                                        date_from=start,
                                        date_to=end,
                                        skus=chunk,
                                        page=p,
                                        page_size=100,
                                    ),
                                    attempts=8,
                                    base_delay=5.0,
                                )
                            except OzonAPIError as exc:
                                if _is_no_data_error(exc):
                                    logger.info(
                                        "analytics_product_queries: no summary data for %s .. %s",
                                        period_start.date().isoformat(),
                                        period_end.date().isoformat(),
                                    )
                                    break
                                raise
                            summary_container = _extract_result_container(summary_payload)
                            summary_items = _extract_list(summary_container, ["items", "rows", "data", "products"])
                            if not summary_items:
                                break
                            api_rows += len(summary_items)
                            for item in summary_items:
                                row = _normalize_summary_row(item, period_start, period_end, sku_ref)
                                if not row:
                                    continue
                                stmt = pg_insert(AnalyticsProductQuerySummary).values(**row)
                                stmt = stmt.on_conflict_do_update(
                                    index_elements=["period_start", "period_end", "granularity", "sku"],
                                    set_={k: v for k, v in row.items() if k not in {"period_start", "period_end", "granularity", "sku"}},
                                )
                                await session.execute(stmt)
                                summary_rows_upserted += 1

                            page_count = _extract_page_count(summary_container)
                            total = _extract_total(summary_container)
                            summary_page += 1
                            if page_count is not None and summary_page >= page_count:
                                break
                            if total is not None and summary_page * 100 >= total:
                                break
                            if len(summary_items) < 100:
                                break

                        details_page = 0
                        while True:
                            try:
                                details_payload = await self._with_rate_limit_retry(
                                    lambda p=details_page, chunk=sku_chunk, start=period_start, end=period_end: self.client.get_analytics_product_queries_details(
                                        date_from=start,
                                        date_to=end,
                                        skus=chunk,
                                        page=p,
                                        page_size=100,
                                        limit_by_sku=max(1, min(int(limit_by_sku), 15)),
                                    ),
                                    attempts=8,
                                    base_delay=5.0,
                                )
                            except OzonAPIError as exc:
                                if _is_no_data_error(exc):
                                    logger.info(
                                        "analytics_product_queries: no details data for %s .. %s",
                                        period_start.date().isoformat(),
                                        period_end.date().isoformat(),
                                    )
                                    break
                                raise
                            details_container = _extract_result_container(details_payload)
                            detail_items = _extract_list(details_container, ["items", "rows", "data", "queries"])
                            if not detail_items:
                                break
                            api_rows += len(detail_items)
                            for item in detail_items:
                                for row in _normalize_detail_rows(item, period_start, period_end, sku_ref):
                                    stmt = pg_insert(AnalyticsProductQueryDetail).values(**row)
                                    stmt = stmt.on_conflict_do_update(
                                        index_elements=["period_start", "period_end", "granularity", "sku", "query_text"],
                                        set_={k: v for k, v in row.items() if k not in {"period_start", "period_end", "granularity", "sku", "query_text"}},
                                    )
                                    await session.execute(stmt)
                                    detail_rows_upserted += 1

                            page_count = _extract_page_count(details_container)
                            total = _extract_total(details_container)
                            details_page += 1
                            if page_count is not None and details_page >= page_count:
                                break
                            if total is not None and details_page * 100 >= total:
                                break
                            if len(detail_items) < 100:
                                break

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=summary_rows_upserted + detail_rows_upserted,
                records_inserted=summary_rows_upserted + detail_rows_upserted,
                records_updated=0,
            )
            logger.info(
                "analytics_product_queries sync completed: summary=%s details=%s windows=%s api_rows=%s",
                summary_rows_upserted,
                detail_rows_upserted,
                windows_processed,
                api_rows,
            )
            return {
                "summary_rows_upserted": summary_rows_upserted,
                "detail_rows_upserted": detail_rows_upserted,
                "windows_processed": windows_processed,
                "api_rows": api_rows,
                "days_back": days_back,
                "granularity": granularity,
                "latest_available_date": latest_available_day.date().isoformat() if latest_available_day else None,
                "requested_last_complete_date": last_complete_day.date().isoformat(),
            }
        except Exception as e:
            logger.error(f"analytics_product_queries sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_fbs_warehouse_stocks(self) -> Dict[str, int]:
        """Sync FBS warehouse stocks via /v1/product/info/warehouse/stocks per warehouse_id."""
        logger.info("Starting fbs_warehouse_stocks sync...")
        sync_log = await self._create_sync_log("fbs_warehouse_stocks")

        try:
            from .config import settings
            wh_ids_raw = settings.fbs_warehouse_ids.strip()
            if not wh_ids_raw:
                raise ValueError(
                    "FBS_WAREHOUSE_IDS not set in .env — "
                    "provide comma-separated warehouse IDs for FBS stock sync"
                )
            # Parse format: "id1:Name1,id2:Name2" or "id1,id2"
            warehouse_ids: List[int] = []
            wh_names: Dict[int, str] = {}
            for part in wh_ids_raw.split(","):
                part = part.strip()
                if not part:
                    continue
                if ":" in part:
                    id_str, name = part.split(":", 1)
                    wh_id = int(id_str.strip())
                    wh_names[wh_id] = name.strip()
                else:
                    wh_id = int(part)
                warehouse_ids.append(wh_id)
            if not warehouse_ids:
                raise ValueError("FBS_WAREHOUSE_IDS contains no valid integer IDs")

            rows: List[Dict[str, Any]] = []
            for wh_id in warehouse_ids:
                logger.info(f"fbs_warehouse_stocks: fetching stocks for warehouse_id={wh_id}")
                wh_stocks = await self.client.get_all_warehouse_stocks(wh_id)
                logger.info(f"fbs_warehouse_stocks: warehouse {wh_id} returned {len(wh_stocks)} items")
                for item in wh_stocks:
                    item["_warehouse_id"] = wh_id
                rows.extend(wh_stocks)

            rows_upserted = 0
            async with db_manager.session() as session:
                await session.execute(delete(FBSWarehouseStock))

                for row in rows:
                    warehouse_id = self._parse_int_flexible(row.get("_warehouse_id"))
                    product_id = self._parse_int_flexible(row.get("product_id"))
                    sku = self._parse_int_flexible(row.get("sku")) or product_id
                    if not warehouse_id or not sku:
                        continue

                    present = self._parse_int_flexible(row.get("present", 0))
                    reserved = self._parse_int_flexible(row.get("reserved", 0))

                    data = {
                        "sku": sku,
                        "offer_id": row.get("offer_id"),
                        "product_id": product_id,
                        "warehouse_id": warehouse_id,
                        "warehouse_name": wh_names.get(warehouse_id) or row.get("warehouse_name"),
                        "present": present,
                        "reserved": reserved,
                        "raw_data": row,
                        "last_synced_at": datetime.now(),
                    }
                    stmt = pg_insert(FBSWarehouseStock).values(**data)
                    stmt = stmt.on_conflict_do_update(
                        constraint="uq_fbs_warehouse_stocks_sku_wh",
                        set_={k: v for k, v in data.items() if k not in {"sku", "warehouse_id"}},
                    )
                    await session.execute(stmt)
                    rows_upserted += 1

            await self._update_sync_log(sync_log, "success", rows_upserted)
            logger.info(f"fbs_warehouse_stocks sync completed: rows_upserted={rows_upserted}")
            return {
                "rows_upserted": rows_upserted,
                "warehouses_queried": len(warehouse_ids),
                "api_items_received": len(rows),
            }
        except Exception as e:
            logger.error(f"fbs_warehouse_stocks sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_analytics_turnover(self, days_back: int = 30) -> Dict[str, int]:
        """Sinhronizacija /v1/analytics/turnover/stocks."""
        logger.info("Starting analytics_turnover sync...")
        sync_log = await self._create_sync_log("analytics_turnover")

        from_date = datetime.now(timezone.utc) - timedelta(days=days_back)
        to_date = datetime.now(timezone.utc)
        limit = 1000
        offset = 0
        records_processed = 0
        rows_to_insert: List[Dict[str, Any]] = []

        try:
            while True:
                result = await self._with_rate_limit_retry(
                    lambda: self.client.get_analytics_turnover(
                        date_from=from_date,
                        date_to=to_date,
                        limit=limit,
                        offset=offset,
                    )
                )
                items = []
                if isinstance(result, dict):
                    items = result.get("items") or result.get("result") or result.get("rows") or result.get("data") or []
                if not isinstance(items, list) or not items:
                    break

                for item in items:
                    item_date = self._parse_datetime_flexible(
                        item.get("date")
                        or item.get("updated_at")
                        or item.get("period_from")
                    ) or datetime.now(timezone.utc)
                    rows_to_insert.append(
                        {
                            "date": item_date,
                            "sku": self._parse_int_flexible(item.get("sku")) or 0,
                            "stock": self._parse_int_flexible(item.get("stock")) or self._parse_int_flexible(item.get("available_stock_count")) or 0,
                            "sales_speed": self._parse_decimal_flexible(item.get("sales_speed") or item.get("ads")),
                            "days_in_stock": self._parse_int_flexible(item.get("days_in_stock") or item.get("days_without_sales")),
                            "recommended_stock": self._parse_int_flexible(item.get("recommended_stock")),
                            "recommended_supply": self._parse_int_flexible(item.get("recommended_supply")),
                            "raw_data": item,
                            "last_synced_at": datetime.now(),
                        }
                    )

                total = self._parse_int_flexible(result.get("total")) if isinstance(result, dict) else None
                if len(items) < limit:
                    break
                offset += limit
                if total is not None and offset >= total:
                    break

            async with db_manager.session() as session:
                await session.execute(
                    delete(AnalyticsTurnover).where(
                        AnalyticsTurnover.date >= from_date,
                        AnalyticsTurnover.date <= to_date,
                    )
                )
                for row in rows_to_insert:
                    if not row["sku"]:
                        continue
                    await session.execute(insert(AnalyticsTurnover).values(**row))
                    records_processed += 1

            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"analytics_turnover sync completed: {records_processed} processed")
            return {"processed": records_processed}
        except Exception as e:
            logger.error(f"analytics_turnover sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_analytics_average_delivery_time(self) -> Dict[str, int]:
        """Sinhronizacija /v1/analytics/average-delivery-time."""
        logger.info("Starting analytics_average_delivery_time sync...")
        sync_log = await self._create_sync_log("analytics_average_delivery_time")

        records_processed = 0
        try:
            result = await self._with_rate_limit_retry(
                lambda: self.client.get_analytics_average_delivery_time({})
            )
            items = result.get("data", []) if isinstance(result, dict) else []
            if not isinstance(items, list):
                items = []

            async with db_manager.session() as session:
                for item in items:
                    metrics = item.get("metrics", {}) if isinstance(item.get("metrics"), dict) else {}
                    orders_count = metrics.get("orders_count", {}) if isinstance(metrics.get("orders_count"), dict) else {}
                    fast = orders_count.get("fast", {}) if isinstance(orders_count.get("fast"), dict) else {}
                    medium = orders_count.get("medium", {}) if isinstance(orders_count.get("medium"), dict) else {}
                    long_ = orders_count.get("long", {}) if isinstance(orders_count.get("long"), dict) else {}

                    row = {
                        "delivery_cluster_id": self._parse_int_flexible(item.get("delivery_cluster_id")) or 0,
                        "average_delivery_time": self._parse_decimal_flexible(metrics.get("average_delivery_time")),
                        "average_delivery_time_status": metrics.get("average_delivery_time_status"),
                        "lost_profit": self._parse_decimal_flexible(metrics.get("lost_profit")),
                        "exact_impact_share": self._parse_decimal_flexible(metrics.get("exact_impact_share")),
                        "attention_level": metrics.get("attention_level"),
                        "recommended_supply": self._parse_int_flexible(metrics.get("recommended_supply")),
                        "orders_total": self._parse_int_flexible(orders_count.get("total")),
                        "orders_fast": self._parse_int_flexible(fast.get("value")),
                        "orders_fast_percent": self._parse_decimal_flexible(fast.get("percent")),
                        "orders_medium": self._parse_int_flexible(medium.get("value")),
                        "orders_medium_percent": self._parse_decimal_flexible(medium.get("percent")),
                        "orders_long": self._parse_int_flexible(long_.get("value")),
                        "orders_long_percent": self._parse_decimal_flexible(long_.get("percent")),
                        "clusters_data": item.get("clusters_data"),
                        "raw_data": item,
                        "last_synced_at": datetime.now(),
                    }
                    if not row["delivery_cluster_id"]:
                        continue
                    stmt = pg_insert(AnalyticsAverageDeliveryTime).values(**row)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["delivery_cluster_id"],
                        set_={k: v for k, v in row.items() if k != "delivery_cluster_id"},
                    )
                    await session.execute(stmt)
                    records_processed += 1

            try:
                await self._capture_delivery_time_daily_snapshot(session)
            except Exception as snap_err:
                logger.error(f"delivery_time_daily_snapshot failed (non-fatal): {snap_err}")

            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"analytics_average_delivery_time sync completed: {records_processed} processed")
            return {"processed": records_processed}
        except Exception as e:
            logger.error(f"analytics_average_delivery_time sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_realization_v2(self, days_back: int = 365) -> Dict[str, int]:
        """Sinhronizacija /v2/finance/realization za diapazon mesjacev."""
        logger.info("Starting realization_v2 sync...")
        sync_log = await self._create_sync_log("realization_v2")

        records_processed = 0
        months_processed = 0
        now_utc = datetime.now(timezone.utc)
        start_date = now_utc - timedelta(days=days_back)
        month_cursor = datetime(start_date.year, start_date.month, 1, tzinfo=timezone.utc)
        end_month = datetime(now_utc.year, now_utc.month, 1, tzinfo=timezone.utc)

        try:
            while month_cursor <= end_month:
                year = month_cursor.year
                month = month_cursor.month
                month_start = datetime(year, month, 1, tzinfo=timezone.utc)
                if month == 12:
                    month_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
                else:
                    month_end = datetime(year, month + 1, 1, tzinfo=timezone.utc)

                try:
                    result = await self._with_rate_limit_retry(
                        lambda y=year, m=month: self.client.get_realization_report(y, m)
                    )
                except OzonAPIError as e:
                    if e.status_code == 404:
                        logger.info(f"realization_v2: month {year}-{month:02d} is not prepared yet, skipping")
                        month_cursor = month_end
                        continue
                    raise

                result_block = result.get("result", {}) if isinstance(result, dict) else {}
                header = result_block.get("header", {}) if isinstance(result_block.get("header"), dict) else {}
                rows = result_block.get("rows", []) if isinstance(result_block.get("rows"), list) else []

                async with db_manager.session() as session:
                    await session.execute(
                        delete(RealizationReport).where(
                            RealizationReport.date >= month_start,
                            RealizationReport.date < month_end,
                        )
                    )

                    doc_date = self._parse_datetime_flexible(header.get("doc_date")) or month_start
                    for row in rows:
                        item = row.get("item", {}) if isinstance(row.get("item"), dict) else {}
                        delivery_commission = row.get("delivery_commission", {}) if isinstance(row.get("delivery_commission"), dict) else {}
                        return_commission = row.get("return_commission", {}) if isinstance(row.get("return_commission"), dict) else {}
                        quantity = self._parse_int_flexible(delivery_commission.get("quantity")) or 0
                        price_per_instance = self._parse_decimal_flexible(row.get("seller_price_per_instance")) or 0
                        total_amount = price_per_instance * quantity if quantity else price_per_instance

                        row_data = {
                            "date": doc_date,
                            "sku": self._parse_int_flexible(item.get("sku")) or 0,
                            "offer_id": item.get("offer_id"),
                            "name": item.get("name"),
                            "quantity": quantity,
                            "price": price_per_instance,
                            "total_amount": total_amount,
                            "commission_percent": self._parse_decimal_flexible(row.get("commission_ratio")) * 100
                            if self._parse_decimal_flexible(row.get("commission_ratio")) is not None
                            else None,
                            "commission_amount": self._parse_decimal_flexible(delivery_commission.get("commission")),
                            "payout_amount": self._parse_decimal_flexible(delivery_commission.get("bonus")),
                            "delivery_cost": self._parse_decimal_flexible(delivery_commission.get("amount")),
                            "total_payout": self._parse_decimal_flexible(delivery_commission.get("total")),
                            "raw_data": {
                                "header": header,
                                "row": row,
                                "return_commission": return_commission,
                            },
                            "last_synced_at": datetime.now(),
                        }
                        if not row_data["sku"]:
                            continue
                        inserted = await session.execute(insert(RealizationReport).returning(RealizationReport.id).values(**row_data))
                        realization_report_id = inserted.scalar_one()
                        await self._upsert_realization_detail(
                            session=session,
                            realization_report_id=realization_report_id,
                            raw_data=row_data["raw_data"],
                        )
                        records_processed += 1

                months_processed += 1
                month_cursor = month_end

            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"realization_v2 sync completed: rows={records_processed}, months={months_processed}")
            return {"processed": records_processed, "months_processed": months_processed}
        except Exception as e:
            logger.error(f"realization_v2 sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== POSTINGS SYNC ====================
    
    async def sync_postings(self, days_back: int = 30) -> Dict[str, int]:
        """Sinhronizacija otpravlenij FBS."""
        logger.info(f"Starting postings sync for last {days_back} days...")
        sync_log = await self._create_sync_log("postings")
        
        records_processed = 0
        
        try:
            to_date = datetime.now(timezone.utc)
            from_date = to_date - timedelta(days=days_back)
            
            async for postings in self.client.get_all_postings(from_date, to_date):
                if isinstance(postings, dict):
                    postings = postings.get("postings") or postings.get("result") or []
                for posting_data in postings:
                    if not isinstance(posting_data, dict):
                        continue
                    posting_number = posting_data.get("posting_number")
                    customer = posting_data.get("customer") or {}
                    address = customer.get("address") or {}
                    delivery_method = posting_data.get("delivery_method") or {}
                    
                    posting_dict = {
                        "posting_number": posting_number,
                        "order_id": posting_data.get("order_id"),
                        "order_number": posting_data.get("order_number"),
                        "status": posting_data.get("status"),
                        "created_at": self._parse_datetime(posting_data.get("created_at")),
                        "in_process_at": self._parse_datetime(posting_data.get("in_process_at")),
                        "shipment_date": self._parse_datetime(posting_data.get("shipment_date")),
                        "delivered_at": self._parse_datetime(posting_data.get("delivered_at")),
                        "delivery_schema": "FBS",
                        "total_price": self._parse_decimal(posting_data.get("price")),
                        "total_discount": self._parse_decimal(posting_data.get("discount_amount")),
                        "tracking_number": posting_data.get("tracking_number"),
                        "delivery_method_name": delivery_method.get("name"),
                        "customer_name": customer.get("name"),
                        "customer_phone": customer.get("phone"),
                        "address": str(address),
                        "city": address.get("city"),
                        "region": address.get("region"),
                        "raw_data": posting_data,
                        "last_synced_at": datetime.now()
                    }
                    
                    async with db_manager.session() as session:
                        stmt = pg_insert(Posting).values(**posting_dict)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["posting_number"],
                            set_={k: v for k, v in posting_dict.items() if k != "posting_number"}
                        )
                        await session.execute(stmt)
                        records_processed += 1
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Postings sync completed: {records_processed} processed")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"Postings sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== POSTINGS FBO SYNC ====================
    
    async def sync_postings_fbo(self, days_back: int = 30) -> Dict[str, int]:
        """Sinhronizacija otpravlenij FBO."""
        logger.info(f"Starting FBO postings sync for last {days_back} days...")
        sync_log = await self._create_sync_log("postings_fbo")
        
        records_processed = 0
        
        try:
            to_date = datetime.now(timezone.utc)
            from_date = to_date - timedelta(days=days_back)
            
            async for postings in self.client.get_all_postings_fbo(from_date, to_date):
                if isinstance(postings, dict):
                    postings = postings.get("postings") or postings.get("result") or []
                for posting_data in postings:
                    if not isinstance(posting_data, dict):
                        continue
                    posting_dict = {
                        "posting_number": posting_data.get("posting_number"),
                        "order_id": posting_data.get("order_id"),
                        "status": posting_data.get("status"),
                        "created_at": self._parse_datetime(posting_data.get("created_at")),
                        "in_process_at": self._parse_datetime(posting_data.get("in_process_at")),
                        "shipment_date": self._parse_datetime(posting_data.get("shipment_date")),
                        "delivered_at": self._parse_datetime(posting_data.get("delivered_at")),
                        "total_price": self._parse_decimal(posting_data.get("price")),
                        "total_discount": self._parse_decimal(posting_data.get("discount_amount")),
                        "warehouse_id": posting_data.get("analytics_data", {}).get("warehouse_id"),
                        "warehouse_name": posting_data.get("analytics_data", {}).get("warehouse_name"),
                        "items": posting_data.get("products", []),
                        "raw_data": posting_data,
                        "last_synced_at": datetime.now()
                    }
                    
                    async with db_manager.session() as session:
                        stmt = pg_insert(PostingFBO).values(**posting_dict)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["posting_number"],
                            set_={k: v for k, v in posting_dict.items() if k != "posting_number"}
                        )
                        await session.execute(stmt)
                        records_processed += 1
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"FBO postings sync completed: {records_processed} processed")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"FBO postings sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== RETURNS SYNC ====================
    
    async def sync_returns(self) -> Dict[str, int]:
        """Sinhronizacija vozvratov FBS (cherez /v1/returns/list)."""
        logger.info("Starting returns sync...")
        sync_log = await self._create_sync_log("returns")
        
        records_processed = 0
        batch = []
        batch_size = 500  # Batch dlja bulk insert
        
        try:
            now_utc = datetime.now(timezone.utc)
            ytd_from = now_utc.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            filter_payload = {
                "visual_status_change_moment": {
                    "time_from": ytd_from.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "time_to": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                }
            }

            async for returns in self.client.get_all_returns(filter_payload=filter_payload, limit=500):
                for return_data in returns:
                    normalized = self._normalize_return_payload(return_data)
                    return_dict = {
                        "return_id": normalized["return_id"],
                        "posting_number": normalized["posting_number"],
                        "sku": normalized["sku"],
                        "offer_id": normalized["offer_id"],
                        "product_name": normalized["product_name"],
                        "quantity": normalized["quantity"],
                        "return_reason": normalized["return_reason"],
                        "status": normalized["status"],
                        "returned_at": normalized["returned_at"],
                        "refund_amount": normalized["refund_amount"],
                        "raw_data": return_data,
                        "last_synced_at": datetime.now()
                    }
                    
                    batch.append(return_dict)
                    
                    # Bulk insert kogda nakopilos' batch_size
                    if len(batch) >= batch_size:
                        async with db_manager.session() as session:
                            for item in batch:
                                stmt = pg_insert(Return).values(**item)
                                stmt = stmt.on_conflict_do_update(
                                    index_elements=["return_id"],
                                    set_={k: v for k, v in item.items() if k != "return_id"}
                                )
                                await session.execute(stmt)
                        records_processed += len(batch)
                        logger.info(f"Returns processed: {records_processed}")
                        batch = []
            
            # Obrabatyvaem ostatok
            if batch:
                async with db_manager.session() as session:
                    for item in batch:
                        stmt = pg_insert(Return).values(**item)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["return_id"],
                            set_={k: v for k, v in item.items() if k != "return_id"}
                        )
                        await session.execute(stmt)
                records_processed += len(batch)
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Returns sync completed: {records_processed} processed")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"Returns sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== RETURNS FBO SYNC ====================
    
    async def sync_returns_fbo(self) -> Dict[str, int]:
        """Sinhronizacija vozvratov FBO."""
        logger.info("Starting FBO returns sync...")
        sync_log = await self._create_sync_log("returns_fbo")
        
        records_processed = 0
        
        try:
            now_utc = datetime.now(timezone.utc)
            ytd_from = now_utc.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            filter_payload = {
                "visual_status_change_moment": {
                    "time_from": ytd_from.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "time_to": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                }
            }
            async for returns in self.client.get_all_returns_fbo(filter_payload=filter_payload, limit=500):
                for return_data in returns:
                    normalized = self._normalize_return_payload(return_data)
                    return_dict = {
                        "return_id": normalized["return_id"],
                        "posting_number": normalized["posting_number"],
                        "sku": normalized["sku"],
                        "offer_id": normalized["offer_id"],
                        "product_name": normalized["product_name"],
                        "quantity": normalized["quantity"],
                        "status": normalized["status"],
                        "return_reason": normalized["return_reason"],
                        "returned_at": normalized["returned_at"],
                        "refund_amount": normalized["refund_amount"],
                        "raw_data": return_data,
                        "last_synced_at": datetime.now()
                    }
                    
                    async with db_manager.session() as session:
                        stmt = pg_insert(ReturnFBO).values(**return_dict)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["return_id"],
                            set_={k: v for k, v in return_dict.items() if k != "return_id"}
                        )
                        await session.execute(stmt)
                        records_processed += 1
            
            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"FBO returns sync completed: {records_processed} processed")
            return {"processed": records_processed}
            
        except Exception as e:
            logger.error(f"FBO returns sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== CASH FLOW SYNC ====================

    async def sync_cash_flow_statements(self, days_back: int = 365) -> Dict[str, int]:
        """Sinhronizacija /v1/finance/cash-flow-statement/list."""
        logger.info("Starting cash flow statements sync...")
        sync_log = await self._create_sync_log("cash_flow_statements")

        records_processed = 0
        to_date = datetime.now(timezone.utc)
        from_date = to_date - timedelta(days=days_back)

        try:
            rows_to_insert: List[Dict[str, Any]] = []
            async for items in self.client.get_all_cash_flow_statements(from_date=from_date, to_date=to_date, page_size=1000):
                for item in items:
                    period = item.get("period", {}) if isinstance(item.get("period"), dict) else {}
                    period_begin = self._parse_datetime(period.get("begin"))
                    rows_to_insert.append(
                        {
                            "date": period_begin or from_date,
                            "revenue": self._parse_decimal(item.get("orders_amount")),
                            "commission": self._parse_decimal(item.get("commission_amount")),
                            "delivery_cost": self._parse_decimal(item.get("item_delivery_and_return_amount")),
                            "return_cost": self._parse_decimal(item.get("returns_amount")),
                            "other_costs": self._parse_decimal(item.get("services_amount")),
                            "net_amount": self._parse_decimal(
                                (item.get("orders_amount") or 0)
                                + (item.get("returns_amount") or 0)
                                + (item.get("commission_amount") or 0)
                                + (item.get("services_amount") or 0)
                                + (item.get("item_delivery_and_return_amount") or 0)
                            ),
                            "raw_data": item,
                            "last_synced_at": datetime.now(),
                        }
                    )

            async with db_manager.session() as session:
                await session.execute(
                    delete(CashFlowStatement).where(
                        CashFlowStatement.date >= from_date,
                        CashFlowStatement.date <= to_date,
                    )
                )
                for row in rows_to_insert:
                    stmt = insert(CashFlowStatement).values(**row)
                    await session.execute(stmt)
                    records_processed += 1

            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Cash flow statements sync completed: {records_processed} processed")
            return {"processed": records_processed}

        except Exception as e:
            logger.error(f"Cash flow statements sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== CAMPAIGNS SYNC ====================
    
    async def sync_campaigns(self) -> Dict[str, Any]:
        """Sinhronizacija reklamnyh kampanij."""
        logger.info("Starting campaigns sync...")
        sync_log = await self._create_sync_log("campaigns")
        skip_result = await self._skip_recent_campaigns_sync_if_fresh(sync_log)
        if skip_result:
            return skip_result
        
        campaigns_processed = 0
        campaigns_inserted = 0
        campaigns_updated = 0
        campaign_objects_processed = 0
        campaign_details_updated = 0
        campaign_objects_skipped = 0
        campaign_objects_not_found = 0
        campaign_objects_errors = 0
        statistics_processed = 0
        statistics_batches = 0
        
        try:
            logger.info("campaigns sync: fetching campaign list from API...")
            campaigns_data = await self.client.get_campaigns()
            logger.info("campaigns sync: got campaign list, parsing rows...")
            campaign_rows = campaigns_data.get("list", []) if isinstance(campaigns_data, dict) else []
            logger.info("campaigns sync: found %d campaigns", len(campaign_rows))
            campaign_id_map: Dict[int, int] = {}

            for campaign_data in campaign_rows:
                external_campaign_id = int(campaign_data.get("id")) if campaign_data.get("id") else None
                if not external_campaign_id:
                    continue

                campaign_dict = {
                    "campaign_id": external_campaign_id,
                    "title": campaign_data.get("title"),
                    "state": campaign_data.get("state"),
                    "adv_object_type": campaign_data.get("advObjectType"),
                    "daily_budget": self._parse_decimal(campaign_data.get("dailyBudget")),
                    "total_budget": self._parse_decimal(campaign_data.get("totalBudget")),
                    "created_at": self._parse_datetime(campaign_data.get("createdAt")),
                    "started_at": self._parse_datetime(campaign_data.get("startedAt")),
                    "ended_at": self._parse_datetime(campaign_data.get("endedAt")),
                    "raw_data": campaign_data,
                    "last_synced_at": datetime.now()
                }
                
                async with db_manager.session() as session:
                    stmt = pg_insert(Campaign).values(**campaign_dict)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["campaign_id"],
                        set_={k: v for k, v in campaign_dict.items() if k != "campaign_id"}
                    ).returning(Campaign.id)
                    db_campaign_id = (await session.execute(stmt)).scalar_one()
                    campaign_id_map[external_campaign_id] = int(db_campaign_id)
                    campaigns_processed += 1
                    campaigns_updated += 1

            logger.info("campaigns sync: upserted %d campaigns to DB, fetching objects...", campaigns_processed)

            obj_idx = 0
            for campaign_data in campaign_rows:
                external_campaign_id = int(campaign_data.get("id")) if campaign_data.get("id") else None
                if not external_campaign_id or external_campaign_id not in campaign_id_map:
                    continue

                obj_idx += 1
                db_campaign_id = campaign_id_map[external_campaign_id]
                campaign_state = str(campaign_data.get("state") or "")
                objects_payload: Dict[str, Any] = {}
                object_rows: List[Dict[str, Any]] = []
                should_fetch_objects = campaign_state not in {
                    "CAMPAIGN_STATE_ARCHIVED",
                    "CAMPAIGN_STATE_FINISHED",
                }

                if not should_fetch_objects:
                    campaign_objects_skipped += 1
                    objects_payload = {
                        "skipped": True,
                        "reason": f"state={campaign_state or 'unknown'}",
                    }

                if should_fetch_objects:
                    logger.info(
                        "campaigns sync: fetching objects for campaign %s (%d/%d, state=%s)",
                        external_campaign_id, obj_idx, len(campaign_rows), campaign_state,
                    )
                    try:
                        objects_payload = await self._with_rate_limit_retry(
                            lambda ext_id=external_campaign_id: self.client.get_campaign_objects(ext_id),
                            attempts=4,
                            base_delay=3.0,
                        )
                        if isinstance(objects_payload, dict):
                            raw_list = objects_payload.get("list")
                            if isinstance(raw_list, list):
                                object_rows = [item for item in raw_list if isinstance(item, dict)]
                        elif isinstance(objects_payload, list):
                            object_rows = [item for item in objects_payload if isinstance(item, dict)]
                            objects_payload = {"list": object_rows}
                    except OzonAPIError as exc:
                        if exc.status_code == 404:
                            campaign_objects_not_found += 1
                        else:
                            campaign_objects_errors += 1
                        logger.warning(
                            "Campaign objects sync skipped for campaign %s: %s",
                            external_campaign_id,
                            exc,
                        )
                        objects_payload = {"error": str(exc)}
                        object_rows = []
                    except Exception as exc:
                        campaign_objects_errors += 1
                        logger.warning(
                            "Campaign objects sync skipped for campaign %s: %s",
                            external_campaign_id,
                            exc,
                        )
                        objects_payload = {"error": str(exc)}
                        object_rows = []

                logger.info("campaigns sync: saving detail+objects for campaign %s", external_campaign_id)
                detail_dict = {
                    "campaign_id": db_campaign_id,
                    "budget": self._parse_decimal_flexible(campaign_data.get("totalBudget")),
                    "daily_budget": self._parse_decimal_flexible(campaign_data.get("dailyBudget")),
                    "start_date": self._parse_datetime_flexible(campaign_data.get("startedAt")),
                    "end_date": self._parse_datetime_flexible(campaign_data.get("endedAt")),
                    "schedule": None,
                    "targeting": None,
                    "objects": objects_payload.get("list") if isinstance(objects_payload, dict) else object_rows,
                    "raw_data": {
                        "campaign": campaign_data,
                        "objects": objects_payload,
                    },
                    "last_synced_at": datetime.now(),
                }

                async with db_manager.session() as session:
                    await session.execute(
                        delete(CampaignDetail).where(CampaignDetail.campaign_id == db_campaign_id)
                    )
                    await session.execute(pg_insert(CampaignDetail).values(**detail_dict))
                    campaign_details_updated += 1

                    await session.execute(
                        delete(CampaignObject).where(CampaignObject.campaign_id == db_campaign_id)
                    )

                    for object_row in object_rows:
                        object_id = self._parse_int_flexible(
                            object_row.get("sku")
                            or object_row.get("id")
                            or object_row.get("objectId")
                        )
                        if object_id is None:
                            continue
                        object_dict = {
                            "campaign_id": db_campaign_id,
                            "sku": object_id,
                            "bid": self._parse_decimal_flexible(object_row.get("bid")),
                            "status": object_row.get("status"),
                            "views": self._parse_int_flexible(object_row.get("views")) or 0,
                            "clicks": self._parse_int_flexible(object_row.get("clicks")) or 0,
                            "spent": self._parse_decimal_flexible(
                                object_row.get("moneySpent") or object_row.get("spent")
                            ),
                            "orders": self._parse_int_flexible(object_row.get("orders")) or 0,
                            "revenue": self._parse_decimal_flexible(
                                object_row.get("ordersMoney") or object_row.get("revenue")
                            ),
                            "raw_data": object_row,
                            "last_synced_at": datetime.now(),
                        }
                        await session.execute(pg_insert(CampaignObject).values(**object_dict))
                        campaign_objects_processed += 1

            logger.info("campaigns sync: all details saved (%d), starting statistics...", campaign_details_updated)
            if campaign_id_map:
                date_to = datetime.now(timezone.utc)
                date_from = date_to - timedelta(days=max(int(settings.sync_days_back or 30), 30))
                external_ids = sorted(campaign_id_map.keys())
                batch_size = 10

                def _is_missing_campaign_report_error(exc: BaseException) -> bool:
                    if not isinstance(exc, OzonAPIError) or exc.status_code != 404:
                        return False
                    response_text = str((exc.response_data or {}).get("text") or "").lower()
                    return "report not found" in response_text

                def _is_campaign_report_unavailable_error(exc: BaseException) -> bool:
                    if isinstance(exc, RetryError) and exc.last_attempt:
                        exc = exc.last_attempt.exception()
                    if isinstance(exc, OzonAPIError):
                        if _is_missing_campaign_report_error(exc):
                            return True
                        message_text = str(exc).lower()
                        return "campaign report" in message_text and "not ready in time" in message_text
                    return False

                def _is_active_campaign_request_limit_error(exc: BaseException) -> bool:
                    if isinstance(exc, RetryError) and exc.last_attempt:
                        exc = exc.last_attempt.exception()
                    if not isinstance(exc, OzonAPIError):
                        return False
                    response_text = str((exc.response_data or {}).get("text") or "").lower()
                    message_text = str(exc).lower()
                    joined = f"{message_text} {response_text}"
                    return (
                        "максимум 1" in joined
                        or "maximum 1" in joined
                        or "active requests" in joined
                        or "лимит активных запросов" in joined
                        or "active request limit" in joined
                        or "( 1)" in joined
                    )

                async def _wait_campaign_report(
                    uuid: str,
                    expected_ids: List[int],
                    attempts: int = 24,
                    delay_seconds: float = 5.0,
                ) -> Dict[str, Any]:
                    last_error: Optional[Exception] = None
                    for _ in range(attempts):
                        await asyncio.sleep(delay_seconds)
                        try:
                            status_resp = await self.client.get_report_status(uuid)
                        except OzonAPIError as exc:
                            last_error = exc
                            if _is_missing_campaign_report_error(exc):
                                continue
                            raise
                        state = (status_resp.get("state") or "").upper() if isinstance(status_resp, dict) else ""
                        if state == "ERROR":
                            raise OzonAPIError(
                                f"Campaign report {uuid} failed: {status_resp.get('error', 'unknown')}"
                            )
                        if state == "OK":
                            report_payload = await self.client.download_campaign_report(uuid)
                            if isinstance(report_payload, dict) and any(
                                str(cid) in report_payload for cid in expected_ids
                            ):
                                return report_payload
                            return report_payload
                    if last_error:
                        if isinstance(last_error, RetryError) and last_error.last_attempt:
                            raise last_error.last_attempt.exception()
                        raise last_error
                    raise OzonAPIError(f"Campaign report {uuid} not ready in time")

                async def _request_campaign_report_with_retry(
                    batch_external_ids: List[int],
                    create_attempts: int = 1,
                ) -> Dict[str, Any]:
                    last_error: Optional[Exception] = None
                    for create_attempt in range(1, create_attempts + 1):
                        try:
                            uuid = await self._with_rate_limit_retry(
                                lambda: self.client.request_campaign_report(
                                    batch_external_ids,
                                    date_from=date_from,
                                    date_to=date_to,
                                    group_by="DATE",
                                ),
                                attempts=4,
                                base_delay=5.0,
                            )
                        except Exception as exc:
                            root_exc = exc.last_attempt.exception() if isinstance(exc, RetryError) and exc.last_attempt else exc
                            if _is_active_campaign_request_limit_error(root_exc):
                                raise root_exc
                            raise
                        if not uuid:
                            raise OzonAPIError(
                                f"Campaign report request returned empty UUID for batch {batch_external_ids}"
                            )
                        try:
                            return await self._with_rate_limit_retry(
                                lambda: _wait_campaign_report(uuid, expected_ids=batch_external_ids),
                                attempts=3,
                                base_delay=4.0,
                            )
                        except (OzonAPIError, RetryError) as exc:
                            last_error = exc
                            root_exc = exc.last_attempt.exception() if isinstance(exc, RetryError) and exc.last_attempt else exc
                            if create_attempt < create_attempts and _is_missing_campaign_report_error(root_exc):
                                logger.warning(
                                    "Campaign report %s for batch %s was not found on attempt %s/%s; requesting a new report",
                                    uuid,
                                    batch_external_ids,
                                    create_attempt,
                                    create_attempts,
                                )
                                continue
                            raise root_exc
                    if last_error:
                        raise last_error
                    raise OzonAPIError(
                        f"Campaign report could not be obtained for batch {batch_external_ids}"
                    )

                total_batches = (len(external_ids) + batch_size - 1) // batch_size
                for idx in range(0, len(external_ids), batch_size):
                    batch_external_ids = external_ids[idx:idx + batch_size]
                    batch_num = idx // batch_size + 1
                    logger.info(
                        "campaigns sync: requesting statistics report batch %d/%d (ids: %s)",
                        batch_num, total_batches, batch_external_ids[:3],
                    )
                    try:
                        report_payload = await _request_campaign_report_with_retry(batch_external_ids)
                    except Exception as exc:
                        root_exc = exc.last_attempt.exception() if isinstance(exc, RetryError) and exc.last_attempt else exc
                        if isinstance(root_exc, RateLimitError) or (
                            isinstance(root_exc, OzonAPIError) and root_exc.status_code == 429
                        ) or _is_active_campaign_request_limit_error(root_exc):
                            warning_message = (
                                "Campaign statistics refresh skipped due to Performance API rate limit; "
                                "campaign list was updated successfully"
                            )
                            logger.warning(
                                "%s. batch=%s error=%s",
                                warning_message,
                                batch_external_ids,
                                root_exc,
                            )
                            await self._update_sync_log(
                                sync_log,
                                "success",
                                records_processed=campaigns_processed + statistics_processed,
                                records_inserted=campaigns_inserted + statistics_processed,
                                records_updated=campaigns_updated,
                                error_message=warning_message,
                            )
                            return {
                                "processed": campaigns_processed + statistics_processed,
                                "campaigns_processed": campaigns_processed,
                                "campaigns_inserted": campaigns_inserted,
                                "campaigns_updated": campaigns_updated,
                                "statistics_processed": statistics_processed,
                                "statistics_batches": statistics_batches,
                                "statistics_skipped": 1,
                                "skip_reason": "rate_limit",
                            }
                        if _is_campaign_report_unavailable_error(root_exc):
                            warning_message = (
                                "Campaign statistics refresh skipped because Performance API report was not available "
                                "in time; campaign list was updated successfully"
                            )
                            logger.warning(
                                "%s. batch=%s error=%s",
                                warning_message,
                                batch_external_ids,
                                root_exc,
                            )
                            await self._update_sync_log(
                                sync_log,
                                "success",
                                records_processed=campaigns_processed + statistics_processed,
                                records_inserted=campaigns_inserted + statistics_processed,
                                records_updated=campaigns_updated,
                                error_message=warning_message,
                            )
                            return {
                                "processed": campaigns_processed + statistics_processed,
                                "campaigns_processed": campaigns_processed,
                                "campaigns_inserted": campaigns_inserted,
                                "campaigns_updated": campaigns_updated,
                                "statistics_processed": statistics_processed,
                                "statistics_batches": statistics_batches,
                                "statistics_skipped": 1,
                                "skip_reason": "report_unavailable",
                            }
                        raise root_exc
                    statistics_batches += 1
                    if idx > 0:
                        await asyncio.sleep(2.5)

                    async with db_manager.session() as session:
                        await session.execute(
                            delete(CampaignStatistic).where(
                                CampaignStatistic.campaign_id.in_([campaign_id_map[cid] for cid in batch_external_ids]),
                                CampaignStatistic.date >= date_from,
                                CampaignStatistic.date <= date_to,
                            )
                        )

                        for external_campaign_id in batch_external_ids:
                            campaign_block = report_payload.get(str(external_campaign_id)) or report_payload.get(external_campaign_id)
                            if not isinstance(campaign_block, dict):
                                continue
                            report_rows = ((campaign_block.get("report") or {}).get("rows") or [])
                            if not isinstance(report_rows, list):
                                continue

                            for row in report_rows:
                                if not isinstance(row, dict):
                                    continue
                                row_date = self._ensure_aware_utc(self._parse_datetime_flexible(row.get("date")))
                                sku = self._parse_int_flexible(row.get("sku"))
                                if row_date is None or sku is None:
                                    continue

                                views = self._parse_int_flexible(row.get("views")) or 0
                                clicks = self._parse_int_flexible(row.get("clicks")) or 0
                                adds_to_cart = self._parse_int_flexible(row.get("toCart") or row.get("hits_tocart")) or 0
                                spent = self._parse_decimal_flexible(row.get("moneySpent")) or 0.0
                                orders = self._parse_int_flexible(row.get("orders")) or 0
                                revenue = self._parse_decimal_flexible(row.get("ordersMoney")) or 0.0
                                ctr_pct = self._parse_decimal_flexible(row.get("ctr"))
                                acos_pct = self._parse_decimal_flexible(row.get("drr"))
                                avg_bid = self._parse_decimal_flexible(row.get("avgBid"))
                                cpc = (spent / clicks) if clicks > 0 else avg_bid
                                roas = (revenue / spent) if spent > 0 else None

                                stat_dict = {
                                    "campaign_id": campaign_id_map[external_campaign_id],
                                    "date": row_date,
                                    "views": views,
                                    "clicks": clicks,
                                    "adds_to_cart": adds_to_cart,
                                    "ctr": (ctr_pct / 100.0) if ctr_pct is not None else None,
                                    "spent": spent,
                                    "avg_bid": avg_bid,
                                    "cpc": cpc,
                                    "orders": orders,
                                    "revenue": revenue,
                                    "roas": roas,
                                    "acos": (acos_pct / 100.0) if acos_pct is not None else None,
                                    "sku": sku,
                                    "raw_data": row,
                                }
                                await session.execute(pg_insert(CampaignStatistic).values(**stat_dict))
                                statistics_processed += 1
            
            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=campaigns_processed + statistics_processed,
                records_inserted=campaigns_inserted + statistics_processed + campaign_objects_processed,
                records_updated=campaigns_updated + campaign_details_updated,
                error_message=(
                    f"objects_loaded={campaign_objects_processed}; "
                    f"objects_skipped={campaign_objects_skipped}; "
                    f"objects_not_found={campaign_objects_not_found}; "
                    f"objects_errors={campaign_objects_errors}"
                ),
            )
            
            logger.info(
                "Campaigns sync completed: campaigns=%s, statistics=%s, batches=%s",
                campaigns_processed,
                statistics_processed,
                statistics_batches,
            )
            return {
                "processed": campaigns_processed + statistics_processed + campaign_objects_processed,
                "campaigns_processed": campaigns_processed,
                "campaigns_inserted": campaigns_inserted,
                "campaigns_updated": campaigns_updated,
                "campaign_details_updated": campaign_details_updated,
                "campaign_objects_processed": campaign_objects_processed,
                "campaign_objects_skipped": campaign_objects_skipped,
                "campaign_objects_not_found": campaign_objects_not_found,
                "campaign_objects_errors": campaign_objects_errors,
                "statistics_processed": statistics_processed,
                "statistics_batches": statistics_batches,
            }
            
        except Exception as e:
            logger.error(f"Campaigns sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== REVIEWS SYNC ====================
    
    async def sync_reviews(self) -> Dict[str, int]:
        """Sinhronizacija otzyvov."""
        logger.info("Starting reviews sync...")
        sync_log = await self._create_sync_log("reviews")
        
        records_processed = 0
        
        try:
            async for reviews in self.client.get_all_reviews():
                for review_data in reviews:
                    review_dict = {
                        "review_id": review_data.get("id"),
                        "sku": review_data.get("sku"),
                        "offer_id": review_data.get("offer_id"),
                        "rating": review_data.get("rating"),
                        "text": review_data.get("text"),
                        "status": review_data.get("status"),
                        "is_buyer": review_data.get("is_buyer", False),
                        "published_at": self._parse_datetime(review_data.get("published_at")),
                        "created_at": self._parse_datetime(review_data.get("created_at")),
                        "helpful_count": review_data.get("helpful_count", 0),
                        "unhelpful_count": review_data.get("unhelpful_count", 0),
                        "raw_data": review_data,
                        "last_synced_at": datetime.now()
                    }
                    
                    async with db_manager.session() as session:
                        stmt = pg_insert(Review).values(**review_dict)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["review_id"],
                            set_={k: v for k, v in review_dict.items() if k != "review_id"}
                        )
                        await session.execute(stmt)
                        records_processed += 1
            
            await self._snapshot_review_ratings()

            await self._update_sync_log(sync_log, "success", records_processed)
            logger.info(f"Reviews sync completed: {records_processed} processed")
            return {"processed": records_processed}

        except Exception as e:
            logger.error(f"Reviews sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def _snapshot_review_ratings(self) -> None:
        """Sohranit segodnjasnij snimok srednej ocenki po kazhdomu sku."""
        from sqlalchemy import select as sa_select, func as sa_func
        today = datetime.now().date()
        async with db_manager.session() as session:
            rows = await session.execute(
                sa_select(
                    Review.sku,
                    sa_func.avg(Review.rating).label("avg_rating"),
                    sa_func.count().label("cnt"),
                ).where(Review.sku.isnot(None)).group_by(Review.sku)
            )
            for sku, avg_rating, cnt in rows.all():
                stmt = pg_insert(ReviewRatingSnapshot).values(
                    sku=sku,
                    snapshot_date=today,
                    avg_rating=avg_rating,
                    reviews_count=cnt,
                )
                stmt = stmt.on_conflict_do_update(
                    index_elements=["sku", "snapshot_date"],
                    set_={"avg_rating": avg_rating, "reviews_count": cnt},
                )
                await session.execute(stmt)
            await session.commit()

    # ==================== SELLER RATING SYNC ====================
    
    async def sync_seller_rating(self) -> Dict[str, int]:
        """Sinhronizacija reitinga prodavca."""
        logger.info("Starting seller rating sync...")
        sync_log = await self._create_sync_log("seller_rating")
        
        try:
            rating_data = await self.client.get_rating_summary()
            
            rating_dict = {
                "date": datetime.now().date(),
                "overall_rating": self._parse_decimal(rating_data.get("rating")),
                "position_in_category": rating_data.get("position_in_category"),
                "price_quality_rating": self._parse_decimal(rating_data.get("price_quality_rating")),
                "delivery_rating": self._parse_decimal(rating_data.get("delivery_rating")),
                "service_rating": self._parse_decimal(rating_data.get("service_rating")),
                "cancellation_rate": self._parse_decimal(rating_data.get("cancellation_rate")),
                "late_shipment_rate": self._parse_decimal(rating_data.get("late_shipment_rate")),
                "return_rate": self._parse_decimal(rating_data.get("return_rate")),
                "raw_data": rating_data,
                "last_synced_at": datetime.now()
            }
            
            async with db_manager.session() as session:
                stmt = pg_insert(SellerRating).values(**rating_dict)
                stmt = stmt.on_conflict_do_update(
                    index_elements=["date"],
                    set_={k: v for k, v in rating_dict.items() if k != "date"}
                )
                await session.execute(stmt)
            
            await self._update_sync_log(sync_log, "success", 1)
            logger.info("Seller rating sync completed")
            return {"processed": 1}
            
        except Exception as e:
            logger.error(f"Seller rating sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== PROMO SYNC ====================

    async def _ensure_promo_schema(self) -> None:
        """Dobavljaet novye kolonki v promo_actions/promo_products, esli ih net."""
        if "postgresql" not in settings.database_url.lower():
            return
        alter_sql = [
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS discount_type VARCHAR(50);",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS discount_value NUMERIC(15, 2);",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS potential_products_count INTEGER;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS participating_products_count INTEGER;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS banned_products_count INTEGER;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS description TEXT;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS with_targeting BOOLEAN DEFAULT false;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS is_voucher_action BOOLEAN DEFAULT false;",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS order_amount NUMERIC(15, 2);",
            "ALTER TABLE promo_actions ADD COLUMN IF NOT EXISTS freeze_date VARCHAR(50);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS max_action_price NUMERIC(15, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS add_mode VARCHAR(50);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS stock INTEGER;",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS min_stock INTEGER;",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS current_boost NUMERIC(10, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS min_boost NUMERIC(10, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS max_boost NUMERIC(10, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS price_min_elastic NUMERIC(15, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS price_max_elastic NUMERIC(15, 2);",
            "ALTER TABLE promo_products ADD COLUMN IF NOT EXISTS first_seen_at TIMESTAMPTZ DEFAULT now();",
            # Таблица событий добавления/удаления товаров в акциях
            """CREATE TABLE IF NOT EXISTS promo_product_events (
                id SERIAL PRIMARY KEY,
                action_id INTEGER NOT NULL,
                sku BIGINT NOT NULL,
                event_type VARCHAR(20) NOT NULL,
                source VARCHAR(20) NOT NULL DEFAULT 'sync',
                detected_at TIMESTAMPTZ DEFAULT now()
            );""",
            "CREATE INDEX IF NOT EXISTS ix_promo_events_action_sku ON promo_product_events (action_id, sku);",
        ]
        async with db_manager.session() as session:
            for stmt in alter_sql:
                await session.execute(text(stmt))

    def _promo_product_row(self, action_db_id: int, product: Dict[str, Any], is_candidate: bool) -> Optional[PromoProduct]:
        sku = product.get("id")
        try:
            sku = int(float(sku)) if sku is not None else None
        except (TypeError, ValueError):
            sku = None
        if sku is None:
            return None

        regular_price = self._parse_decimal(product.get("price"))
        action_price = self._parse_decimal(product.get("action_price"))
        discount_percent = None
        if regular_price and action_price is not None and regular_price > 0 and action_price > 0:
            discount_percent = ((regular_price - action_price) / regular_price) * 100

        def _int(v):
            try:
                return int(float(v)) if v is not None else None
            except (TypeError, ValueError):
                return None

        return PromoProduct(
            action_id=action_db_id,
            sku=sku,
            offer_id=product.get("offer_id"),
            regular_price=regular_price,
            action_price=action_price,
            discount_percent=discount_percent,
            is_participating=not is_candidate,
            is_candidate=is_candidate,
            orders_count=0,
            revenue=None,
            max_action_price=self._parse_decimal(product.get("max_action_price")),
            add_mode=product.get("add_mode"),
            stock=_int(product.get("stock")),
            min_stock=_int(product.get("min_stock")),
            current_boost=self._parse_decimal(product.get("current_boost")),
            min_boost=self._parse_decimal(product.get("min_boost")),
            max_boost=self._parse_decimal(product.get("max_boost")),
            price_min_elastic=self._parse_decimal(product.get("price_min_elastic")),
            price_max_elastic=self._parse_decimal(product.get("price_max_elastic")),
            raw_data=product,
            last_synced_at=datetime.now(),
        )

    async def sync_promo(self) -> Dict[str, int]:
        """Sinhronizacija akcij Ozon, tovarov v akcijah i tovarov-kandidatov."""
        logger.info("Starting promo sync...")
        sync_log = await self._create_sync_log("promo")

        await self._ensure_promo_schema()

        actions_processed = 0
        products_processed = 0
        candidates_processed = 0

        try:
            response = await self.client.get_actions_list()
            actions = response.get("result", []) if isinstance(response, dict) else []

            # Собираем action_id которые вернул API
            api_action_ids = set()
            for a in actions:
                try:
                    api_action_ids.add(int(float(a["id"])))
                except (TypeError, ValueError, KeyError):
                    pass

            # Сбрасываем is_participating для акций, которых нет в текущем ответе API
            if api_action_ids:
                async with db_manager.session() as session:
                    await session.execute(
                        update(PromoAction)
                        .where(PromoAction.is_participating == True)
                        .where(PromoAction.action_id.notin_(api_action_ids))
                        .values(is_participating=False, status="ENDED", last_synced_at=datetime.now())
                    )
            else:
                # API вернул пустой список — сбрасываем все
                async with db_manager.session() as session:
                    await session.execute(
                        update(PromoAction)
                        .where(PromoAction.is_participating == True)
                        .values(is_participating=False, status="ENDED", last_synced_at=datetime.now())
                    )

            for action in actions:
                external_action_id = action.get("id")
                if external_action_id is None:
                    continue

                try:
                    action_id_int = int(float(external_action_id))
                except (TypeError, ValueError):
                    logger.warning(f"Skipping promo action with invalid id: {external_action_id}")
                    continue

                promo_action_data = {
                    "action_id": action_id_int,
                    "title": action.get("title"),
                    "action_type": action.get("action_type"),
                    "date_start": self._parse_datetime(action.get("date_start")),
                    "date_end": self._parse_datetime(action.get("date_end")),
                    "status": "ACTIVE" if action.get("is_participating") else "AVAILABLE",
                    "is_participating": bool(action.get("is_participating", False)),
                    "discount_percent": self._parse_decimal(action.get("discount_value")),
                    "max_quantity": action.get("potential_products_count"),
                    "discount_type": action.get("discount_type"),
                    "discount_value": self._parse_decimal(action.get("discount_value")),
                    "potential_products_count": action.get("potential_products_count"),
                    "participating_products_count": action.get("participating_products_count"),
                    "banned_products_count": action.get("banned_products_count"),
                    "description": action.get("description"),
                    "with_targeting": bool(action.get("with_targeting", False)),
                    "is_voucher_action": bool(action.get("is_voucher_action", False)),
                    "order_amount": self._parse_decimal(action.get("order_amount")),
                    "freeze_date": action.get("freeze_date") or None,
                    "raw_data": action,
                    "last_synced_at": datetime.now(),
                }

                # Tovary, uchastvujushhie v akcii
                all_products: List[Dict[str, Any]] = []
                products_fetch_ok = False
                try:
                    async for products_page in self.client.get_all_action_products(action_id=action_id_int, limit=100):
                        all_products.extend(products_page)
                    products_fetch_ok = True
                except Exception as e:
                    logger.warning(f"Failed to fetch action products for action {action_id_int}: {e}")

                # Tovary-kandidaty (mogut uchastvovat, no ne dobavleny)
                all_candidates: List[Dict[str, Any]] = []
                candidates_fetch_ok = False
                try:
                    offset = 0
                    while True:
                        rc = await self.client.get_action_candidates(action_id=action_id_int, limit=100, offset=offset)
                        result = rc.get("result") if isinstance(rc, dict) else None
                        if isinstance(result, dict):
                            page = result.get("products", []) or []
                        elif isinstance(result, list):
                            page = result
                        else:
                            page = []
                        if not page:
                            break
                        all_candidates.extend(page)
                        if len(page) < 100:
                            break
                        offset += 100
                        if offset > 10000:
                            break
                    candidates_fetch_ok = True
                except Exception as e:
                    logger.warning(f"Failed to fetch action candidates for action {action_id_int}: {e}")

                async with db_manager.session() as session:
                    stmt_action = pg_insert(PromoAction).values(**promo_action_data)
                    stmt_action = stmt_action.on_conflict_do_update(
                        index_elements=["action_id"],
                        set_={k: v for k, v in promo_action_data.items() if k != "action_id"},
                    ).returning(PromoAction.id)
                    action_db_id = (await session.execute(stmt_action)).scalar_one()

                    # Diff-подход: сохраняем first_seen_at для существующих товаров
                    if products_fetch_ok and candidates_fetch_ok:
                        # Получаем текущие SKU и их first_seen_at + статус
                        existing_rows = await session.execute(
                            select(PromoProduct.sku, PromoProduct.first_seen_at, PromoProduct.is_participating)
                            .where(PromoProduct.action_id == action_db_id)
                        )
                        existing_first_seen = {}
                        old_participating_skus = set()
                        for r in existing_rows:
                            existing_first_seen[int(r.sku)] = r.first_seen_at
                            if r.is_participating:
                                old_participating_skus.add(int(r.sku))

                        # Удаляем старые данные
                        await session.execute(
                            delete(PromoProduct).where(PromoProduct.action_id == action_db_id)
                        )

                        rows: List[PromoProduct] = []
                        seen_skus = set()
                        action_participating = 0
                        action_candidates = 0
                        now_ts = datetime.now()
                        for product in all_products:
                            row = self._promo_product_row(action_db_id, product, is_candidate=False)
                            if row is not None and row.sku not in seen_skus:
                                # Сохраняем first_seen_at если товар уже был в акции
                                row.first_seen_at = existing_first_seen.get(row.sku, now_ts)
                                rows.append(row)
                                seen_skus.add(row.sku)
                                action_participating += 1
                        for product in all_candidates:
                            row = self._promo_product_row(action_db_id, product, is_candidate=True)
                            if row is not None and row.sku not in seen_skus:
                                row.first_seen_at = existing_first_seen.get(row.sku, now_ts)
                                rows.append(row)
                                seen_skus.add(row.sku)
                                action_candidates += 1

                        if rows:
                            session.add_all(rows)

                        # Трекинг событий: сравниваем старый и новый набор участников
                        new_participating_skus = set()
                        for row in rows:
                            if row.is_participating:
                                new_participating_skus.add(int(row.sku))
                        old_participating = old_participating_skus
                        new_participating = new_participating_skus
                        added_skus = new_participating - old_participating
                        removed_skus = old_participating - new_participating
                        if added_skus or removed_skus:
                            from src.models import PromoProductEvent
                            for sku in added_skus:
                                session.add(PromoProductEvent(
                                    action_id=action_id_int,
                                    sku=sku,
                                    event_type="ADDED",
                                    source="sync",
                                    detected_at=now_ts,
                                ))
                            for sku in removed_skus:
                                session.add(PromoProductEvent(
                                    action_id=action_id_int,
                                    sku=sku,
                                    event_type="REMOVED",
                                    source="sync",
                                    detected_at=now_ts,
                                ))
                            logger.info(
                                f"Action {action_id_int}: {len(added_skus)} added, "
                                f"{len(removed_skus)} removed"
                            )

                        products_processed += action_participating
                        candidates_processed += action_candidates
                    else:
                        logger.warning(
                            f"Skipping product update for action {action_id_int} — "
                            f"products_ok={products_fetch_ok}, candidates_ok={candidates_fetch_ok}. "
                            f"Keeping existing data."
                        )

                actions_processed += 1

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=actions_processed + products_processed,
                records_inserted=products_processed,
                records_updated=actions_processed,
            )
            logger.info(
                f"Promo sync completed: actions={actions_processed}, products={products_processed}, candidates={candidates_processed}"
            )
            return {
                "actions_processed": actions_processed,
                "products_processed": products_processed,
                "candidates_processed": candidates_processed,
            }

        except Exception as e:
            logger.error(f"Promo sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== REPORT POSTINGS SYNC ====================

    async def sync_postings_report(self, pages: int = 10, page_size: int = 100) -> Dict[str, Any]:
        """Skachivanie seller_postings otchetov i zagruzka v async_reports/fact_orders."""
        logger.info("Starting report_postings sync...")
        sync_log = await self._create_sync_log("report_postings")
        skip_result = await self._skip_recent_async_report_sync_if_fresh(sync_log, "report_postings")
        if skip_result:
            return skip_result
        await self._ensure_fact_orders_cluster_schema()

        reports_processed = 0
        orders_upserted = 0
        unknown_schema_count = 0
        unknown_schema_examples: List[str] = []

        try:
            # Regular postings sync refreshes only a recent rolling window.
            now_utc = datetime.now(timezone.utc)
            report_days_back = max(int(settings.report_postings_days_back or 60), 1)
            report_from = now_utc - timedelta(days=report_days_back)
            max_chunk_days = 90
            date_chunks: List[Tuple[datetime, datetime]] = []
            chunk_start = report_from
            while chunk_start < now_utc:
                chunk_end = min(chunk_start + timedelta(days=max_chunk_days), now_utc)
                date_chunks.append((chunk_start, chunk_end))
                chunk_start = chunk_end
            created_codes: Dict[str, str] = {}
            for schema in ("fbo", "fbs"):
                for chunk_from, chunk_to in date_chunks:
                    try:
                        created = await self.client.create_report_postings(
                            date_from=chunk_from,
                            date_to=chunk_to,
                            posting_type=schema,
                        )
                        code = created.get("result", {}).get("code") or created.get("code")
                        if code:
                            created_codes[code] = schema.upper()
                            logger.info(
                                "Created seller_postings report for %s (%s..%s): %s",
                                schema,
                                chunk_from.date(),
                                chunk_to.date(),
                                code,
                            )
                    except Exception as exc:
                        logger.warning(
                            "Failed to create seller_postings report for %s (%s..%s): %s",
                            schema,
                            chunk_from.date(),
                            chunk_to.date(),
                            exc,
                        )

            # Wait a bit for newly created reports to become downloadable.
            for code in created_codes.keys():
                for _ in range(20):
                    info = await self.client.get_report_info(code)
                    status = str(info.get("result", {}).get("status", "")).lower()
                    if status in {"success", "error"}:
                        break
                    await asyncio.sleep(2)

            for page in range(1, pages + 1):
                reports_page = await self.client.get_reports_list(page=page, page_size=page_size)
                reports = reports_page.get("result", {}).get("reports", [])
                if not reports:
                    break

                for report in reports:
                    report_type = str(report.get("report_type", "")).lower()
                    report_code = str(report.get("code", "")).strip()
                    status = str(report.get("status", "")).lower()

                    if not report_code:
                        continue
                    if report_type not in {"seller_postings", "postings"} and "posting" not in report_code.lower():
                        continue
                    # Prefer processing reports created in current run to keep schema deterministic.
                    if created_codes and report_code not in created_codes:
                        continue

                    info = await self.client.get_report_info(report_code)
                    info_result = info.get("result", {}) if isinstance(info, dict) else {}
                    file_url = info_result.get("file") or report.get("file")
                    final_status = (info_result.get("status") or status or "").lower()

                    if not file_url or final_status != "success":
                        continue

                    try:
                        file_bytes = await self.client.download_file(file_url)
                    except Exception as exc:
                        logger.warning(f"Skipping report {report_code}: file download failed: {exc}")
                        continue
                    # Fresh seller_postings reports are currently exported as UTF-8 with BOM.
                    text = file_bytes.decode("utf-8-sig", errors="replace")
                    reader = csv.DictReader(io.StringIO(text), delimiter=";")
                    rows = list(reader)

                    report_id = self._report_code_to_id(report_code)
                    created_at = self._parse_datetime_flexible(info_result.get("created_at") or report.get("created_at"))
                    completed_at = self._parse_datetime_flexible(info_result.get("updated_at") or info_result.get("created_at"))

                    async with db_manager.session() as session:
                        report_data = {
                            "report_id": report_id,
                            "report_type": report_type or "seller_postings",
                            "status": final_status,
                            "date_from": None,
                            "date_to": None,
                            "filters": info_result.get("params") or report.get("params"),
                            "file_url": file_url,
                            "file_size": len(file_bytes),
                            "row_count": len(rows),
                            "created_at": created_at,
                            "completed_at": completed_at,
                            "raw_data": info_result or report,
                            "last_synced_at": datetime.now(),
                        }
                        stmt_report = pg_insert(AsyncReport).values(**report_data)
                        stmt_report = stmt_report.on_conflict_do_update(
                            index_elements=["report_id"],
                            set_={k: v for k, v in report_data.items() if k != "report_id"},
                        )
                        await session.execute(stmt_report)

                        grouped: Dict[str, List[Dict[str, Any]]] = {}
                        for row in rows:
                            posting_number = (row.get("Номер отправления") or "").strip()
                            if not posting_number:
                                continue
                            grouped.setdefault(posting_number, []).append(row)

                        for posting_number, group_rows in grouped.items():
                            first = group_rows[0]
                            order_id = str(first.get("Номер заказа") or posting_number)
                            items_total = self._parse_decimal_flexible(first.get("Сумма отправления"))
                            discount_total = sum(
                                (self._parse_decimal_flexible(r.get("Скидка руб")) or 0.0) for r in group_rows
                            )
                            delivery_cost = sum(
                                (self._parse_decimal_flexible(r.get("Стоимость доставки")) or 0.0) for r in group_rows
                            )
                            items = []
                            for r in group_rows:
                                items.append(
                                    {
                                        "name": r.get("Название товара"),
                                        "sku": r.get("SKU"),
                                        "offer_id": r.get("Артикул"),
                                        "quantity": self._parse_decimal_flexible(r.get("Количество")),
                                        "price": self._parse_decimal_flexible(r.get("Ваша цена")),
                                        "buyer_paid": self._parse_decimal_flexible(r.get("Оплачено покупателем")),
                                    }
                                )

                            report_schema = created_codes.get(report_code)
                            if not report_schema:
                                params = info_result.get("params") or report.get("params") or {}
                                if isinstance(params, dict):
                                    ds = params.get("delivery_schema")
                                    if isinstance(ds, list) and ds:
                                        report_schema = str(ds[0]).upper()
                                    elif isinstance(ds, str) and ds:
                                        report_schema = ds.upper()
                            if report_schema not in {"FBO", "FBS"}:
                                report_schema = "UNKNOWN"
                                unknown_schema_count += 1
                                if len(unknown_schema_examples) < 10:
                                    unknown_schema_examples.append(
                                        f"{report_code}:{order_id}:{posting_number}"
                                    )

                            fact_data = {
                                "order_id": order_id,
                                "posting_number": posting_number,
                                "delivery_schema": report_schema,
                                "status": first.get("Статус"),
                                "substatus": None,
                                "created_at": self._parse_datetime_flexible(first.get("Принят в обработку")),
                                "in_process_at": self._parse_datetime_flexible(first.get("Принят в обработку")),
                                "shipment_date": self._parse_datetime_flexible(
                                    first.get("Фактическая дата передачи в доставку")
                                    or first.get("Дата отгрузки")
                                ),
                                "delivered_at": self._parse_datetime_flexible(first.get("Дата доставки")),
                                "cancelled_at": None,
                                "items_total": items_total,
                                "discount_total": discount_total,
                                "delivery_cost": delivery_cost,
                                "commission_total": None,
                                "payout_total": None,
                                "is_on_time": None,
                                "sla_hours": None,
                                "is_returned": None,
                                "return_amount": None,
                                "items": items,
                                "customer_name": None,
                                "region": None,
                                "city": None,
                                "delivery_cluster_from": (
                                    first.get("Кластер отгрузки")
                                    or first.get("Кластер отправки")
                                    or first.get("Shipment cluster")
                                    or first.get("cluster_from")
                                ),
                                "delivery_cluster_to": (
                                    first.get("Кластер назначения")
                                    or first.get("Кластер доставки")
                                    or first.get("Delivery cluster")
                                    or first.get("cluster_to")
                                ),
                                "shipping_warehouse_name": (
                                    first.get("Склад")
                                    or first.get("Склад отгрузки")
                                    or first.get("warehouse")
                                ),
                                "raw_data": {"report_code": report_code, "rows": group_rows},
                                "last_synced_at": datetime.now(),
                            }

                            stmt_fact = pg_insert(FactOrder).values(**fact_data)
                            stmt_fact = stmt_fact.on_conflict_do_update(
                                index_elements=["order_id"],
                                set_={k: v for k, v in fact_data.items() if k != "order_id"},
                            )
                            await session.execute(stmt_fact)
                            await self._replace_fact_order_items(
                                session=session,
                                order_id=order_id,
                                posting_number=posting_number,
                                items=items,
                            )
                            orders_upserted += 1

                    reports_processed += 1

            cluster_sync_result: Dict[str, Any] = {}
            try:
                cluster_sync_result = await self.sync_order_delivery_clusters(days_back=max(settings.sync_days_back, 35))
            except Exception as cluster_exc:
                logger.warning(f"order delivery clusters enrichment skipped: {cluster_exc}")

            if unknown_schema_count > 0:
                error_message = (
                    f"UNKNOWN delivery_schema detected: {unknown_schema_count}. "
                    f"examples={unknown_schema_examples}"
                )
                logger.error(error_message)
                await self._update_sync_log(
                    sync_log,
                    "error",
                    records_processed=orders_upserted,
                    records_inserted=orders_upserted,
                    records_updated=0,
                    error_message=error_message,
                )
            else:
                await self._update_sync_log(
                    sync_log,
                    "success",
                    records_processed=orders_upserted,
                    records_inserted=orders_upserted,
                    records_updated=0,
                )
            logger.info(
                f"report_postings sync completed: reports={reports_processed}, "
                f"orders_upserted={orders_upserted}, unknown_schema_count={unknown_schema_count}"
            )
            return {
                "reports_processed": reports_processed,
                "orders_upserted": orders_upserted,
                "unknown_schema_count": unknown_schema_count,
                "cluster_sync": cluster_sync_result,
            }
        except Exception as e:
            logger.error(f"report_postings sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== REPORT PRODUCTS SYNC ====================
    async def _backfill_report_products_offer_ids(self, report_ids: List[int]) -> int:
        """Backfill pustyh offer_id v report_products_items po SKU cherez /v3/product/info/list."""
        if not report_ids:
            return 0

        target_rows: List[Tuple[int, int]] = []
        async with db_manager.session() as session:
            result = await session.execute(
                select(
                    ReportProductItem.id,
                    ReportProductItem.fbo_sku_id,
                    ReportProductItem.fbs_sku_id,
                ).where(
                    ReportProductItem.report_id.in_(report_ids),
                    (ReportProductItem.offer_id.is_(None) | (text("btrim(coalesce(offer_id, '')) = ''"))),
                    (ReportProductItem.fbo_sku_id.is_not(None) | ReportProductItem.fbs_sku_id.is_not(None)),
                )
            )
            for row in result.fetchall():
                row_id = int(row[0])
                sku_val = self._parse_int_flexible(row[1]) or self._parse_int_flexible(row[2])
                if row_id and sku_val:
                    target_rows.append((row_id, int(sku_val)))

        if not target_rows:
            return 0

        sku_set = sorted({sku for _, sku in target_rows})
        sku_to_offer: Dict[int, str] = {}
        chunk_size = 100
        for i in range(0, len(sku_set), chunk_size):
            chunk = sku_set[i : i + chunk_size]
            try:
                resp = await self.client.get_product_info_list_v3(chunk)
                payload_items = resp.get("items") or (resp.get("result") or {}).get("items") or []
                for item in payload_items:
                    sku_val = self._parse_int_flexible(item.get("sku"))
                    offer_id = str(item.get("offer_id") or "").strip()
                    if sku_val and offer_id:
                        sku_to_offer[int(sku_val)] = offer_id
            except Exception as e:
                logger.warning(f"report_products backfill: failed /v3/product/info/list chunk {i // chunk_size + 1}: {e}")

        if not sku_to_offer:
            return 0

        updated = 0
        now_ts = datetime.now()
        async with db_manager.session() as session:
            for row_id, sku_val in target_rows:
                offer_id = sku_to_offer.get(sku_val)
                if not offer_id:
                    continue
                await session.execute(
                    update(ReportProductItem)
                    .where(ReportProductItem.id == row_id)
                    .values(offer_id=offer_id, last_synced_at=now_ts)
                )
                updated += 1

        return updated

    async def sync_products_report(self, pages: int = 10, page_size: int = 100) -> Dict[str, Any]:
        """Skachivanie seller_products otcheta i zagruzka strok v report_products_items."""
        logger.info("Starting report_products sync...")
        sync_log = await self._create_sync_log("report_products")
        skip_result = await self._skip_recent_async_report_sync_if_fresh(sync_log, "report_products")
        if skip_result:
            recent_report_ids: List[int] = []
            async with db_manager.session() as session:
                recent_ids_res = await session.execute(
                    text(
                        """
                        SELECT DISTINCT report_id
                        FROM report_products_items
                        ORDER BY report_id DESC
                        LIMIT 5
                        """
                    )
                )
                recent_report_ids = [int(r[0]) for r in recent_ids_res.fetchall() if r and r[0] is not None]
            offer_backfilled = await self._backfill_report_products_offer_ids(recent_report_ids)
            return {**skip_result, "offer_backfilled": offer_backfilled}

        reports_processed = 0
        rows_upserted = 0
        processed_report_ids: List[int] = []

        try:
            created = await self.client.create_report_products(language="DEFAULT", visibility="ALL")
            created_code = (created.get("result") or {}).get("code") or created.get("code")
            created_codes = {created_code} if created_code else set()
            if created_code:
                logger.info(f"Created seller_products report: {created_code}")
                for _ in range(30):
                    info = await self.client.get_report_info(created_code)
                    status = str((info.get("result") or {}).get("status", "")).lower()
                    if status in {"success", "error"}:
                        break
                    await asyncio.sleep(2)

            for page in range(1, pages + 1):
                reports_page = await self.client.get_reports_list(page=page, page_size=page_size)
                reports = (reports_page.get("result") or {}).get("reports", [])
                if not reports:
                    break

                for report in reports:
                    report_code = str(report.get("code", "")).strip()
                    if not report_code:
                        continue
                    if created_codes and report_code not in created_codes:
                        continue

                    report_type = str(report.get("report_type", "")).lower()
                    is_product_report = report_type in {"products", "seller_products"} or "products" in report_code.lower()
                    if not is_product_report:
                        continue

                    info = await self.client.get_report_info(report_code)
                    info_result = info.get("result", {}) if isinstance(info, dict) else {}
                    final_status = str(info_result.get("status") or report.get("status") or "").lower()
                    file_url = info_result.get("file") or report.get("file")
                    if final_status != "success" or not file_url:
                        continue

                    file_bytes = await self.client.download_file(file_url)
                    try:
                        csv_text = file_bytes.decode("utf-8-sig")
                    except UnicodeDecodeError:
                        csv_text = file_bytes.decode("cp1251", errors="replace")
                    rows: List[Dict[str, Any]] = []
                    parsed = False
                    for delim in (";", ","):
                        try:
                            reader = csv.DictReader(io.StringIO(csv_text, newline=""), delimiter=delim)
                            rows = list(reader)
                            parsed = True
                            break
                        except csv.Error:
                            continue
                    if not parsed:
                        raise ValueError("Unable to parse products report CSV")

                    report_id = self._report_code_to_id(report_code)
                    processed_report_ids.append(report_id)
                    created_at = self._parse_datetime_flexible(info_result.get("created_at") or report.get("created_at"))
                    completed_at = self._parse_datetime_flexible(info_result.get("updated_at") or info_result.get("created_at"))

                    async with db_manager.session() as session:
                        report_data = {
                            "report_id": report_id,
                            "report_type": report_type or "seller_products",
                            "status": final_status,
                            "date_from": None,
                            "date_to": None,
                            "filters": info_result.get("params") or report.get("params"),
                            "file_url": file_url,
                            "file_size": len(file_bytes),
                            "row_count": len(rows),
                            "created_at": created_at,
                            "completed_at": completed_at,
                            "raw_data": info_result or report,
                            "last_synced_at": datetime.now(),
                        }
                        stmt_report = pg_insert(AsyncReport).values(**report_data)
                        stmt_report = stmt_report.on_conflict_do_update(
                            index_elements=["report_id"],
                            set_={k: v for k, v in report_data.items() if k != "report_id"},
                        )
                        await session.execute(stmt_report)

                        await session.execute(
                            delete(ReportProductItem).where(ReportProductItem.report_id == report_id)
                        )

                        for i, row in enumerate(rows, start=1):
                            row_data = {
                                "report_id": report_id,
                                "line_no": i,
                                "offer_id": self._pick_value(row, ["Offer ID", "РђСЂС‚РёРєСѓР»", "РђСЂС‚РёРєСѓР» С‚РѕРІР°СЂР°", "Р С’РЎР‚РЎвЂљР С‘Р С”РЎС“Р В»"]),
                                "product_name": self._pick_value(row, ["Name", "РќР°Р·РІР°РЅРёРµ С‚РѕРІР°СЂР°", "РќР°РёРјРµРЅРѕРІР°РЅРёРµ С‚РѕРІР°СЂР°", "Р СњР В°Р В·Р Р†Р В°Р Р…Р С‘Р Вµ РЎвЂљР С•Р Р†Р В°РЎР‚Р В°"]),
                                "ozon_product_id": self._parse_int_flexible(
                                    self._pick_value(row, ["Ozon Product ID", "Ozon product id", "Ozon ID", "Ozon Product Id"])
                                ),
                                "fbo_sku_id": self._parse_int_flexible(
                                    self._pick_value(row, ["FBO Ozon SKU ID", "SKU FBO", "FBO SKU ID"])
                                    or self._pick_value(row, ["SKU"])
                                ),
                                "fbs_sku_id": self._parse_int_flexible(
                                    self._pick_value(row, ["FBS Ozon SKU ID", "SKU FBS", "FBS SKU ID"])
                                    or self._pick_value(row, ["SKU"])
                                ),
                                "crossborder_sku": self._pick_value(row, ["CrossBorder Ozon SKU", "CrossBorder SKU"]),
                                "barcode": self._pick_value(row, ["Barcode", "РЁС‚СЂРёС…РєРѕРґ", "Barcode С‚РѕРІР°СЂР°"]),
                                "product_status": self._pick_value(row, ["РЎС‚Р°С‚СѓСЃ С‚РѕРІР°СЂР°", "Product status", "Status"]),
                                "stock_fbo_available": self._parse_int_flexible(
                                    self._pick_value(row, ["Р”РѕСЃС‚СѓРїРЅРѕ РЅР° СЃРєР»Р°РґРµ Ozon, С€С‚", "Available in Ozon warehouse, pcs"])
                                    or self._pick_by_contains(row, ["РґРѕСЃС‚СѓРїРЅРѕ Рє РїСЂРѕРґР°Р¶Рµ РїРѕ СЃС…РµРјРµ fbo"])
                                ),
                                "stock_reserved": self._parse_int_flexible(
                                    self._pick_value(row, ["Р—Р°СЂРµР·РµСЂРІРёСЂРѕРІР°РЅРѕ, С€С‚", "Reserved, pcs"])
                                    or self._pick_by_contains(row, ["Р·Р°СЂРµР·РµСЂРІРёСЂРѕРІР°РЅРѕ"])
                                ),
                                "price_current": self._parse_decimal_flexible(
                                    self._pick_value(row, ["РўРµРєСѓС‰Р°СЏ С†РµРЅР° СЃ СѓС‡С‘С‚РѕРј СЃРєРёРґРєРё, СЂСѓР±.", "Current price with discount, RUB"])
                                    or self._pick_by_contains(row, ["С‚РµРєСѓС‰Р°СЏ С†РµРЅР°", "current price"])
                                ),
                                "price_base": self._parse_decimal_flexible(
                                    self._pick_value(row, ["Р‘Р°Р·РѕРІР°СЏ С†РµРЅР° (С†РµРЅР° РґРѕ СЃРєРёРґРѕРє), СЂСѓР±.", "Base price (before discounts), RUB"])
                                    or self._pick_by_contains(row, ["Р±Р°Р·РѕРІР°СЏ С†РµРЅР°", "base price"])
                                ),
                                "price_premium": self._parse_decimal_flexible(
                                    self._pick_value(row, ["Р¦РµРЅР° Premium, СЂСѓР±.", "Premium price, RUB"])
                                    or self._pick_by_contains(row, ["premium"])
                                ),
                                "price_recommended": self._parse_decimal_flexible(
                                    self._pick_value(row, ["Р РµРєРѕРјРµРЅРґРѕРІР°РЅРЅР°СЏ С†РµРЅР°, СЂСѓР±.", "Recommended price, RUB"])
                                    or self._pick_by_contains(row, ["СЂРµРєРѕРјРµРЅРґРѕРІР°РЅРЅР°СЏ С†РµРЅР°", "recommended price"])
                                ),
                                "recommended_price_link": self._pick_value(
                                    row, ["РђРєС‚СѓР°Р»СЊРЅР°СЏ СЃСЃС‹Р»РєР° РЅР° СЂРµРєРѕРјРµРЅРґРѕРІР°РЅРЅСѓСЋ С†РµРЅСѓ", "Actual link to recommended price"]
                                ) or self._pick_by_contains(row, ["СЃСЃС‹Р»РєР°", "link"]),
                                "raw_data": row,
                                "last_synced_at": datetime.now(),
                            }
                            stmt_item = pg_insert(ReportProductItem).values(**row_data)
                            stmt_item = stmt_item.on_conflict_do_update(
                                constraint="uq_report_products_items_report_line",
                                set_={k: v for k, v in row_data.items() if k not in {"report_id", "line_no"}},
                            )
                            await session.execute(stmt_item)
                            rows_upserted += 1

                    reports_processed += 1

            offer_backfilled = await self._backfill_report_products_offer_ids(processed_report_ids)

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
            )
            logger.info(
                f"report_products sync completed: reports={reports_processed}, rows_upserted={rows_upserted}, offer_backfilled={offer_backfilled}"
            )
            return {"reports_processed": reports_processed, "rows_upserted": rows_upserted, "offer_backfilled": offer_backfilled}
        except Exception as e:
            logger.error(f"report_products sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== REPORT RETURNS SYNC ====================

    async def sync_returns_report(self, pages: int = 10, page_size: int = 100) -> Dict[str, Any]:
        """Skachivanie seller_returns otcheta i zagruzka strok v report_returns_items."""
        logger.info("Starting report_returns sync...")
        sync_log = await self._create_sync_log("report_returns")
        skip_result = await self._skip_recent_async_report_sync_if_fresh(sync_log, "report_returns")
        if skip_result:
            return skip_result

        reports_processed = 0
        rows_upserted = 0
        download_failed = 0

        try:
            now_utc = datetime.now(timezone.utc)
            ytd_from = now_utc.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

            # Ozon limits report period to ~60 days — split into chunks
            max_chunk_days = 60
            date_chunks: List[Tuple[datetime, datetime]] = []
            chunk_start = ytd_from
            while chunk_start < now_utc:
                chunk_end = min(chunk_start + timedelta(days=max_chunk_days), now_utc)
                date_chunks.append((chunk_start, chunk_end))
                chunk_start = chunk_end

            created_codes: Dict[str, str] = {}
            valid_statuses = ("MovingToSeller", "MovingToOzon", "ReturnedToOzon")
            for schema in ("fbo", "fbs"):
                for status in valid_statuses:
                    for chunk_from, chunk_to in date_chunks:
                        created = await self.client.create_report_returns(
                            date_from=chunk_from,
                            date_to=chunk_to,
                            delivery_schema=schema,
                            status=status,
                            language="DEFAULT",
                        )
                        code = (created.get("result") or {}).get("code") or created.get("code")
                        if code:
                            created_codes[code] = schema.upper()
                            logger.info(f"Created returns report for {schema}/{status} ({chunk_from.date()}..{chunk_to.date()}): {code}")

            for code in created_codes.keys():
                for _ in range(30):
                    info = await self.client.get_report_info(code)
                    status = str((info.get("result") or {}).get("status", "")).lower()
                    if status in {"success", "error"}:
                        break
                    await asyncio.sleep(2)

            for page in range(1, pages + 1):
                reports_page = await self.client.get_reports_list(page=page, page_size=page_size)
                reports = (reports_page.get("result") or {}).get("reports", [])
                if not reports:
                    break

                for report in reports:
                    report_code = str(report.get("code", "")).strip()
                    if not report_code:
                        continue
                    if created_codes and report_code not in created_codes:
                        continue

                    report_type = str(report.get("report_type", "")).lower()
                    is_returns_report = report_type in {"returns", "seller_returns"} or "returns" in report_code.lower()
                    if not is_returns_report:
                        continue

                    info = await self.client.get_report_info(report_code)
                    info_result = info.get("result", {}) if isinstance(info, dict) else {}
                    final_status = str(info_result.get("status") or report.get("status") or "").lower()
                    file_url = info_result.get("file") or report.get("file")
                    if final_status != "success" or not file_url:
                        continue

                    rows: List[Dict[str, Any]] = []
                    try:
                        temp_file_path = await self.client.download_file_to_tempfile(file_url, suffix=".xlsx")
                    except Exception as exc:
                        download_failed += 1
                        logger.warning(f"Skipping returns report {report_code}: file download failed: {exc}")
                        await self._enqueue_report_download_retry(
                            report_code=report_code,
                            report_type=report_type or "seller_returns",
                            file_url=file_url,
                            error_message=str(exc),
                            raw_data=info_result or report,
                        )
                        continue
                    file_size = os.path.getsize(temp_file_path)
                    try:
                        with open(temp_file_path, "rb") as f:
                            signature = f.read(8)
                        zip_signature = b"PK\x03\x04"
                        looks_like_zip = signature[:4] == zip_signature or zip_signature in signature

                        if looks_like_zip:
                            wb = openpyxl.load_workbook(temp_file_path, read_only=True, data_only=True)
                            ws = wb.active
                            all_rows = list(ws.iter_rows(values_only=True))
                            if all_rows:
                                header_idx = 0
                                header_map: Dict[str, int] = {}

                                def _normalize_header(value: Any) -> str:
                                    return " ".join(str(value or "").strip().lower().replace("\xa0", " ").split())

                                def _build_header_map(row_vals: Any) -> Dict[str, int]:
                                    mapping: Dict[str, int] = {}
                                    for col_idx, cell_value in enumerate(row_vals):
                                        header_name = _normalize_header(cell_value)
                                        if header_name and header_name not in mapping:
                                            mapping[header_name] = col_idx
                                    return mapping

                                def _pick_cell(row_vals: Any, *header_names: str) -> Any:
                                    for header_name in header_names:
                                        col_idx = header_map.get(_normalize_header(header_name))
                                        if col_idx is None or col_idx >= len(row_vals):
                                            continue
                                        value = row_vals[col_idx]
                                        if value not in (None, ""):
                                            return value
                                    return None

                                for idx, row_vals in enumerate(all_rows):
                                    row_text = " | ".join(str(v).strip().lower() for v in row_vals if v not in (None, ""))
                                    if ("номер отправления" in row_text) or ("артикул товара" in row_text):
                                        header_idx = idx
                                        header_map = _build_header_map(row_vals)
                                        break
                                for vals in all_rows[header_idx + 1:]:
                                    if not vals:
                                        continue
                                    posting_number = _pick_cell(
                                        vals,
                                        "Номер отправления",
                                        "Posting number",
                                    )
                                    if posting_number in (None, ""):
                                        continue
                                    rows.append(
                                        {
                                            "delivery_schema": _pick_cell(
                                                vals,
                                                "Схема доставки",
                                                "Схема",
                                                "Delivery schema",
                                            ),
                                            "product_name": _pick_cell(
                                                vals,
                                                "Название товара",
                                                "Наименование товара",
                                                "Product name",
                                            ),
                                            "posting_number": posting_number,
                                            "order_id": _pick_cell(
                                                vals,
                                                "Номер заказа",
                                                "Order ID",
                                            ),
                                            "offer_id": _pick_cell(
                                                vals,
                                                "Артикул товара",
                                                "Артикул",
                                                "Offer ID",
                                            ),
                                            "sku": _pick_cell(
                                                vals,
                                                "SKU",
                                                "OZON SKU ID",
                                                "Ozon sku id",
                                            ),
                                            "returned_at": _pick_cell(
                                                vals,
                                                "Дата возврата",
                                                "Дата приёмки возврата",
                                                "Return date",
                                            ),
                                            "status": _pick_cell(
                                                vals,
                                                "Статус возврата",
                                                "Статус",
                                                "Status",
                                            ),
                                            "quantity": _pick_cell(
                                                vals,
                                                "Количество",
                                                "Количество возвращаемых товаров",
                                                "Quantity",
                                            ),
                                            "refund_amount": _pick_cell(
                                                vals,
                                                "Сумма возврата",
                                                "Сумма к возврату",
                                                "Стоимость товара",
                                                "Цена без комиссии",
                                                "Refund amount",
                                            ),
                                        }
                                    )
                            wb.close()

                        if not rows:
                            with open(temp_file_path, "rb") as f:
                                file_bytes = f.read()
                            try:
                                text = file_bytes.decode("utf-8-sig")
                            except UnicodeDecodeError:
                                text = file_bytes.decode("cp1251", errors="replace")
                            parsed = False
                            for delim in (";", ","):
                                try:
                                    reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=delim)
                                    rows = list(reader)
                                    parsed = True
                                    break
                                except csv.Error:
                                    continue
                            if not parsed:
                                raise ValueError("Unable to parse returns report file")
                    finally:
                        if os.path.exists(temp_file_path):
                            os.remove(temp_file_path)

                    report_id = self._report_code_to_id(report_code)
                    created_at = self._parse_datetime_flexible(info_result.get("created_at") or report.get("created_at"))
                    completed_at = self._parse_datetime_flexible(info_result.get("updated_at") or info_result.get("created_at"))

                    async with db_manager.session() as session:
                        report_data = {
                            "report_id": report_id,
                            "report_type": report_type or "seller_returns",
                            "status": final_status,
                            "date_from": ytd_from,
                            "date_to": now_utc,
                            "filters": info_result.get("params") or report.get("params"),
                            "file_url": file_url,
                            "file_size": file_size,
                            "row_count": len(rows),
                            "created_at": created_at,
                            "completed_at": completed_at,
                            "raw_data": info_result or report,
                            "last_synced_at": datetime.now(),
                        }
                        stmt_report = pg_insert(AsyncReport).values(**report_data)
                        stmt_report = stmt_report.on_conflict_do_update(
                            index_elements=["report_id"],
                            set_={k: v for k, v in report_data.items() if k != "report_id"},
                        )
                        await session.execute(stmt_report)

                        await session.execute(
                            delete(ReportReturnItem).where(ReportReturnItem.report_id == report_id)
                        )

                        for i, row in enumerate(rows, start=1):
                            raw_row = {
                                k: (v.isoformat() if isinstance(v, datetime) else v)
                                for k, v in row.items()
                            }
                            delivery_schema = (
                                self._pick_value(row, ["delivery_schema", "Схема доставки", "Схема", "Delivery schema"])
                                or created_codes.get(report_code)
                                or "UNKNOWN"
                            )
                            returned_at = self._parse_datetime_flexible(
                                self._pick_value(row, ["returned_at", "Дата возврата", "Дата приёмки возврата", "Return date"])
                            )
                            row_data = {
                                "report_id": report_id,
                                "line_no": i,
                                "return_id": self._parse_int_flexible(
                                    self._pick_value(row, ["return_id", "ID возврата", "Return ID"])
                                ),
                                "posting_number": self._pick_value(row, ["posting_number", "Номер отправления", "Posting number"]),
                                "order_id": self._pick_value(row, ["order_id", "Номер заказа", "Order ID"]),
                                "delivery_schema": str(delivery_schema).upper(),
                                "status": self._pick_value(row, ["status", "Статус", "Статус возврата", "Status"]),
                                "offer_id": self._pick_value(row, ["offer_id", "Артикул", "Артикул товара", "Offer ID"]),
                                "sku": self._parse_int_flexible(self._pick_value(row, ["sku", "SKU", "OZON SKU ID"])),
                                "product_name": self._pick_value(row, ["product_name", "Название товара", "Наименование товара", "Product name", "name"]),
                                "quantity": self._parse_int_flexible(
                                    self._pick_value(row, ["quantity", "Количество", "Количество возвращаемых товаров", "Quantity"])
                                ),
                                "refund_amount": self._parse_decimal_flexible(
                                    self._pick_value(
                                        row,
                                        ["refund_amount", "Сумма возврата", "Сумма к возврату", "Стоимость товара", "Цена без комиссии", "Refund amount"],
                                    )
                                ),
                                "returned_at": returned_at,
                                "raw_data": raw_row,
                                "last_synced_at": datetime.now(),
                            }
                            stmt_item = pg_insert(ReportReturnItem).values(**row_data)
                            stmt_item = stmt_item.on_conflict_do_update(
                                constraint="uq_report_returns_items_report_line",
                                set_={k: v for k, v in row_data.items() if k not in {"report_id", "line_no"}},
                            )
                            await session.execute(stmt_item)
                            rows_upserted += 1

                    reports_processed += 1

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
                error_message=(
                    f"download_failed={download_failed}" if download_failed else None
                ),
            )
            logger.info(
                f"report_returns sync completed: reports={reports_processed}, rows_upserted={rows_upserted}, "
                f"download_failed={download_failed}"
            )
            return {
                "reports_processed": reports_processed,
                "rows_upserted": rows_upserted,
                "download_failed": download_failed,
            }
        except Exception as e:
            logger.error(f"report_returns sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== REPORT WAREHOUSE STOCK SYNC ====================

    async def sync_compensation_reports(self, year: Optional[int] = None) -> Dict[str, Any]:
        """Skachivanie async-otchetov po kompensacijam/dekompensacijam v BD."""
        logger.info("Starting report_compensation sync...")
        sync_log = await self._create_sync_log("report_compensation")
        skip_result = await self._skip_recent_async_report_sync_if_fresh(sync_log, "report_compensation")
        if skip_result:
            return skip_result

        reports_requested = 0
        reports_processed = 0
        rows_upserted = 0
        download_failed = 0

        now_utc = datetime.now(timezone.utc)
        if year is None:
            from_date = now_utc.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            to_date = now_utc
        else:
            from_date = datetime(year, 1, 1, tzinfo=timezone.utc)
            if year < now_utc.year:
                to_date = datetime(year, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
            else:
                to_date = now_utc

        month_starts = list(self._iter_report_months(from_date, to_date))
        # Ozon reports often not available for the current open month.
        current_month_start = now_utc.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_starts = [m for m in month_starts if m < current_month_start]
        header_tokens = [
            "дата",
            "сум",
            "компенса",
            "декомпенса",
            "артикул",
            "sku",
            "товар",
            "posting",
            "отправлен",
            "основание",
        ]
        created_reports: List[Dict[str, Any]] = []

        try:
            creators = {
                "compensation": self.client.create_report_compensation,
                "decompensation": self.client.create_report_decompensation,
            }

            decomp_available = True
            for report_kind, create_report in creators.items():
                if report_kind == "decompensation" and not decomp_available:
                    continue
                for month_start in month_starts:
                    month_key = month_start.strftime("%Y-%m")
                    try:
                        created = await create_report(month_key, language="RU")
                    except Exception as exc:
                        if isinstance(exc, OzonAPIError) and exc.status_code == 404:
                            if report_kind == "decompensation":
                                decomp_available = False
                                logger.warning("Decompensation reports are not available (HTTP 404). Skipping.")
                                break
                            logger.warning(f"{report_kind} report not available for {month_key} (HTTP 404). Skipping.")
                            continue
                        logger.warning(f"Failed to create {report_kind} report for {month_key}: {exc}")
                        continue

                    code = (created.get("result") or {}).get("code") or created.get("code")
                    if not code:
                        logger.warning(f"{report_kind} report for {month_key} returned without code")
                        continue

                    created_reports.append(
                        {
                            "report_kind": report_kind,
                            "month_key": month_key,
                            "month_start": month_start,
                            "report_code": code,
                        }
                    )
                    reports_requested += 1
                    logger.info(f"Created {report_kind} report for {month_key}: {code}")

            for report_meta in created_reports:
                report_kind = str(report_meta["report_kind"])
                month_key = str(report_meta["month_key"])
                month_start = report_meta["month_start"]
                report_code = str(report_meta["report_code"])

                info_result = await self._wait_report_info(report_code)
                final_status = str(info_result.get("status") or "").lower()
                file_url = info_result.get("file")

                if final_status != "success" or not file_url:
                    logger.warning(
                        f"Skipping {report_kind} report {report_code} for {month_key}: "
                        f"status={final_status or 'unknown'}, file_url={bool(file_url)}"
                    )
                    continue

                try:
                    file_bytes = await self.client.download_file(file_url)
                except Exception as exc:
                    download_failed += 1
                    await self._enqueue_report_download_retry(
                        report_code=report_code,
                        report_type=report_kind,
                        file_url=file_url,
                        error_message=str(exc),
                        raw_data=info_result,
                    )
                    logger.warning(f"Failed to download {report_kind} report {report_code}: {exc}")
                    continue

                rows = self._parse_tabular_report_file(file_bytes, header_tokens)
                report_id = self._report_code_to_id(report_code)
                created_at = self._parse_datetime_flexible(info_result.get("created_at"))
                completed_at = self._parse_datetime_flexible(
                    info_result.get("updated_at") or info_result.get("created_at")
                )
                last_day = monthrange(month_start.year, month_start.month)[1]
                month_end = month_start.replace(day=last_day, hour=23, minute=59, second=59, microsecond=0)

                async with db_manager.session() as session:
                    report_data = {
                        "report_id": report_id,
                        "report_type": info_result.get("report_type") or f"{report_kind}_report",
                        "status": final_status,
                        "date_from": month_start,
                        "date_to": month_end,
                        "filters": info_result.get("params") or {"date": month_key, "language": "RU"},
                        "file_url": file_url,
                        "file_size": len(file_bytes),
                        "row_count": len(rows),
                        "created_at": created_at,
                        "completed_at": completed_at,
                        "raw_data": info_result,
                        "last_synced_at": datetime.now(),
                    }
                    stmt_report = pg_insert(AsyncReport).values(**report_data)
                    stmt_report = stmt_report.on_conflict_do_update(
                        index_elements=["report_id"],
                        set_={k: v for k, v in report_data.items() if k != "report_id"},
                    )
                    await session.execute(stmt_report)

                    await session.execute(
                        delete(ReportCompensationItem).where(ReportCompensationItem.report_id == report_id)
                    )
                    await session.execute(
                        delete(ReportCompensationItem).where(
                            ReportCompensationItem.report_kind == report_kind,
                            ReportCompensationItem.report_month == self._ensure_aware_utc(month_start),
                        )
                    )

                    for line_no, row in enumerate(rows, start=1):
                        normalized = self._normalize_compensation_report_row(row, report_kind, month_start)
                        if not normalized:
                            continue

                        row_data = {
                            "report_id": report_id,
                            "line_no": line_no,
                            "report_kind": report_kind,
                            **normalized,
                            "last_synced_at": datetime.now(),
                        }
                        stmt_item = pg_insert(ReportCompensationItem).values(**row_data)
                        stmt_item = stmt_item.on_conflict_do_update(
                            constraint="uq_report_compensation_items_report_line",
                            set_={k: v for k, v in row_data.items() if k not in {"report_id", "line_no"}},
                        )
                        await session.execute(stmt_item)
                        rows_upserted += 1

                reports_processed += 1

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
                error_message=(f"download_failed={download_failed}" if download_failed else None),
            )
            logger.info(
                "report_compensation sync completed: "
                f"reports_requested={reports_requested}, reports_processed={reports_processed}, "
                f"rows_upserted={rows_upserted}, download_failed={download_failed}"
            )
            return {
                "reports_requested": reports_requested,
                "reports_processed": reports_processed,
                "rows_upserted": rows_upserted,
                "download_failed": download_failed,
            }
        except Exception as e:
            logger.error(f"report_compensation sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    async def sync_warehouse_stock_report(self, pages: int = 10, page_size: int = 100) -> Dict[str, Any]:
        """Skachivanie otcheta po ostatkam skladov i zagruzka strok v report_warehouse_stock_items."""
        logger.info("Starting report_warehouse_stock sync...")
        sync_log = await self._create_sync_log("report_warehouse_stock")
        skip_result = await self._skip_recent_async_report_sync_if_fresh(sync_log, "report_warehouse_stock")
        if skip_result:
            return skip_result

        reports_processed = 0
        rows_upserted = 0

        try:
            warehouse_ids = set()

            try:
                wh = await self.client.get_warehouses_list()
                wh_items = wh.get("warehouses", []) if isinstance(wh, dict) else []
                for item in wh_items:
                    wh_id = self._parse_int_flexible(
                        item.get("seller_warehouse_id") or item.get("warehouse_id") or item.get("id")
                    )
                    if wh_id:
                        warehouse_ids.add(wh_id)
            except Exception as e:
                logger.warning(f"Failed to load warehouse IDs from warehouses list: {e}")

            if not warehouse_ids:
                try:
                    dm = await self.client.get_delivery_methods_list(limit=1000, offset=0)
                    methods = dm.get("result", []) if isinstance(dm, dict) else []
                    for item in methods:
                        wh_id = self._parse_int_flexible(item.get("warehouse_id"))
                        if wh_id:
                            warehouse_ids.add(wh_id)
                except Exception as e:
                    logger.warning(f"Failed to load warehouse IDs from delivery methods: {e}")

            if not warehouse_ids:
                raise ValueError("No warehouse_id values found to build /v1/report/warehouse/stock request")

            created = await self.client.create_report_warehouse_stock(
                warehouse_ids=sorted(warehouse_ids),
                language="DEFAULT",
            )
            created_code = (created.get("result") or {}).get("code") or created.get("code")
            created_codes = {created_code} if created_code else set()
            if created_code:
                logger.info(f"Created warehouse stock report: {created_code}")
                for _ in range(45):
                    info = await self.client.get_report_info(created_code)
                    status = str((info.get("result") or {}).get("status", "")).lower()
                    if status in {"success", "error"}:
                        break
                    await asyncio.sleep(2)

            for page in range(1, pages + 1):
                reports_page = await self.client.get_reports_list(page=page, page_size=page_size)
                reports = (reports_page.get("result") or {}).get("reports", [])
                if not reports:
                    break

                for report in reports:
                    report_code = str(report.get("code", "")).strip()
                    if not report_code:
                        continue
                    if created_codes and report_code not in created_codes:
                        continue

                    info = await self.client.get_report_info(report_code)
                    info_result = info.get("result", {}) if isinstance(info, dict) else {}
                    report_type = str(info_result.get("report_type") or report.get("report_type") or "").lower()
                    final_status = str(info_result.get("status") or report.get("status") or "").lower()
                    file_url = info_result.get("file") or report.get("file")
                    if final_status != "success" or not file_url:
                        continue

                    is_stock_report = (
                        "stock" in report_type
                        or "warehouse" in report_type
                        or "stocks" in report_code.lower()
                        or "warehouse" in report_code.lower()
                    )
                    if not is_stock_report and created_codes:
                        is_stock_report = report_code in created_codes
                    if not is_stock_report:
                        continue

                    file_bytes = await self.client.download_file(file_url)
                    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                    ws = wb.active
                    sheet_rows = list(ws.iter_rows(values_only=True))
                    wb.close()

                    rows: List[Dict[str, Any]] = []
                    headers: List[str] = []
                    header_idx = 0
                    for i, vals in enumerate(sheet_rows):
                        if not vals:
                            continue
                        non_empty = [v for v in vals if v not in (None, "")]
                        if len(non_empty) < 1:
                            continue
                        row_text = " | ".join(str(v).strip().lower() for v in non_empty)
                        if any(token in row_text for token in ("склад", "warehouse", "артикул", "offer", "sku")):
                            header_idx = i
                            break

                    if sheet_rows:
                        raw_headers = sheet_rows[header_idx]
                        headers = [
                            (str(v).strip() if v not in (None, "") else f"column_{idx + 1}")
                            for idx, v in enumerate(raw_headers)
                        ]
                        for vals in sheet_rows[header_idx + 1:]:
                            if not vals:
                                continue
                            if all(v in (None, "") for v in vals):
                                continue
                            row_map: Dict[str, Any] = {}
                            for idx, val in enumerate(vals):
                                key = headers[idx] if idx < len(headers) else f"column_{idx + 1}"
                                row_map[key] = val.isoformat() if isinstance(val, datetime) else val
                            rows.append(row_map)

                    report_id = self._report_code_to_id(report_code)
                    created_at = self._parse_datetime_flexible(info_result.get("created_at") or report.get("created_at"))
                    completed_at = self._parse_datetime_flexible(info_result.get("updated_at") or info_result.get("created_at"))

                    async with db_manager.session() as session:
                        report_data = {
                            "report_id": report_id,
                            "report_type": report_type or "warehouse_stock",
                            "status": final_status,
                            "date_from": None,
                            "date_to": None,
                            "filters": info_result.get("params") or report.get("params"),
                            "file_url": file_url,
                            "file_size": len(file_bytes),
                            "row_count": len(rows),
                            "created_at": created_at,
                            "completed_at": completed_at,
                            "raw_data": info_result or report,
                            "last_synced_at": datetime.now(),
                        }
                        stmt_report = pg_insert(AsyncReport).values(**report_data)
                        stmt_report = stmt_report.on_conflict_do_update(
                            index_elements=["report_id"],
                            set_={k: v for k, v in report_data.items() if k != "report_id"},
                        )
                        await session.execute(stmt_report)

                        await session.execute(
                            delete(ReportWarehouseStockItem).where(ReportWarehouseStockItem.report_id == report_id)
                        )

                        for i, row in enumerate(rows, start=1):
                            warehouse_id = self._parse_int_flexible(
                                self._pick_value(row, ["Идентификатор склада", "Warehouse ID", "warehouse_id"])
                                or self._pick_by_contains(row, ["идентификатор склада", "warehouse id", "warehouse_id", "id склада"])
                            )
                            warehouse_name = (
                                self._pick_value(row, ["Склад", "Название склада", "Warehouse"])
                                or self._pick_by_contains(row, ["название склада", "склад", "warehouse"])
                            )
                            stock_total = self._parse_int_flexible(
                                self._pick_value(row, ["Остаток", "Доступно", "Stock", "Quantity"])
                                or self._pick_by_contains(row, ["остаток", "доступно", "stock", "quantity", "количество"])
                            )
                            row_data = {
                                "report_id": report_id,
                                "line_no": i,
                                "warehouse_id": warehouse_id,
                                "warehouse_name": warehouse_name,
                                "offer_id": self._pick_value(row, ["Offer ID", "Артикул", "Артикул товара"])
                                or self._pick_by_contains(row, ["offer id", "артикул"]),
                                "sku": self._parse_int_flexible(
                                    self._pick_value(row, ["SKU", "Ozon SKU ID", "FBO Ozon SKU ID", "FBS Ozon SKU ID"])
                                    or self._pick_by_contains(row, ["sku"])
                                ),
                                "product_name": self._pick_value(row, ["Название товара", "Наименование товара", "Product name", "Name"])
                                or self._pick_by_contains(row, ["название", "наименование", "product"]),
                                "stock_total": stock_total,
                                "raw_data": row,
                                "last_synced_at": datetime.now(),
                            }
                            stmt_item = pg_insert(ReportWarehouseStockItem).values(**row_data)
                            stmt_item = stmt_item.on_conflict_do_update(
                                constraint="uq_report_warehouse_stock_items_report_line",
                                set_={k: v for k, v in row_data.items() if k not in {"report_id", "line_no"}},
                            )
                            await session.execute(stmt_item)
                            rows_upserted += 1

                    reports_processed += 1

            await self._update_sync_log(
                sync_log,
                "success",
                records_processed=rows_upserted,
                records_inserted=rows_upserted,
                records_updated=0,
            )
            logger.info(
                f"report_warehouse_stock sync completed: reports={reports_processed}, rows_upserted={rows_upserted}"
            )
            return {"reports_processed": reports_processed, "rows_upserted": rows_upserted}
        except Exception as e:
            logger.error(f"report_warehouse_stock sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise
    
    # ==================== PRODUCT DIMENSIONS SYNC ====================

    async def sync_product_dimensions(self) -> Dict[str, int]:
        """Заполняет product_dimensions из products.raw_data.product_attributes_v4.

        Ozon возвращает height/depth/width в мм, weight в г (dimension_unit='mm',
        weight_unit='g'). Конвертируем в см/кг. volume_l = len*wid*hei / 1000 (см³→л).
        Источник = 'ozon_api_v4' (реальные поля из уже синхронизированного Attribute-ответа).
        """
        logger.info("Starting product dimensions sync (from products.raw_data)")
        sync_log = await self._create_sync_log("product_dimensions")

        processed = inserted = updated = skipped = 0

        try:
            async with db_manager.session() as session:
                rows = await session.execute(
                    _sql_text(
                        "SELECT offer_id, "
                        "       raw_data->'product_attributes_v4' AS v4, "
                        "       raw_data->'resolved_ids'->>'sku' AS sku "
                        "FROM products"
                    )
                )
                all_rows = rows.fetchall()

            for row in all_rows:
                offer_id = row[0]
                v4 = row[1] or {}
                sku_str = row[2]

                if not offer_id or not isinstance(v4, dict):
                    skipped += 1
                    continue

                raw_h = v4.get("height")
                raw_d = v4.get("depth")
                raw_w = v4.get("width")
                raw_wt = v4.get("weight")
                dim_unit = (v4.get("dimension_unit") or "mm").lower()
                wt_unit = (v4.get("weight_unit") or "g").lower()

                if raw_h is None or raw_d is None or raw_w is None:
                    skipped += 1
                    continue

                try:
                    h = float(raw_h); d = float(raw_d); w = float(raw_w)
                except (TypeError, ValueError):
                    skipped += 1
                    continue

                # Конвертация в см
                if dim_unit == "mm":
                    length_cm = d / 10.0
                    width_cm = w / 10.0
                    height_cm = h / 10.0
                elif dim_unit in ("cm", "см"):
                    length_cm, width_cm, height_cm = d, w, h
                elif dim_unit in ("m", "м"):
                    length_cm, width_cm, height_cm = d * 100, w * 100, h * 100
                else:
                    length_cm = d / 10.0
                    width_cm = w / 10.0
                    height_cm = h / 10.0

                # Вес в кг
                weight_kg = None
                if raw_wt is not None:
                    try:
                        wt = float(raw_wt)
                        if wt_unit == "g":
                            weight_kg = wt / 1000.0
                        elif wt_unit == "kg":
                            weight_kg = wt
                        else:
                            weight_kg = wt / 1000.0
                    except (TypeError, ValueError):
                        weight_kg = None

                volume_l = round(length_cm * width_cm * height_cm / 1000.0, 3)

                sku_val = None
                if sku_str:
                    try:
                        sku_val = int(sku_str)
                    except (TypeError, ValueError):
                        sku_val = None

                data = {
                    "offer_id": str(offer_id),
                    "sku": sku_val,
                    "length_cm": round(length_cm, 2),
                    "width_cm": round(width_cm, 2),
                    "height_cm": round(height_cm, 2),
                    "weight_kg": round(weight_kg, 3) if weight_kg is not None else None,
                    "volume_l": volume_l,
                    "source": "ozon_api_v4",
                }

                async with db_manager.session() as session:
                    exists_row = await session.execute(
                        select(ProductDimension.id).where(ProductDimension.offer_id == data["offer_id"])
                    )
                    existed = exists_row.scalar_one_or_none() is not None

                    stmt = pg_insert(ProductDimension).values(**data)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["offer_id"],
                        set_={k: v for k, v in data.items() if k != "offer_id"},
                    )
                    await session.execute(stmt)

                processed += 1
                if existed:
                    updated += 1
                else:
                    inserted += 1

            await self._update_sync_log(sync_log, "success", processed, inserted, updated)
            logger.info(
                "Product dimensions sync done: %d processed (%d new, %d updated), %d skipped",
                processed, inserted, updated, skipped,
            )
            return {"processed": processed, "inserted": inserted, "updated": updated, "skipped": skipped}
        except Exception as e:
            logger.error(f"Product dimensions sync failed: {e}")
            await self._update_sync_log(sync_log, "error", error_message=str(e))
            raise

    # ==================== FULL SYNC ====================

    async def full_sync(self, days_back: int = 30):
        """Polnaja sinhronizacija vseh dannyh."""
        logger.info("Starting full sync...")
        
        results = {}
        
        # Osnovnye dannye
        try:
            results["products"] = await self.sync_products()
        except Exception as e:
            logger.error(f"Products sync failed: {e}")
            results["products"] = {"error": str(e)}
        
        try:
            results["analytics_data"] = await self.sync_analytics_data(days_back=days_back)
        except Exception as e:
            logger.error(f"Analytics data sync failed: {e}")
            results["analytics_data"] = {"error": str(e)}

        try:
            results["analytics_stocks"] = await self.sync_analytics_stocks()
        except Exception as e:
            logger.error(f"Analytics stocks sync failed: {e}")
            results["analytics_stocks"] = {"error": str(e)}

        try:
            results["fbs_warehouse_stocks"] = await self.sync_fbs_warehouse_stocks()
        except Exception as e:
            logger.error(f"FBS warehouse stocks sync failed: {e}")
            results["fbs_warehouse_stocks"] = {"error": str(e)}

        try:
            results["analytics_turnover"] = await self.sync_analytics_turnover(days_back=days_back)
        except Exception as e:
            logger.error(f"Analytics turnover sync failed: {e}")
            results["analytics_turnover"] = {"error": str(e)}

        try:
            results["average_delivery_time"] = await self.sync_analytics_average_delivery_time()
        except Exception as e:
            logger.error(f"Average delivery time sync failed: {e}")
            results["average_delivery_time"] = {"error": str(e)}

        try:
            results["realization_v2"] = await self.sync_realization_v2(days_back=max(days_back, 365))
        except Exception as e:
            logger.error(f"Realization v2 sync failed: {e}")
            results["realization_v2"] = {"error": str(e)}
        
        try:
            results["returns"] = await self.sync_returns()
        except Exception as e:
            logger.error(f"Returns sync failed: {e}")
            results["returns"] = {"error": str(e)}
        
        try:
            results["returns_fbo"] = await self.sync_returns_fbo()
        except Exception as e:
            logger.error(f"FBO returns sync failed: {e}")
            results["returns_fbo"] = {"error": str(e)}

        try:
            results["cash_flow"] = await self.sync_cash_flow_statements(days_back=max(days_back, 365))
        except Exception as e:
            logger.error(f"Cash flow sync failed: {e}")
            results["cash_flow"] = {"error": str(e)}

        try:
            results["promo"] = await self.sync_promo()
        except Exception as e:
            logger.error(f"Promo sync failed: {e}")
            results["promo"] = {"error": str(e)}

        try:
            results["report_postings"] = await self.sync_postings_report()
        except Exception as e:
            logger.error(f"Report postings sync failed: {e}")
            results["report_postings"] = {"error": str(e)}

        try:
            results["report_products"] = await self.sync_products_report()
        except Exception as e:
            logger.error(f"Report products sync failed: {e}")
            results["report_products"] = {"error": str(e)}

        try:
            results["report_returns"] = await self.sync_returns_report()
        except Exception as e:
            logger.error(f"Report returns sync failed: {e}")
            results["report_returns"] = {"error": str(e)}

        try:
            results["report_compensation"] = await self.sync_compensation_reports()
        except Exception as e:
            logger.error(f"Report compensation sync failed: {e}")
            results["report_compensation"] = {"error": str(e)}

        try:
            results["report_warehouse_stock"] = await self.sync_warehouse_stock_report()
        except Exception as e:
            logger.error(f"Report warehouse stock sync failed: {e}")
            results["report_warehouse_stock"] = {"error": str(e)}

        logger.info("Full sync completed")
        return results


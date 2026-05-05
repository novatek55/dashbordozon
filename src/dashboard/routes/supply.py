"""Dashboard routes/supply.py handlers."""
import asyncio
import io
import json
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import asyncpg
import aiohttp
import openpyxl
import pandas as pd
from aiohttp import web

from src.config import settings
from src.dashboard.constants import (
    BASE_DIR, MSK,
    SUPPLY_CLUSTER_MARKUP_DEFAULTS, SUPPLY_MACROLOCAL_CLUSTER_FALLBACKS,
    SUPPLY_ACCEPTANCE_CACHE_MAX_ENTRIES,
)
from src.dashboard import state
from src.dashboard.helpers import (
    clean_nan_values, as_float, normalize_offer_id,
    _to_int, _ozon_supply_post, _get_env_from_dotenv, _get_ozon_credentials,
    _normalize_cluster_name, _normalize_text_key, _extract_supply_clusters,
)
from src.palletization.calculator import calculate_pallets_from_supply_plan, calculate_pallets_for_cluster, filter_small_pallets
from src.dashboard.routes.stocks import get_analytics_stocks
from src.dashboard.routes.palletization_routes import _build_palletization_products_map


async def sync_cluster_warehouses_to_db(request: web.Request) -> web.Response:
    body = await request.json() if request.body_exists else {}
    payload_clusters = body.get("clusters")
    source = "request_body"

    if not isinstance(payload_clusters, list):
        client_id = (
            os.getenv("OZON_CLIENT_ID")
            or getattr(settings, "ozon_client_id", "")
            or _get_env_from_dotenv("OZON_CLIENT_ID")
            or ""
        ).strip()
        api_key = (
            os.getenv("OZON_SUPPLY_API_KEY")
            or os.getenv("OZON_API_KEY")
            or _get_env_from_dotenv("OZON_SUPPLY_API_KEY")
            or _get_env_from_dotenv("OZON_API_KEY")
            or getattr(settings, "ozon_api_key", "")
            or ""
        ).strip()
        if not client_id or not api_key:
            return web.json_response(
                {"success": False, "error": "Missing OZON_CLIENT_ID/OZON_SUPPLY_API_KEY"},
                status=400,
            )
        headers = {
            "Client-Id": client_id,
            "Api-Key": api_key,
            "Content-Type": "application/json",
        }
        timeout = aiohttp.ClientTimeout(total=60)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            status_code, data = await _ozon_supply_post(
                session,
                "/v1/cluster/list",
                headers,
                {"cluster_type": 1},
            )
        if status_code != 200:
            return web.json_response(
                {
                    "success": False,
                    "error": "Failed to load clusters from Ozon",
                    "status": status_code,
                    "details": data,
                },
                status=502,
            )
        payload_clusters = data.get("clusters") if isinstance(data, dict) else None
        source = "ozon_api"

    if not isinstance(payload_clusters, list) or not payload_clusters:
        return web.json_response(
            {"success": False, "error": "clusters list is empty"},
            status=400,
        )

    pool: asyncpg.Pool = request.app["pool"]
    clusters_saved = 0
    warehouses_saved = 0
    skipped_warehouses = 0

    async with pool.acquire() as conn:
        async with conn.transaction():
            for cluster in payload_clusters:
                if not isinstance(cluster, dict):
                    continue
                macrolocal_cluster_id = _to_int(cluster.get("macrolocal_cluster_id"))
                if not macrolocal_cluster_id:
                    continue
                cluster_id = _to_int(cluster.get("id"))
                cluster_name = str(cluster.get("name") or "").strip() or f"cluster_{macrolocal_cluster_id}"
                cluster_type = str(cluster.get("type") or "").strip() or None

                await conn.execute(
                    """
                    INSERT INTO ozon_clusters_directory (
                        macrolocal_cluster_id, cluster_id, cluster_name, cluster_type, raw_data, updated_at
                    )
                    VALUES ($1, $2, $3, $4, $5::jsonb, now())
                    ON CONFLICT (macrolocal_cluster_id)
                    DO UPDATE SET
                        cluster_id = EXCLUDED.cluster_id,
                        cluster_name = EXCLUDED.cluster_name,
                        cluster_type = EXCLUDED.cluster_type,
                        raw_data = EXCLUDED.raw_data,
                        updated_at = now()
                    """,
                    macrolocal_cluster_id,
                    cluster_id,
                    cluster_name,
                    cluster_type,
                    json.dumps(cluster, ensure_ascii=False),
                )
                clusters_saved += 1

                # Полностью обновляем набор складов по кластеру текущим снимком.
                await conn.execute(
                    "DELETE FROM ozon_cluster_warehouses WHERE macrolocal_cluster_id = $1",
                    macrolocal_cluster_id,
                )

                logistic_clusters = cluster.get("logistic_clusters") or []
                if not isinstance(logistic_clusters, list):
                    logistic_clusters = []

                for lc_idx, logistic_cluster in enumerate(logistic_clusters):
                    warehouses = (logistic_cluster or {}).get("warehouses") or []
                    if not isinstance(warehouses, list):
                        continue
                    for warehouse in warehouses:
                        warehouse_id = _to_int((warehouse or {}).get("warehouse_id"))
                        if not warehouse_id:
                            skipped_warehouses += 1
                            continue
                        warehouse_type = str((warehouse or {}).get("type") or "").strip() or None
                        warehouse_name = str((warehouse or {}).get("name") or "").strip() or None
                        await conn.execute(
                            """
                            INSERT INTO ozon_cluster_warehouses (
                                macrolocal_cluster_id,
                                logistic_cluster_index,
                                warehouse_id,
                                warehouse_type,
                                warehouse_name,
                                raw_data,
                                updated_at
                            )
                            VALUES ($1, $2, $3, $4, $5, $6::jsonb, now())
                            ON CONFLICT (macrolocal_cluster_id, logistic_cluster_index, warehouse_id)
                            DO UPDATE SET
                                warehouse_type = EXCLUDED.warehouse_type,
                                warehouse_name = EXCLUDED.warehouse_name,
                                raw_data = EXCLUDED.raw_data,
                                updated_at = now()
                            """,
                            macrolocal_cluster_id,
                            int(lc_idx),
                            warehouse_id,
                            warehouse_type,
                            warehouse_name,
                            json.dumps(warehouse, ensure_ascii=False),
                        )
                        warehouses_saved += 1

    return web.json_response(
        {
            "success": True,
            "source": source,
            "clusters_saved": clusters_saved,
            "warehouses_saved": warehouses_saved,
            "skipped_warehouses": skipped_warehouses,
        }
    )


async def upload_supply_file(request: web.Request) -> web.Response:
    """POST /api/supply-plan/upload-supply-file — парсит Excel-файл поставки Ozon.

    Ожидает multipart с полем 'file' (xlsx).
    Возвращает {items: [{article, name, quantity, ozon_id, barcode}]}.
    """
    reader = await request.multipart()
    file_part = await reader.next()
    if file_part is None or file_part.name != "file":
        return web.json_response({"error": "Требуется multipart field 'file'."}, status=400)

    content = await file_part.read(decode=False)
    if not content:
        return web.json_response({"error": "Файл пустой."}, status=400)

    try:
        frame = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        return web.json_response({"error": f"Не удалось прочитать Excel: {exc}"}, status=400)

    cols_lower = {str(c).strip().lower().replace("ё", "е"): c for c in frame.columns if c is not None}

    art_col = None
    for candidate in ("артикул", "offer_id", "article"):
        if candidate in cols_lower:
            art_col = cols_lower[candidate]
            break

    qty_col = None
    for candidate in ("количество", "quantity", "кол-во", "qty"):
        if candidate in cols_lower:
            qty_col = cols_lower[candidate]
            break

    name_col = None
    for candidate in ("имя (необязательно)", "имя", "название", "name"):
        if candidate in cols_lower:
            name_col = cols_lower[candidate]
            break

    ozon_id_col = None
    for candidate in ("ozon id", "ozon_id", "sku"):
        if candidate in cols_lower:
            ozon_id_col = cols_lower[candidate]
            break

    barcode_col = None
    for candidate in ("штрихкод", "barcode"):
        if candidate in cols_lower:
            barcode_col = cols_lower[candidate]
            break

    if art_col is None or qty_col is None:
        return web.json_response(
            {
                "error": "Нужны колонки: 'артикул' и 'количество'.",
                "columns": [str(c) for c in frame.columns],
            },
            status=400,
        )

    items: list = []
    for _, row in frame.iterrows():
        article = str(row.get(art_col, "")).strip()
        if not article or article.lower() == "nan":
            continue
        qty_raw = row.get(qty_col, 0)
        try:
            qty = int(float(qty_raw))
        except (ValueError, TypeError):
            qty = 0
        if qty <= 0:
            continue
        item: dict = {"article": article, "quantity": qty}
        if name_col is not None:
            val = str(row.get(name_col, "")).strip()
            if val and val.lower() != "nan":
                item["name"] = val
        if ozon_id_col is not None:
            val = str(row.get(ozon_id_col, "")).strip()
            if val and val.lower() != "nan":
                item["ozon_id"] = val
        if barcode_col is not None:
            val = str(row.get(barcode_col, "")).strip()
            if val and val.lower() != "nan":
                item["barcode"] = val
        items.append(item)

    if not items:
        return web.json_response({"error": "Не найдено строк с положительным количеством."}, status=400)

    return web.json_response({"items": items, "total": len(items)})


async def save_supply_plan_state(request: web.Request) -> web.Response:
    try:
        payload = await request.json()
    except Exception:
        return web.json_response({"error": "Ожидается JSON body."}, status=400)

    offer_id_raw = str(payload.get("offer_id") or "")
    offer_id = normalize_offer_id(offer_id_raw)
    product_id = _to_int(payload.get("product_id"))
    if not offer_id and not product_id:
        return web.json_response({"error": "Нужно передать 'offer_id' или 'product_id'."}, status=400)

    supply_stock_raw = payload.get("supply_stock")
    hidden_raw = payload.get("hidden")
    if supply_stock_raw is None and hidden_raw is None:
        return web.json_response({"error": "Нужно передать 'supply_stock' или 'hidden'."}, status=400)

    current_supply_stock = max(0, int(as_float(supply_stock_raw, default=0.0))) if supply_stock_raw is not None else None
    current_hidden = bool(hidden_raw) if hidden_raw is not None else None

    pool: asyncpg.Pool = request.app["pool"]
    async with pool.acquire() as conn:
        canonical_offer_id = offer_id_raw.strip() or offer_id
        canonical_product_id = product_id
        product_row = None
        if product_id:
            product_row = await conn.fetchrow(
                """
                SELECT product_id, offer_id
                FROM products
                WHERE product_id = $1
                """,
                product_id,
            )
            if product_row is None and not offer_id:
                return web.json_response({"error": "Указан неизвестный product_id."}, status=400)
        if product_row is None and offer_id:
            product_row = await conn.fetchrow(
                """
                SELECT product_id, offer_id
                FROM products
                WHERE offer_id = $1
                ORDER BY updated_at DESC NULLS LAST, product_id DESC
                LIMIT 1
                """,
                offer_id_raw.strip(),
            )
        if product_row:
            canonical_product_id = _to_int(product_row["product_id"])
            canonical_offer_id = str(product_row["offer_id"] or "").strip() or canonical_offer_id
            offer_id = normalize_offer_id(canonical_offer_id)
        elif product_id:
            canonical_product_id = None

        matching_rows = await conn.fetch(
            """
            SELECT offer_id, product_id, supply_stock, hidden, updated_at
            FROM supply_plan_state
            """
        )
        normalized_matches = [
            row
            for row in matching_rows
            if (
                canonical_product_id
                and _to_int(row["product_id"]) == canonical_product_id
            )
            or normalize_offer_id(row["offer_id"]) == offer_id
        ]
        existing = None
        if normalized_matches:
            normalized_matches.sort(
                key=lambda row: row["updated_at"] or datetime.min.replace(tzinfo=timezone.utc),
                reverse=True,
            )
            existing = normalized_matches[0]
        final_supply_stock = current_supply_stock if current_supply_stock is not None else int(existing["supply_stock"] or 0) if existing else 0
        final_hidden = current_hidden if current_hidden is not None else bool(existing["hidden"]) if existing else False
        await conn.execute(
            """
            DELETE FROM supply_plan_state
            WHERE offer_id = any($1::text[])
            """,
            [str(row["offer_id"] or "") for row in normalized_matches],
        )
        await conn.execute(
            """
            INSERT INTO supply_plan_state (offer_id, product_id, supply_stock, hidden, updated_at)
            VALUES ($1, $2, $3, $4, now())
            ON CONFLICT (offer_id) DO UPDATE
            SET product_id = EXCLUDED.product_id,
                supply_stock = EXCLUDED.supply_stock,
                hidden = EXCLUDED.hidden,
                updated_at = now()
            """,
            canonical_offer_id or offer_id,
            canonical_product_id,
            final_supply_stock,
            final_hidden,
        )

    return web.json_response(
        {
            "status": "ok",
            "offer_id": canonical_offer_id or offer_id,
            "product_id": canonical_product_id,
            "supply_stock": final_supply_stock,
            "hidden": final_hidden,
        }
    )


async def reset_hidden_supply_plan_items(request: web.Request) -> web.Response:
    pool: asyncpg.Pool = request.app["pool"]
    async with pool.acquire() as conn:
        result = await conn.execute(
            """
            UPDATE supply_plan_state
            SET hidden = FALSE,
                updated_at = now()
            WHERE hidden = TRUE
            """
        )
    return web.json_response({"status": "ok", "result": result})


async def fill_supply_plan_from_availability_report(request: web.Request) -> web.Response:
    try:
        form = await request.post()
        file_field = form.get("file")
        if file_field is None:
            return web.json_response(
                {
                    "error": "Нужно загрузить XLSX-файл отчета. Поиск файла по имени больше не используется.",
                },
                status=400,
            )

        filename = str(getattr(file_field, "filename", "") or "").strip()
        if not filename.lower().endswith(".xlsx"):
            return web.json_response({"error": "Поддерживаются только файлы .xlsx"}, status=400)

        content = file_field.file.read()
        if not content:
            return web.json_response({"error": "Загруженный файл пустой"}, status=400)

        targets, validation = _load_supply_targets_from_availability_bytes(content, filename)
    except Exception as e:
        return web.json_response({"error": str(e)}, status=400)
    state.SUPPLY_PLAN_UPLOADED_TARGETS = targets
    state.SUPPLY_PLAN_UPLOADED_META = {
        **validation,
        "uploaded_at": datetime.now(MSK).isoformat(),
    }

    return web.json_response(
        {
            "status": "ok",
            "report_file": filename,
            "articles": len(targets),
            "validation": validation,
        }
    )


async def get_supply_plan(request: web.Request) -> web.Response:
    analytics_response = await get_analytics_stocks(request)
    payload = json.loads(analytics_response.text)
    items = payload.get("items", [])
    summary = payload.get("summary", {})

    unified_moscow_name = "Москва, МО и Дальние регионы"
    normalized_unified_moscow = _normalize_cluster_name(unified_moscow_name)
    direct_moscow_warehouses = {
        _normalize_cluster_name("ЖУКОВСКИЙ_РФЦ"),
        _normalize_cluster_name("НОГИНСК_РФЦ"),
        _normalize_cluster_name("ПУШКИНО_1_РФЦ"),
    }
    unified_spb_name = "Санкт-Петербург и СЗО"
    normalized_unified_spb = _normalize_cluster_name(unified_spb_name)
    direct_spb_warehouses = {
        _normalize_cluster_name("САНКТ_ПЕТЕРБУРГ_РФЦ"),
        _normalize_cluster_name("САНКТ-ПЕТЕРБУРГ_РФЦ"),
        _normalize_cluster_name("САНКТ-ПЕТЕРБУРГ РФЦ"),
    }

    supply_skus: List[int] = sorted(
        {
            _to_int(item.get("sku"))
            for item in items
            if _to_int(item.get("sku"))
        }
    )

    pool: asyncpg.Pool = request.app["pool"]
    sku_price_map: Dict[int, float] = {}
    async with pool.acquire() as conn:
        state_rows = await conn.fetch(
            """
            SELECT offer_id, product_id, supply_stock, hidden, updated_at
            FROM supply_plan_state
            """
        )
        product_rows = await conn.fetch(
            """
            SELECT product_id, offer_id
            FROM products
            WHERE coalesce(trim(offer_id), '') <> ''
            """
        )
        location_rows = await conn.fetch(
            """
            SELECT DISTINCT
                coalesce(nullif(trim(cluster_name), ''), nullif(trim(warehouse_name), '')) AS location_name
            FROM analytics_stocks
            WHERE coalesce(trim(cluster_name), '') <> ''
               OR coalesce(trim(warehouse_name), '') <> ''
            ORDER BY 1
            """
        )
        if supply_skus:
            price_rows = await conn.fetch(
                """
                WITH src AS (
                    SELECT fbo_sku_id::bigint AS sku, price_current, price_base, last_synced_at
                    FROM report_products_items
                    WHERE fbo_sku_id IS NOT NULL
                    UNION ALL
                    SELECT fbs_sku_id::bigint AS sku, price_current, price_base, last_synced_at
                    FROM report_products_items
                    WHERE fbs_sku_id IS NOT NULL
                ),
                latest_products AS (
                    SELECT DISTINCT ON (sku)
                        sku,
                        CASE
                            WHEN coalesce(price_current, 0) > 0 THEN price_current
                            WHEN coalesce(price_base, 0) > 0 THEN price_base
                            ELSE NULL
                        END AS price_value
                    FROM src
                    WHERE coalesce(price_current, 0) > 0 OR coalesce(price_base, 0) > 0
                    ORDER BY sku, last_synced_at DESC NULLS LAST
                ),
                latest_orders AS (
                    SELECT
                        sku::bigint AS sku,
                        avg(price)::numeric(12, 2) AS price_value
                    FROM fact_order_items
                    WHERE price IS NOT NULL
                      AND price > 0
                      AND sku = any($1::bigint[])
                      AND last_synced_at >= now() - interval '90 days'
                    GROUP BY sku
                )
                SELECT
                    s.sku,
                    coalesce(lp.price_value, lo.price_value) AS price_value
                FROM unnest($1::bigint[]) AS s(sku)
                LEFT JOIN latest_products lp ON lp.sku = s.sku
                LEFT JOIN latest_orders lo ON lo.sku = s.sku
                """,
                supply_skus,
            )
            for row in price_rows:
                sku_key = _to_int(row.get("sku"))
                price_value = as_float(row.get("price_value"), default=0.0)
                if sku_key and price_value > 0:
                    sku_price_map[sku_key] = float(price_value)

    product_id_by_offer_norm: Dict[str, int] = {}
    for row in product_rows:
        offer_norm = normalize_offer_id(row["offer_id"])
        product_id = _to_int(row["product_id"])
        if offer_norm and product_id:
            product_id_by_offer_norm[offer_norm] = product_id

    state_map: Dict[str, Dict[str, Any]] = {}
    for row in state_rows:
        offer_norm = normalize_offer_id(row["offer_id"])
        row_product_id = _to_int(row["product_id"])
        if not offer_norm and not row_product_id:
            continue
        state_key = f"product:{row_product_id}" if row_product_id else f"offer:{offer_norm}"
        existing_state = state_map.get(state_key)
        row_updated_at = row["updated_at"]
        existing_updated_at = existing_state.get("_updated_at_raw") if existing_state else None
        if existing_state and existing_updated_at and row_updated_at and existing_updated_at >= row_updated_at:
            continue
        if existing_state and existing_updated_at and not row_updated_at:
            continue
        state_map[state_key] = {
            "product_id": row_product_id,
            "supply_stock": int(row["supply_stock"] or 0),
            "hidden": bool(row["hidden"]),
            "updated_at": row_updated_at.isoformat() if row_updated_at else None,
            "_updated_at_raw": row_updated_at,
        }
    for entry in state_map.values():
        entry.pop("_updated_at_raw", None)

    planning_locations: List[Dict[str, str]] = []
    seen_location_keys: set[str] = set()
    for row in location_rows:
        location_name = str(row["location_name"] or "").strip()
        if not location_name:
            continue
        normalized_location = _normalize_cluster_name(location_name)
        is_moscow_group = (
            normalized_location == normalized_unified_moscow
            or normalized_location in direct_moscow_warehouses
            or "моск" in normalized_location
            or "moscow" in normalized_location
        )
        is_spb_group = (
            normalized_location == normalized_unified_spb
            or normalized_location in direct_spb_warehouses
            or "санкт петербург" in normalized_location
            or normalized_location.startswith("спб")
        )
        if is_moscow_group:
            location_name = unified_moscow_name
            normalized_location = normalized_unified_moscow
        elif is_spb_group:
            location_name = unified_spb_name
            normalized_location = normalized_unified_spb
        if not normalized_location or normalized_location in seen_location_keys:
            continue
        seen_location_keys.add(normalized_location)
        planning_locations.append(
            {
                "warehouse_name": location_name,
                "cluster_name": location_name,
                "norm": normalized_location,
            }
        )

    visible_items: List[Dict[str, Any]] = []
    hidden_count = 0
    for item in items:
        offer_norm = normalize_offer_id(item.get("offer_id"))
        item_product_id = product_id_by_offer_norm.get(offer_norm)
        item_state: Dict[str, Any] = {}
        if item_product_id:
            item_state = state_map.get(f"product:{item_product_id}", {})
        if not item_state:
            item_state = state_map.get(f"offer:{offer_norm}", {})
        if item_state.get("hidden"):
            hidden_count += 1
            continue

        supply_stock = int(item_state.get("supply_stock") or 0)
        remaining = supply_stock
        detail_rows: List[Dict[str, Any]] = []
        offer_uploaded_targets = state.SUPPLY_PLAN_UPLOADED_TARGETS.get(offer_norm, {})
        fbo_details = [
            detail
            for detail in item.get("details", [])
            if detail.get("stock_type") == "FBO"
        ]
        merged_fbo_details: Dict[str, Dict[str, Any]] = {}
        for detail in fbo_details:
            warehouse_name = str(detail.get("warehouse_name") or "")
            cluster_name = str(detail.get("cluster_name") or "")
            normalized_warehouse = _normalize_cluster_name(warehouse_name)
            normalized_cluster = _normalize_cluster_name(cluster_name)
            is_moscow_group = (
                normalized_cluster == normalized_unified_moscow
                or normalized_warehouse in direct_moscow_warehouses
            )
            is_spb_group = (
                normalized_cluster == normalized_unified_spb
                or normalized_warehouse in direct_spb_warehouses
                or "санкт петербург" in normalized_cluster
                or "санкт петербург" in normalized_warehouse
                or normalized_cluster.startswith("спб")
                or normalized_warehouse.startswith("спб")
            )
            effective_name = cluster_name or warehouse_name
            effective_norm = normalized_cluster or normalized_warehouse
            if is_moscow_group:
                effective_name = unified_moscow_name
                effective_norm = normalized_unified_moscow
            elif is_spb_group:
                effective_name = unified_spb_name
                effective_norm = normalized_unified_spb

            if not effective_norm:
                effective_norm = _normalize_cluster_name(effective_name)
            if not effective_norm:
                merged_fbo_details[str(len(merged_fbo_details))] = dict(detail)
                continue

            target_detail = merged_fbo_details.get(effective_norm)
            if target_detail is None:
                target_detail = {
                    **detail,
                    "warehouse_name": effective_name,
                    "cluster_name": effective_name,
                    "avg_daily_sales": 0.0,
                    "recommended_supply": 0,
                    "fbo_stock": 0,
                }
                merged_fbo_details[effective_norm] = target_detail

            target_detail["avg_daily_sales"] = float(target_detail.get("avg_daily_sales") or 0.0) + float(
                detail.get("avg_daily_sales") or 0.0
            )
            target_detail["recommended_supply"] = int(target_detail.get("recommended_supply") or 0) + int(
                detail.get("recommended_supply") or 0
            )
            target_detail["fbo_stock"] = int(target_detail.get("fbo_stock") or 0) + int(detail.get("fbo_stock") or 0)
            if target_detail.get("sku") is None and detail.get("sku") is not None:
                target_detail["sku"] = detail.get("sku")
        fbo_details = list(merged_fbo_details.values())

        existing_location_keys: set[str] = set()
        for detail in fbo_details:
            norm_key = _normalize_cluster_name(
                str(detail.get("cluster_name") or detail.get("warehouse_name") or "")
            )
            if norm_key:
                existing_location_keys.add(norm_key)
        for location in planning_locations:
            norm_key = location.get("norm")
            if not norm_key or norm_key in existing_location_keys:
                continue
            fbo_details.append(
                {
                    "stock_type": "FBO",
                    "sku": item.get("sku"),
                    "warehouse_name": location.get("warehouse_name"),
                    "cluster_name": location.get("cluster_name"),
                    "avg_daily_sales": 0.0,
                    "recommended_supply": 0,
                }
            )
            existing_location_keys.add(norm_key)

        allocation_rows: List[Tuple[int, Dict[str, Any]]] = []
        for detail in fbo_details:
            location_norm = _normalize_cluster_name(
                str(detail.get("cluster_name") or detail.get("warehouse_name") or "")
            )
            uploaded_target = None
            if location_norm and location_norm in offer_uploaded_targets:
                uploaded_target = int(offer_uploaded_targets.get(location_norm) or 0)
            required = int(detail.get("recommended_supply") or 0)
            target_supply = uploaded_target if uploaded_target is not None else required
            allocation_rows.append(
                (
                    max(0, int(target_supply)),
                    detail,
                )
            )

        allocation_rows.sort(
            key=lambda row: (
                -float((row[1] or {}).get("avg_daily_sales") or 0.0),
                -int(row[0] or 0),
                (row[1] or {}).get("warehouse_name") or (row[1] or {}).get("cluster_name") or "",
            )
        )
        for target_supply, detail in allocation_rows:
            required = int(detail.get("recommended_supply") or 0)
            allocated = min(remaining, int(target_supply or 0))
            remaining -= allocated
            detail_rows.append(
                {
                    "sku": detail.get("sku"),
                    "warehouse_name": detail.get("warehouse_name"),
                    "cluster_name": detail.get("cluster_name"),
                    "fbo_stock": int(detail.get("fbo_stock") or 0),
                    "avg_daily_sales": detail.get("avg_daily_sales"),
                    "recommended_supply": required,
                    "allocated_supply": allocated,
                }
            )

        visible_items.append(
            {
                "product_id": item_product_id,
                "offer_id": item.get("offer_id"),
                "sku": item.get("sku"),
                "price_current": float(
                    sku_price_map.get(_to_int(item.get("sku")) or 0)
                    or as_float(item.get("price_current"), default=0.0)
                    or 0.0
                ),
                "article_tags": item.get("article_tags") or [],
                "recommended_supply": int(item.get("recommended_supply") or 0),
                "supply_stock": supply_stock,
                "remaining_supply": remaining,
                "turnover_grade": item.get("turnover_grade"),
                "avg_daily_sales": item.get("avg_daily_sales"),
                "stock_fbo": int(item.get("stock_fbo") or 0),
                "stock_fbs": int(item.get("stock_fbs") or 0),
                "stock_fbo_available": int(item.get("stock_fbo_available") or 0),
                "stock_fbo_supply": int(item.get("stock_fbo_supply") or 0),
                "stock_fbo_transit": int(item.get("stock_fbo_transit") or 0),
                "stock_fbo_acceptance": int(item.get("stock_fbo_acceptance") or 0),
                "sales_series_28": item.get("sales_series_28"),
                "sales_dates_28": item.get("sales_dates_28"),
                "depletion_within_horizon": item.get("depletion_within_horizon"),
                "depletion_days_left": item.get("depletion_days_left"),
                "depletion_date_label": item.get("depletion_date_label"),
                "last_synced_at": item.get("last_synced_at"),
                "details": detail_rows,
            }
        )

    visible_items.sort(key=lambda item: (-(item["recommended_supply"] or 0), item["offer_id"] or ""))
    supply_summary = {
        "articles": len(visible_items),
        "hidden_articles": hidden_count,
        "recommended_supply": sum(int(item["recommended_supply"] or 0) for item in visible_items),
        "supply_stock": sum(int(item["supply_stock"] or 0) for item in visible_items),
        "remaining_supply": sum(int(item["remaining_supply"] or 0) for item in visible_items),
        "target_days": summary.get("target_days", 28),
        "stock_fbo": summary.get("stock_fbo", 0),
        "stock_fbs": summary.get("stock_fbs", 0),
        "stock_fbo_available": summary.get("stock_fbo_available", 0),
        "stock_fbo_supply": summary.get("stock_fbo_supply", 0),
        "stock_fbo_transit": summary.get("stock_fbo_transit", 0),
        "stock_fbo_acceptance": summary.get("stock_fbo_acceptance", 0),
        "uploaded_targets": bool(state.SUPPLY_PLAN_UPLOADED_TARGETS),
        "uploaded_targets_meta": state.SUPPLY_PLAN_UPLOADED_META if state.SUPPLY_PLAN_UPLOADED_TARGETS else None,
    }
    return web.json_response({"count": len(visible_items), "items": visible_items, "summary": supply_summary})


async def calculate_supply_plan_pallets(request: web.Request) -> web.Response:
    """Рассчитать паллетизацию для данных поставки"""
    import traceback
    try:
        body = await request.json()
        supply_items = body.get("items", [])
        
        if not supply_items:
            return web.json_response({"success": True, "clusters": []})
        
        # Получаем справочник товаров из основной БД
        products_db = await _build_palletization_products_map(request.app["pool"])
        
        # Рассчитываем паллеты
        results = calculate_pallets_from_supply_plan(supply_items, products_db)
        tariffs, tariff_file = _load_crossdock_tariffs_sc_rows()
        pickup_points = await _resolve_crossdock_pickup_points_from_env(request.app.get("pool"))
        cluster_markup_map = await _load_cluster_markup_tariffs(request.app.get("pool"))
        if tariffs:
            _estimate_crossdock_costs_for_pallet_clusters(results, tariffs, pickup_points, cluster_markup_map)

        return web.json_response(
            {
                "success": True,
                "clusters": results,
                "crossdock_tariff": {
                    "file": tariff_file,
                    "rows": len(tariffs),
                    "pickup_points": pickup_points,
                    "enabled": bool(tariffs),
                },
            }
        )
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"ERROR in calculate_supply_plan_pallets: {error_details}")
        return web.json_response({"success": False, "error": str(e), "details": error_details}, status=500)


async def build_supply_plan_acceptance(request: web.Request) -> web.Response:
    """Построить отчет по принимающим складам (создает черновики в Ozon)."""
    import traceback
    try:
        body = await request.json() if request.body_exists else {}
        supply_items = body.get("items", [])
        use_multi_cluster_draft = bool(body.get("use_multi_cluster_draft"))
        if not isinstance(supply_items, list) or not supply_items:
            return web.json_response({"success": False, "error": "items is required"}, status=400)

        warehouse_acceptance = await build_supply_acceptance_report(
            supply_items,
            fetch_timeslots=False,
            pool=request.app.get("pool"),
            use_multi_cluster_draft=use_multi_cluster_draft,
        )
        return web.json_response({"success": True, "warehouse_acceptance": warehouse_acceptance})
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"ERROR in build_supply_plan_acceptance: {error_details}")
        return web.json_response({"success": False, "error": str(e), "details": error_details}, status=500)


async def filter_supply_plan_pallets(request: web.Request) -> web.Response:
    """POST /api/supply-plan/pallets/filter — удалить паллеты ниже указанной высоты.

    Body:
        clusters: list (текущие кластеры из паллетизации)
        min_height: float (минимальная высота в метрах, по умолчанию 0.4)
    """
    try:
        body = await request.json() if request.body_exists else {}
        clusters_data = body.get("clusters") or []
        min_height = float(body.get("min_height", 0.4) or 0.4)

        if not clusters_data:
            return web.json_response({"success": False, "error": "Нет данных кластеров"}, status=400)

        result = filter_small_pallets(clusters_data, min_height_any=min_height)
        removed_total = sum(len(c.get("filtered_out", [])) for c in result)
        return web.json_response({
            "success": True,
            "clusters": result,
            "removed_total": removed_total,
            "min_height": min_height,
        })
    except Exception as e:
        import traceback
        return web.json_response({"success": False, "error": str(e), "details": traceback.format_exc()}, status=500)


async def export_supply_plan_pallets(request: web.Request) -> web.StreamResponse:
    """POST /api/supply-plan/pallets/export — сохранить текущую паллетизацию в XLSX."""
    try:
        body = await request.json() if request.body_exists else {}
        clusters_data = body.get("clusters") or []
        if not isinstance(clusters_data, list) or not clusters_data:
            return web.json_response({"success": False, "error": "Нет данных кластеров"}, status=400)

        workbook = openpyxl.Workbook()
        summary_ws = workbook.active
        summary_ws.title = "Сводка"
        summary_ws.append(
            [
                "Кластер",
                "Паллет",
                "SKU на паллетах",
                "Всего штук",
                "Высота паллет, м",
                "Вес паллет, кг",
                "Ошибок",
                "Отсутствующих товаров",
            ]
        )

        details_ws = workbook.create_sheet("Паллеты")
        details_ws.append(
            [
                "Кластер",
                "Паллета",
                "Артикул",
                "SKU",
                "Название",
                "Количество",
                "Высота, м",
                "Вес, кг",
                "Слоев",
            ]
        )

        errors_ws = workbook.create_sheet("Ошибки")
        errors_ws.append(["Кластер", "Ошибка"])

        missing_ws = workbook.create_sheet("Отсутствуют")
        missing_ws.append(["Кластер", "Артикул / SKU"])

        for cluster in clusters_data:
            cluster_name = str(cluster.get("cluster") or "")
            pallets = cluster.get("pallets") or []
            errors = cluster.get("errors") or []
            missing_products = cluster.get("missing_products") or []
            pallet_items = [
                item
                for pallet in pallets
                for item in (pallet.get("items") or [])
                if isinstance(item, dict)
            ]
            total_qty = sum(_to_int(item.get("quantity")) or 0 for item in pallet_items)
            total_height = sum(as_float(pallet.get("total_height"), default=0.0) for pallet in pallets)
            total_weight = sum(as_float(pallet.get("total_weight"), default=0.0) for pallet in pallets)

            summary_ws.append(
                [
                    cluster_name,
                    len(pallets),
                    len(pallet_items),
                    total_qty,
                    round(total_height, 3),
                    round(total_weight, 2),
                    len(errors),
                    len(missing_products),
                ]
            )

            for pallet in pallets:
                pallet_number = _to_int(pallet.get("pallet_number")) or 0
                for item in pallet.get("items") or []:
                    if not isinstance(item, dict):
                        continue
                    details_ws.append(
                        [
                            cluster_name,
                            pallet_number,
                            item.get("offer_id"),
                            item.get("sku"),
                            item.get("name"),
                            _to_int(item.get("quantity")) or 0,
                            as_float(item.get("height"), default=0.0),
                            as_float(item.get("weight"), default=0.0),
                            _to_int(item.get("layers")) or 0,
                        ]
                    )

            for error_text in errors:
                errors_ws.append([cluster_name, str(error_text or "")])

            for missing in missing_products:
                missing_ws.append([cluster_name, str(missing or "")])

        for ws in (summary_ws, details_ws, errors_ws, missing_ws):
            for column_cells in ws.columns:
                max_len = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"supply_palletization_{datetime.now(MSK).strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        return web.Response(
            body=output.getvalue(),
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
    except Exception as e:
        import traceback
        return web.json_response({"success": False, "error": str(e), "details": traceback.format_exc()}, status=500)


async def repack_supply_plan_cluster(request: web.Request) -> web.Response:
    """Пересчитать один кластер паллет по переданным SKU и количествам."""
    import traceback
    try:
        body = await request.json() if request.body_exists else {}
        cluster_name = str(body.get("cluster_name") or "Unknown").strip() or "Unknown"
        items_raw = body.get("items", [])
        if not isinstance(items_raw, list):
            return web.json_response({"success": False, "error": "items is required"}, status=400)

        cluster_items: List[Dict[str, Any]] = []
        for raw in items_raw:
            if not isinstance(raw, dict):
                continue
            sku = str(raw.get("sku") or "").strip()
            qty = _to_int(raw.get("quantity")) or 0
            if not sku or qty <= 0:
                continue
            cluster_items.append(
                {
                    "cluster": cluster_name,
                    "sku": sku,
                    "quantity": int(qty),
                }
            )

        if not cluster_items:
            return web.json_response(
                {
                    "success": True,
                    "cluster": {
                        "cluster": cluster_name,
                        "pallets": [],
                        "errors": [],
                        "missing_products": [],
                        "split_sku_count": 0,
                    },
                }
            )

        products_db = await _build_palletization_products_map(request.app["pool"])
        cluster_result = calculate_pallets_for_cluster(cluster_items, products_db)
        tariffs, _tariff_file = _load_crossdock_tariffs_sc_rows()
        pickup_points = await _resolve_crossdock_pickup_points_from_env(request.app.get("pool"))
        cluster_markup_map = await _load_cluster_markup_tariffs(request.app.get("pool"))
        if tariffs:
            _estimate_crossdock_costs_for_pallet_clusters([cluster_result], tariffs, pickup_points, cluster_markup_map)
        return web.json_response({"success": True, "cluster": cluster_result})
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"ERROR in repack_supply_plan_cluster: {error_details}")
        return web.json_response({"success": False, "error": str(e), "details": error_details}, status=500)


def _resolve_availability_report_path(filename: Optional[str] = None) -> Path:
    if filename:
        candidate = Path(filename.strip())
        if not candidate.is_absolute():
            candidate = BASE_DIR / candidate
        if candidate.exists() and candidate.is_file():
            return candidate
        raise FileNotFoundError(f"Файл не найден: {candidate}")

    candidates = sorted(
        [
            p
            for p in BASE_DIR.glob("Доступность товаров_*.xlsx")
            if p.is_file() and not p.name.startswith("~$")
        ],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("Не найден файл вида 'Доступность товаров_*.xlsx'")
    return candidates[0]


def _load_supply_targets_from_availability_report(path: Path) -> Dict[str, Dict[str, int]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    targets, _validation = _extract_supply_targets_from_availability_workbook(wb, path.name)
    return targets


def _resolve_crossdock_tariff_file_path() -> Optional[Path]:
    candidates = sorted(
        [p for p in BASE_DIR.glob("tariffs-cross-dock-sc-*.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _resolve_crossdock_tariff_file_path() -> Optional[Path]:
    candidates = sorted(
        [p for p in BASE_DIR.glob("tariffs-cross-dock-sc-*.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _load_crossdock_tariffs_sc_rows() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    path = _resolve_crossdock_tariff_file_path()
    if path is None:
        return [], None
    mtime = path.stat().st_mtime
    if (
        state._CROSSDOCK_TARIFFS_CACHE_ROWS
        and state._CROSSDOCK_TARIFFS_CACHE_PATH == str(path)
        and state._CROSSDOCK_TARIFFS_CACHE_MTIME == mtime
    ):
        return state._CROSSDOCK_TARIFFS_CACHE_ROWS, path.name

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        tariff_zone = str(row[2 - 1] or "").strip()
        pickup_point = str(row[3 - 1] or "").strip()
        recipient_cluster = str(row[4 - 1] or "").strip()
        recipient_warehouse = str(row[5 - 1] or "").strip()
        shipment_pallet = as_float(row[7 - 1], default=0.0)
        transport_pallet = as_float(row[9 - 1], default=0.0)
        lead_time = str(row[10 - 1] or "").strip()
        if not tariff_zone or not recipient_cluster:
            continue
        rows.append(
            {
                "tariff_zone": tariff_zone,
                "tariff_zone_norm": _normalize_text_key(tariff_zone),
                "pickup_point": pickup_point,
                "pickup_point_norm": _normalize_text_key(pickup_point),
                "recipient_cluster": recipient_cluster,
                "recipient_cluster_norm": _normalize_text_key(recipient_cluster),
                "recipient_warehouse": recipient_warehouse,
                "shipment_pallet": float(max(0.0, shipment_pallet)),
                "transport_pallet": float(max(0.0, transport_pallet)),
                "total_pallet": float(max(0.0, shipment_pallet + transport_pallet)),
                "lead_time": lead_time,
            }
        )
    state._CROSSDOCK_TARIFFS_CACHE_PATH = str(path)
    state._CROSSDOCK_TARIFFS_CACHE_MTIME = mtime
    state._CROSSDOCK_TARIFFS_CACHE_ROWS = rows
    return rows, path.name


async def _resolve_crossdock_pickup_points_from_env(pool: Optional[asyncpg.Pool]) -> List[str]:
    explicit_zones_raw = (
        os.getenv("SUPPLY_CROSSDOCK_TARIFF_ZONES")
        or os.getenv("SUPPLY_CROSSDOCK_TARIFF_ZONE")
        or _get_env_from_dotenv("SUPPLY_CROSSDOCK_TARIFF_ZONES")
        or _get_env_from_dotenv("SUPPLY_CROSSDOCK_TARIFF_ZONE")
        or ""
    ).strip()
    if explicit_zones_raw:
        explicit_zones = [z.strip() for z in explicit_zones_raw.split(",") if z.strip()]
        if explicit_zones:
            return explicit_zones

    # По умолчанию считаем тариф для основной зоны отгрузки.
    # Это убирает усреднение по всем зонам, если нет явного соответствия.
    return ["Москва и МО"]


async def _load_cluster_markup_tariffs(pool: Optional[asyncpg.Pool]) -> Dict[str, float]:
    if pool is None:
        return {}
    try:
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT cluster_name_norm, markup_pct
                FROM supply_cluster_markup_tariffs
                """
            )
    except Exception:
        return {}
    result: Dict[str, float] = {}
    for row in rows:
        key = _normalize_cluster_name(str(row.get("cluster_name_norm") or ""))
        if not key:
            continue
        result[key] = float(as_float(row.get("markup_pct"), default=0.0) or 0.0)
    return result


def _select_crossdock_tariff_for_cluster(
    cluster_name: str,
    tariffs: List[Dict[str, Any]],
    pickup_points: List[str],
) -> Optional[Dict[str, Any]]:
    cluster_norm = _normalize_text_key(cluster_name)
    if not cluster_norm:
        return None
    matches = [row for row in tariffs if row.get("recipient_cluster_norm") == cluster_norm]
    if not matches:
        return None

    pickup_norms = {_normalize_text_key(name) for name in pickup_points if name}
    pickup_matches = []
    if pickup_norms:
        pickup_matches = [
            row
            for row in matches
            if (row.get("tariff_zone_norm") in pickup_norms) or (row.get("pickup_point_norm") in pickup_norms)
        ]
    target_rows = pickup_matches if pickup_matches else matches
    if not target_rows:
        return None

    shipment_rate = sum(float(row.get("shipment_pallet") or 0.0) for row in target_rows) / len(target_rows)
    transport_rate = sum(float(row.get("transport_pallet") or 0.0) for row in target_rows) / len(target_rows)
    total_rate = shipment_rate + transport_rate
    lead_times = sorted({str(row.get("lead_time") or "").strip() for row in target_rows if row.get("lead_time")})
    return {
        "shipment_rate_pallet": float(shipment_rate),
        "transport_rate_pallet": float(transport_rate),
        "total_rate_pallet": float(total_rate),
        "lead_time": lead_times[0] if lead_times else "",
        "rows_used": len(target_rows),
        "pickup_matched": bool(pickup_matches),
    }


def _estimate_crossdock_costs_for_pallet_clusters(
    clusters: List[Dict[str, Any]],
    tariffs: List[Dict[str, Any]],
    pickup_points: List[str],
    cluster_markup_map: Optional[Dict[str, float]] = None,
) -> None:
    pallet_base_area_m2 = 1.2 * 0.8
    liters_per_meter = pallet_base_area_m2 * 1000.0
    min_pallet_liters_raw = os.getenv("SUPPLY_CROSSDOCK_MIN_PALLET_LITERS", "1900").strip()
    try:
        min_pallet_liters = max(0.0, float(min_pallet_liters_raw))
    except Exception:
        min_pallet_liters = 1900.0
    for cluster in clusters:
        cluster_name = str(cluster.get("cluster") or cluster.get("cluster_name") or "").strip()
        cluster_norm = _normalize_cluster_name(cluster_name)
        cluster_markup_pct = float((cluster_markup_map or {}).get(cluster_norm, 0.0) or 0.0)
        tariff = _select_crossdock_tariff_for_cluster(cluster_name, tariffs, pickup_points)
        if not tariff:
            cluster["crossdock_estimate"] = {
                "available": False,
                "reason": "tariff_not_found",
                "cluster_markup_pct": float(cluster_markup_pct),
            }
            continue

        sku_liters: Dict[str, float] = {}
        sku_billed_liters: Dict[str, float] = {}
        total_billed_liters = 0.0
        for pallet in cluster.get("pallets") or []:
            pallet_items = list(pallet.get("items") or [])
            pallet_actual_liters = 0.0
            pallet_item_liters: List[Tuple[Dict[str, Any], str, float]] = []
            for item in pallet_items:
                sku = str(item.get("sku") or "").strip()
                if not sku:
                    continue
                item_height_m = float(as_float(item.get("height"), default=0.0) or 0.0)
                item_liters = max(0.0, item_height_m * liters_per_meter)
                item["crossdock_liters"] = float(item_liters)
                sku_liters[sku] = sku_liters.get(sku, 0.0) + item_liters
                pallet_actual_liters += item_liters
                pallet_item_liters.append((item, sku, item_liters))

            if not pallet_item_liters:
                continue

            pallet_billed_liters = max(min_pallet_liters, pallet_actual_liters)
            total_billed_liters += pallet_billed_liters

            if pallet_actual_liters > 0:
                for item, sku, item_liters in pallet_item_liters:
                    billed_liters = pallet_billed_liters * (item_liters / pallet_actual_liters)
                    item["crossdock_billed_liters"] = float(billed_liters)
                    sku_billed_liters[sku] = sku_billed_liters.get(sku, 0.0) + billed_liters
            else:
                equal_share = pallet_billed_liters / len(pallet_item_liters)
                for item, sku, _item_liters in pallet_item_liters:
                    item["crossdock_billed_liters"] = float(equal_share)
                    sku_billed_liters[sku] = sku_billed_liters.get(sku, 0.0) + equal_share

        total_liters = float(sum(sku_liters.values()))
        rate_total = float(tariff.get("total_rate_pallet") or 0.0)
        total_cost = float(total_billed_liters * rate_total)

        sku_costs: Dict[str, float] = {}
        if total_billed_liters > 0:
            for sku, billed_liters in sku_billed_liters.items():
                sku_costs[sku] = total_cost * (billed_liters / total_billed_liters)
        else:
            for sku in sku_billed_liters:
                sku_costs[sku] = 0.0

        for pallet in cluster.get("pallets") or []:
            for item in pallet.get("items") or []:
                sku = str(item.get("sku") or "").strip()
                item_billed_liters = float(as_float(item.get("crossdock_billed_liters"), default=0.0) or 0.0)
                sku_total_liters = float(sku_billed_liters.get(sku, 0.0))
                sku_total_cost = float(sku_costs.get(sku, 0.0))
                item["crossdock_cost"] = float(
                    sku_total_cost * (item_billed_liters / sku_total_liters) if sku_total_liters > 0 else 0.0
                )
                item["crossdock_rate_per_liter"] = rate_total
                item["cluster_markup_pct"] = float(cluster_markup_pct)

        cluster["crossdock_estimate"] = {
            "available": True,
            "shipment_rate_pallet": float(tariff["shipment_rate_pallet"]),
            "transport_rate_pallet": float(tariff["transport_rate_pallet"]),
            "total_rate_pallet": rate_total,
            "total_liters": total_liters,
            "total_billed_liters": float(total_billed_liters),
            "min_pallet_liters": float(min_pallet_liters),
            "total_cost": total_cost,
            "cluster_markup_pct": float(cluster_markup_pct),
            "lead_time": tariff.get("lead_time") or "",
            "rows_used": int(tariff.get("rows_used") or 0),
            "pickup_matched": bool(tariff.get("pickup_matched")),
        }


def _load_supply_targets_from_availability_bytes(content: bytes, source_name: str) -> Tuple[Dict[str, Dict[str, int]], Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    return _extract_supply_targets_from_availability_workbook(wb, source_name)


def _extract_supply_targets_from_availability_workbook(
    wb: Any, source_name: str
) -> Tuple[Dict[str, Dict[str, int]], Dict[str, Any]]:
    ws = wb[wb.sheetnames[0]]

    header_row = None
    max_scan_rows = min(ws.max_row, 30)
    for row_idx in range(1, max_scan_rows + 1):
        row_values = [str(v or "").strip().lower() for v in next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))]
        if "sku" in row_values and any("артикул" in v for v in row_values):
            header_row = row_idx
            break
    if header_row is None:
        raise ValueError(
            f"Неверная структура файла '{source_name}': не найдена строка заголовка (ожидаются колонки SKU/Артикул)."
        )

    header_cells = [str(v or "").strip().lower() for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    article_col = None
    cluster_col = None
    supply_col = None
    for idx, name in enumerate(header_cells, start=1):
        if article_col is None and "артикул" in name:
            article_col = idx
        if cluster_col is None and "кластер" in name:
            cluster_col = idx
        if supply_col is None and "рекомендуемая поставка" in name and "28" in name:
            supply_col = idx
    if not article_col or not cluster_col or not supply_col:
        found_headers = [h for h in header_cells if h]
        found_preview = ", ".join(found_headers[:10]) if found_headers else "пусто"
        raise ValueError(
            "Неверная структура XLSX: нужны колонки "
            "'Артикул', 'Кластер', 'Рекомендуемая поставка ... 28'. "
            f"Найдено: {found_preview}"
        )

    targets: Dict[str, Dict[str, int]] = {}
    matched_rows = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        article_raw = row[article_col - 1] if len(row) >= article_col else None
        cluster_raw = row[cluster_col - 1] if len(row) >= cluster_col else None
        supply_raw = row[supply_col - 1] if len(row) >= supply_col else None
        article_key = normalize_offer_id(article_raw)
        cluster_key = _normalize_cluster_name(str(cluster_raw or ""))
        if not article_key or not cluster_key:
            continue
        supply_qty = int(max(0, as_float(supply_raw, default=0.0)))
        targets.setdefault(article_key, {})[cluster_key] = supply_qty
        matched_rows += 1

    if matched_rows == 0:
        raise ValueError(
            "Структура файла распознана, но не найдено ни одной строки данных "
            "с заполненными полями 'Артикул' и 'Кластер'."
        )

    validation = {
        "source": source_name,
        "sheet": ws.title,
        "header_row": header_row,
        "article_column": article_col,
        "cluster_column": cluster_col,
        "supply_column": supply_col,
        "matched_rows": matched_rows,
    }
    return targets, validation


async def _load_crossdock_dropoff_candidates_from_db(
    pool: Optional[asyncpg.Pool],
    source_macrolocal_cluster_id: int = 4039,
    limit: int = 12,
) -> List[Dict[str, Any]]:
    if pool is None:
        return []
    safe_limit = max(1, min(200, int(limit)))
    try:
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT warehouse_id, warehouse_name
                FROM ozon_cluster_warehouses
                WHERE macrolocal_cluster_id = $1
                  AND upper(coalesce(warehouse_type, '')) = 'SORTING_CENTER'
                GROUP BY warehouse_id, warehouse_name
                ORDER BY warehouse_id
                LIMIT $2
                """,
                int(source_macrolocal_cluster_id),
                safe_limit,
            )
    except Exception:
        return []
    result: List[Dict[str, Any]] = []
    for row in rows:
        warehouse_id = _to_int(row.get("warehouse_id"))
        if not warehouse_id:
            continue
        result.append(
            {
                "warehouse_id": warehouse_id,
                "warehouse_name": str(row.get("warehouse_name") or "").strip() or None,
            }
        )
    return result


async def _load_warehouse_names_from_db(
    pool: Optional[asyncpg.Pool],
    warehouse_ids: List[int],
) -> Dict[int, str]:
    if pool is None:
        return {}
    cleaned_ids = [wid for wid in {int(w) for w in warehouse_ids if _to_int(w)} if wid > 0]
    if not cleaned_ids:
        return {}
    try:
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT warehouse_id, max(warehouse_name) AS warehouse_name
                FROM ozon_cluster_warehouses
                WHERE warehouse_id = ANY($1::bigint[])
                GROUP BY warehouse_id
                """,
                cleaned_ids,
            )
    except Exception:
        return {}
    names: Dict[int, str] = {}
    for row in rows:
        wid = _to_int(row.get("warehouse_id"))
        name = str(row.get("warehouse_name") or "").strip()
        if wid and name:
            names[wid] = name
    return names


async def build_supply_acceptance_report(
    supply_items: List[Dict[str, Any]],
    fetch_timeslots: bool = False,
    pool: Optional[asyncpg.Pool] = None,
    use_multi_cluster_draft: bool = False,
) -> Dict[str, Any]:
    client_id = (
        os.getenv("OZON_CLIENT_ID")
        or getattr(settings, "ozon_client_id", "")
        or _get_env_from_dotenv("OZON_CLIENT_ID")
        or ""
    ).strip()
    api_key = (
        os.getenv("OZON_SUPPLY_API_KEY")
        or os.getenv("OZON_API_KEY")
        or _get_env_from_dotenv("OZON_SUPPLY_API_KEY")
        or _get_env_from_dotenv("OZON_API_KEY")
        or getattr(settings, "ozon_api_key", "")
        or ""
    ).strip()

    if not client_id or not api_key:
        return {
            "success": False,
            "error": "Missing OZON_CLIENT_ID/OZON_SUPPLY_API_KEY",
            "clusters": [],
        }

    source_clusters = _extract_supply_clusters(supply_items)
    if not source_clusters:
        return {"success": True, "clusters": [], "note": "No allocated supply items"}

    cache_ttl_raw = os.getenv("SUPPLY_ACCEPTANCE_CACHE_TTL_SEC", "900").strip()
    try:
        cache_ttl_sec = max(0, int(cache_ttl_raw))
    except Exception:
        cache_ttl_sec = 900

    # Включаем crossdock env-параметры в cache key, чтобы после их изменения
    # не возвращать устаревший результат из кэша.
    crossdock_cache_key = {
        "default_seller": (
            os.getenv("OZON_CROSSDOCK_SELLER_WAREHOUSE_ID")
            or _get_env_from_dotenv("OZON_CROSSDOCK_SELLER_WAREHOUSE_ID")
            or ""
        ).strip(),
        "default_dropoff": (
            os.getenv("OZON_CROSSDOCK_DROPOFF_WAREHOUSE_ID")
            or _get_env_from_dotenv("OZON_CROSSDOCK_DROPOFF_WAREHOUSE_ID")
            or ""
        ).strip(),
        "spb_seller": (
            os.getenv("OZON_CROSSDOCK_SPB_SELLER_WAREHOUSE_ID")
            or _get_env_from_dotenv("OZON_CROSSDOCK_SPB_SELLER_WAREHOUSE_ID")
            or ""
        ).strip(),
        "spb_dropoff": (
            os.getenv("OZON_CROSSDOCK_SPB_DROPOFF_WAREHOUSE_ID")
            or _get_env_from_dotenv("OZON_CROSSDOCK_SPB_DROPOFF_WAREHOUSE_ID")
            or ""
        ).strip(),
    }

    cache_key_payload = {
        "fetch_timeslots": bool(fetch_timeslots),
        "use_multi_cluster_draft": bool(use_multi_cluster_draft),
        "crossdock_env": crossdock_cache_key,
        "clusters": [
            {
                "cluster_name": c["cluster_name"],
                "allocated_total": c["allocated_total"],
                "skus": c["skus"],
            }
            for c in source_clusters
        ],
    }
    cache_key = json.dumps(cache_key_payload, ensure_ascii=False, sort_keys=True)
    now_utc = datetime.now(timezone.utc)
    if cache_ttl_sec > 0:
        cached = state.SUPPLY_ACCEPTANCE_CACHE.get(cache_key)
        if cached:
            cached_at = cached.get("created_at")
            cached_data = cached.get("data")
            if isinstance(cached_at, datetime) and isinstance(cached_data, dict):
                age_sec = (now_utc - cached_at).total_seconds()
                if 0 <= age_sec <= cache_ttl_sec:
                    cloned = json.loads(json.dumps(cached_data, ensure_ascii=False))
                    cloned["from_cache"] = True
                    return cloned

    max_clusters_raw = os.getenv("SUPPLY_REPORT_MAX_CLUSTERS", "10").strip()
    try:
        max_clusters = max(1, int(max_clusters_raw))
    except Exception:
        max_clusters = 10
    selected_clusters = source_clusters[:max_clusters]
    clusters_truncated = len(source_clusters) > max_clusters
    unknown_cluster_fallback_raw = os.getenv("SUPPLY_UNKNOWN_CLUSTER_FALLBACK_MACROLOCAL", "4039").strip()
    try:
        unknown_cluster_fallback_macrolocal = _to_int(unknown_cluster_fallback_raw) or 4039
    except Exception:
        unknown_cluster_fallback_macrolocal = 4039

    # Определяем KGT-политику по SKU из article_characteristics:
    # если все SKU кластера явно non-KGT, то исключаем негабаритные склады.
    cluster_kgt_policy: Dict[str, Dict[str, Any]] = {}
    if pool is not None and selected_clusters:
        all_skus = sorted(
            {
                _to_int(sku)
                for cluster in selected_clusters
                for sku in (cluster.get("skus") or [])
                if _to_int(sku)
            }
        )
        sku_is_kgt: Dict[int, Optional[bool]] = {}
        if all_skus:
            try:
                async with pool.acquire() as conn:
                    rows = await conn.fetch(
                        """
                        SELECT sku, is_kgt
                        FROM article_characteristics
                        WHERE sku = ANY($1::bigint[])
                        """,
                        all_skus,
                    )
                for row in rows:
                    sku = _to_int(row.get("sku"))
                    if sku:
                        value = row.get("is_kgt")
                        sku_is_kgt[sku] = bool(value) if value is not None else None
            except Exception:
                sku_is_kgt = {}

        for cluster in selected_clusters:
            cluster_name = str(cluster.get("cluster_name") or "")
            skus = [_to_int(sku) for sku in (cluster.get("skus") or []) if _to_int(sku)]
            if not skus:
                cluster_kgt_policy[cluster_name] = {
                    "known_skus": 0,
                    "total_skus": 0,
                    "has_kgt": False,
                    "all_known_non_kgt": False,
                }
                continue
            flags: List[Optional[bool]] = [sku_is_kgt.get(sku) for sku in skus]
            known_flags = [flag for flag in flags if flag is not None]
            has_kgt = any(flag is True for flag in known_flags)
            all_known_non_kgt = bool(known_flags) and len(known_flags) == len(flags) and not has_kgt
            cluster_kgt_policy[cluster_name] = {
                "known_skus": len(known_flags),
                "total_skus": len(flags),
                "has_kgt": has_kgt,
                "all_known_non_kgt": all_known_non_kgt,
            }

    # Для /v1/draft/crossdock/create обязательна delivery_info.
    default_crossdock_seller_wh = _to_int(
        os.getenv("OZON_CROSSDOCK_SELLER_WAREHOUSE_ID")
        or _get_env_from_dotenv("OZON_CROSSDOCK_SELLER_WAREHOUSE_ID")
        or 23785825652000
    )
    default_crossdock_dropoff_wh = _to_int(
        os.getenv("OZON_CROSSDOCK_DROPOFF_WAREHOUSE_ID")
        or _get_env_from_dotenv("OZON_CROSSDOCK_DROPOFF_WAREHOUSE_ID")
        or 23969023230000
    )
    spb_crossdock_seller_wh = _to_int(
        os.getenv("OZON_CROSSDOCK_SPB_SELLER_WAREHOUSE_ID")
        or _get_env_from_dotenv("OZON_CROSSDOCK_SPB_SELLER_WAREHOUSE_ID")
        # Если SPB override не задан, используем базовый seller warehouse.
        or default_crossdock_seller_wh
    )
    spb_crossdock_dropoff_wh = _to_int(
        os.getenv("OZON_CROSSDOCK_SPB_DROPOFF_WAREHOUSE_ID")
        or _get_env_from_dotenv("OZON_CROSSDOCK_SPB_DROPOFF_WAREHOUSE_ID")
        # Проверенный валидный drop-off в СПб-кластере (не DOES_NOT_EXIST),
        # но слоты зависят от пары seller_warehouse_id + даты.
        or 1020001419077000
    )

    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }

    timeout = aiohttp.ClientTimeout(total=60)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        cluster_status, cluster_data = await _ozon_supply_post(
            session, "/v1/cluster/list", headers, {"cluster_type": 1}
        )
        if cluster_status != 200:
            return {
                "success": False,
                "error": "Failed to load cluster list",
                "status": cluster_status,
                "details": cluster_data,
                "clusters": [],
            }

        name_to_macrolocal: Dict[str, int] = {}
        normalized_name_to_macrolocal: Dict[str, int] = {}
        for cluster in cluster_data.get("clusters", []) or []:
            name = str(cluster.get("name") or "").strip()
            macrolocal_cluster_id = _to_int(cluster.get("macrolocal_cluster_id"))
            if name and macrolocal_cluster_id:
                name_to_macrolocal[name] = macrolocal_cluster_id
                normalized_name_to_macrolocal[_normalize_cluster_name(name)] = macrolocal_cluster_id

        date_from = datetime.now(timezone.utc).date()
        date_to = date_from + timedelta(days=6)

        # Экспериментальный быстрый путь: один multi-cluster draft на все кластеры.
        # Если что-то идет не так, ниже автоматически сработает обычный по-кластерный путь.
        if use_multi_cluster_draft and not fetch_timeslots and selected_clusters:
            prepared: List[Dict[str, Any]] = []
            early_errors: List[Dict[str, Any]] = []
            for source in selected_clusters:
                cluster_name = source["cluster_name"]
                normalized_cluster_name = _normalize_cluster_name(cluster_name)
                macrolocal_cluster_id = name_to_macrolocal.get(cluster_name) or normalized_name_to_macrolocal.get(
                    normalized_cluster_name
                )
                if not macrolocal_cluster_id and "моск" in normalized_cluster_name:
                    macrolocal_cluster_id = 4039
                if not macrolocal_cluster_id and (
                    ("санкт" in normalized_cluster_name and "петербург" in normalized_cluster_name)
                    or ("спб" in normalized_cluster_name)
                ):
                    macrolocal_cluster_id = 4007
                if not macrolocal_cluster_id:
                    macrolocal_cluster_id = SUPPLY_MACROLOCAL_CLUSTER_FALLBACKS.get(normalized_cluster_name)
                if not macrolocal_cluster_id:
                    macrolocal_cluster_id = unknown_cluster_fallback_macrolocal
                prepared.append(
                    {
                        "cluster_name": cluster_name,
                        "allocated_total": source["allocated_total"],
                        "skus": source["skus"],
                        "macrolocal_cluster_id": int(macrolocal_cluster_id),
                        "supply_type": "DIRECT" if "моск" in normalized_cluster_name else "CROSS_DOCK",
                    }
                )

            # Multi-cluster draft применяем только для DIRECT-кластеров.
            # Для CROSS_DOCK (например, СПб через Щербинку и другие drop-off варианты)
            # остается штатный по-кластерный путь с отдельными delivery_info attempts.
            can_use_multi = bool(prepared) and all(
                str(p.get("supply_type") or "").upper() == "DIRECT" for p in prepared
            )
            if can_use_multi:
                multi_body: Dict[str, Any] = {
                    "clusters_info": [
                        {
                            "macrolocal_cluster_id": int(p["macrolocal_cluster_id"]),
                            "items": [{"sku": int(p["skus"][0]), "quantity": 1}],
                        }
                        for p in prepared
                    ],
                    "deletion_sku_mode": "PARTIAL",
                    "delivery_info": {
                        "type": "DROPOFF",
                        "seller_warehouse_id": int(default_crossdock_seller_wh),
                        "drop_off_warehouse": {
                            "warehouse_id": int(default_crossdock_dropoff_wh),
                            "warehouse_type": "DELIVERY_POINT",
                        },
                    },
                }
                try:
                    multi_status, multi_data = await _ozon_supply_post(
                        session, "/v1/draft/multi-cluster/create", headers, multi_body
                    )
                    multi_draft_id = _to_int((multi_data or {}).get("draft_id")) or 0
                    if multi_status == 200 and multi_draft_id > 0:
                        info_data: Dict[str, Any] = {}
                        info_status = 0
                        for _ in range(10):
                            info_status, info_data = await _ozon_supply_post(
                                session, "/v2/draft/create/info", headers, {"draft_id": multi_draft_id}
                            )
                            if info_status == 200 and (info_data.get("status") or "") != "IN_PROGRESS":
                                break
                            await asyncio.sleep(1.0)

                        if info_status == 200:
                            info_clusters = info_data.get("clusters", []) or []
                            report_clusters: List[Dict[str, Any]] = list(early_errors)
                            for p in prepared:
                                candidate_warehouses: List[Dict[str, Any]] = []
                                for info_cluster in info_clusters:
                                    if _to_int(info_cluster.get("macrolocal_cluster_id")) != int(p["macrolocal_cluster_id"]):
                                        continue
                                    for wh in info_cluster.get("warehouses", []) or []:
                                        availability = wh.get("availability_status") or {}
                                        state = str(availability.get("state") or "").upper()
                                        storage = wh.get("storage_warehouse") or {}
                                        warehouse_id = _to_int(storage.get("warehouse_id"))
                                        if not warehouse_id:
                                            continue
                                        candidate_warehouses.append(
                                            {
                                                "warehouse_id": warehouse_id,
                                                "warehouse_name": storage.get("name"),
                                                "warehouse_address": storage.get("address"),
                                                "state": state,
                                                "total_rank": wh.get("total_rank"),
                                                "total_score": wh.get("total_score"),
                                                "invalid_reason": availability.get("invalid_reason"),
                                                "slots_by_date": [],
                                                "slots_count": 0,
                                                "first_slot": None,
                                            }
                                        )
                                accepted_warehouses = [
                                    warehouse
                                    for warehouse in candidate_warehouses
                                    if str(warehouse.get("state") or "").upper() in {"FULL_AVAILABLE", "AVAILABLE"}
                                ]
                                report_clusters.append(
                                    {
                                        "cluster_name": p["cluster_name"],
                                        "macrolocal_cluster_id": int(p["macrolocal_cluster_id"]),
                                        "allocated_total": int(p["allocated_total"]),
                                        "draft_id": int(multi_draft_id),
                                        "supply_type": p["supply_type"],
                                        "status": info_data.get("status"),
                                        "accepted_warehouses": accepted_warehouses,
                                        "accepted_count": len(accepted_warehouses),
                                        "delivery_info_attempts": [],
                                        "timeslot_candidates_count": len(candidate_warehouses),
                                        "timeslot_filtered_out_count": 0,
                                        "timeslot_filtered_out": [],
                                        "kgt_policy": {
                                            "skip_kgt_warehouses_direct": False,
                                            "known_skus": 0,
                                            "total_skus": 0,
                                            "has_kgt": False,
                                            "all_known_non_kgt": False,
                                        },
                                        "timeslots_truncated": False,
                                        "timeslots_requested": False,
                                        "timeslot_period": {
                                            "date_from": date_from.isoformat(),
                                            "date_to": date_to.isoformat(),
                                            "days": 7,
                                        },
                                    }
                                )

                            result = {
                                "success": True,
                                "generated_at": datetime.now(timezone.utc).isoformat(),
                                "clusters": report_clusters,
                                "clusters_truncated": clusters_truncated,
                                "from_cache": False,
                            }
                            if cache_ttl_sec > 0:
                                state.SUPPLY_ACCEPTANCE_CACHE[cache_key] = {"created_at": datetime.now(timezone.utc), "data": result}
                            return result
                except Exception:
                    pass

        report_clusters: List[Dict[str, Any]] = []
        draft_create_timestamps: List[datetime] = []
        for source in selected_clusters:
            cluster_name = source["cluster_name"]
            normalized_cluster_name = _normalize_cluster_name(cluster_name)
            is_moscow_cluster = "моск" in normalized_cluster_name
            supply_type = "DIRECT" if is_moscow_cluster else "CROSS_DOCK"
            macrolocal_cluster_id = name_to_macrolocal.get(cluster_name)
            if not macrolocal_cluster_id:
                macrolocal_cluster_id = normalized_name_to_macrolocal.get(
                    _normalize_cluster_name(cluster_name)
                )
            # Fallback для стандартного московского кластера, если имя не совпало из-за кодировки.
            if not macrolocal_cluster_id and "моск" in _normalize_cluster_name(cluster_name):
                macrolocal_cluster_id = 4039
            # Fallback для стандартного СПб кластера, если имя пришло в нестандартном формате
            # (например, "Санкт_Петербург_РФЦ", "СПБ РФЦ" и т.д.).
            normalized_for_fallback = _normalize_cluster_name(cluster_name)
            if not macrolocal_cluster_id and (
                ("санкт" in normalized_for_fallback and "петербург" in normalized_for_fallback)
                or ("спб" in normalized_for_fallback)
            ):
                macrolocal_cluster_id = 4007
            if not macrolocal_cluster_id:
                macrolocal_cluster_id = SUPPLY_MACROLOCAL_CLUSTER_FALLBACKS.get(normalized_for_fallback)
            if not macrolocal_cluster_id:
                for key, fallback_id in SUPPLY_MACROLOCAL_CLUSTER_FALLBACKS.items():
                    if key and (key in normalized_for_fallback or normalized_for_fallback in key):
                        macrolocal_cluster_id = fallback_id
                        break

            if not macrolocal_cluster_id:
                macrolocal_cluster_id = unknown_cluster_fallback_macrolocal

            draft_endpoint = "/v1/draft/direct/create" if supply_type == "DIRECT" else "/v1/draft/crossdock/create"
            draft_body: Dict[str, Any] = {
                "cluster_info": {
                    "macrolocal_cluster_id": macrolocal_cluster_id,
                    "items": [{"sku": int(source["skus"][0]), "quantity": 1}],
                },
                "deletion_sku_mode": "PARTIAL",
            }
            if supply_type == "CROSS_DOCK":
                is_spb_cluster = macrolocal_cluster_id == 4007
                crossdock_seller_wh = spb_crossdock_seller_wh if is_spb_cluster else default_crossdock_seller_wh
                crossdock_dropoff_wh = spb_crossdock_dropoff_wh if is_spb_cluster else default_crossdock_dropoff_wh
                if not crossdock_seller_wh or not crossdock_dropoff_wh:
                    report_clusters.append(
                        {
                            "cluster_name": cluster_name,
                            "macrolocal_cluster_id": macrolocal_cluster_id,
                            "allocated_total": source["allocated_total"],
                            "error": "Missing crossdock delivery_info warehouse ids",
                            "accepted_warehouses": [],
                        }
                    )
                    continue
                draft_body["delivery_info"] = {
                    "type": "DROPOFF",
                    "seller_warehouse_id": crossdock_seller_wh,
                    "drop_off_warehouse": {
                        "warehouse_id": crossdock_dropoff_wh,
                        "warehouse_type": "DELIVERY_POINT",
                    },
                }

            delivery_info_attempts: List[Optional[Dict[str, Any]]] = [draft_body.get("delivery_info")]
            dropoff_name_by_id: Dict[int, str] = {}
            use_crossdock_multi_drafts = False
            crossdock_success_drafts: List[Dict[str, Any]] = []
            if supply_type == "CROSS_DOCK":
                base_delivery_info = draft_body.get("delivery_info") or {}
                drop_off_wh = ((base_delivery_info.get("drop_off_warehouse") or {}).get("warehouse_id"))
                seller_wh = _to_int(base_delivery_info.get("seller_warehouse_id"))
                if drop_off_wh:
                    delivery_info_attempts = []
                    tried_keys: set = set()
                    seller_candidates: List[int] = []
                    if seller_wh is not None:
                        seller_candidates.append(int(seller_wh))
                    if 0 not in seller_candidates:
                        seller_candidates.append(0)

                    dropoff_candidates: List[Dict[str, Any]] = []
                    if macrolocal_cluster_id == 4007:
                        # Для СПб вначале пробуем московские сортировочные центры из локального справочника.
                        crossdock_dropoffs_limit_raw = os.getenv(
                            "SUPPLY_CROSSDOCK_SPB_DROPOFF_LIMIT", "3"
                        ).strip()
                        try:
                            crossdock_dropoffs_limit = max(1, int(crossdock_dropoffs_limit_raw))
                        except Exception:
                            crossdock_dropoffs_limit = 3
                        dropoff_candidates = await _load_crossdock_dropoff_candidates_from_db(
                            pool,
                            source_macrolocal_cluster_id=4039,
                            limit=crossdock_dropoffs_limit,
                        )
                        preferred_ids_raw = os.getenv(
                            "SUPPLY_CROSSDOCK_SPB_PREFERRED_DROPOFF_IDS",
                            "1020000996024000,1020001419077000",
                        ).strip()
                        preferred_ids: List[int] = []
                        for token in preferred_ids_raw.split(","):
                            parsed = _to_int(token.strip())
                            if parsed and parsed not in preferred_ids:
                                preferred_ids.append(parsed)
                        preferred_dropoff = _to_int(drop_off_wh)
                        if preferred_dropoff and preferred_dropoff not in preferred_ids:
                            preferred_ids.append(preferred_dropoff)
                        rank_map = {wid: idx for idx, wid in enumerate(preferred_ids)}
                        existing_ids = {
                            _to_int(item.get("warehouse_id")) or 0 for item in dropoff_candidates
                        }
                        preferred_names = await _load_warehouse_names_from_db(pool, preferred_ids)
                        for preferred_id in preferred_ids:
                            if preferred_id not in existing_ids:
                                dropoff_candidates.append(
                                    {
                                        "warehouse_id": preferred_id,
                                        "warehouse_name": preferred_names.get(preferred_id),
                                    }
                                )
                        dropoff_candidates.sort(
                            key=lambda item: (
                                rank_map.get(_to_int(item.get("warehouse_id")) or 0, 10**6),
                                _to_int(item.get("warehouse_id")) or 0,
                            )
                        )
                        dropoff_candidates = dropoff_candidates[:crossdock_dropoffs_limit]
                        if not fetch_timeslots:
                            use_crossdock_multi_drafts = True
                    if not dropoff_candidates:
                        dropoff_candidates = [
                            {
                                "warehouse_id": int(drop_off_wh),
                                "warehouse_name": None,
                            }
                        ]

                    warehouse_type_candidates = (
                        ["SORTING_CENTER"] if macrolocal_cluster_id == 4007 else ["DELIVERY_POINT", "SORTING_CENTER"]
                    )
                    for dropoff in dropoff_candidates:
                        candidate_dropoff_wh = _to_int(dropoff.get("warehouse_id"))
                        if not candidate_dropoff_wh:
                            continue
                        dropoff_name = str(dropoff.get("warehouse_name") or "").strip()
                        if dropoff_name:
                            dropoff_name_by_id[candidate_dropoff_wh] = dropoff_name
                        for candidate_seller_wh in seller_candidates:
                            for candidate_wh_type in warehouse_type_candidates:
                                key = (
                                    candidate_seller_wh,
                                    int(candidate_dropoff_wh),
                                    candidate_wh_type,
                                )
                                if key in tried_keys:
                                    continue
                                tried_keys.add(key)
                                delivery_info_attempts.append(
                                    {
                                        "type": "DROPOFF",
                                        "seller_warehouse_id": candidate_seller_wh,
                                        "drop_off_warehouse": {
                                            "warehouse_id": int(candidate_dropoff_wh),
                                            "warehouse_type": candidate_wh_type,
                                        },
                                    }
                                )

            info_poll_attempts_raw = os.getenv("SUPPLY_CREATE_INFO_POLL_ATTEMPTS", "8").strip()
            info_poll_delay_raw = os.getenv("SUPPLY_CREATE_INFO_POLL_DELAY_SEC", "1.0").strip()
            try:
                info_poll_attempts = max(0, int(info_poll_attempts_raw))
            except Exception:
                info_poll_attempts = 8
            try:
                info_poll_delay = max(0.1, float(info_poll_delay_raw))
            except Exception:
                info_poll_delay = 1.0

            selected_draft_id = 0
            selected_info_data: Dict[str, Any] = {}
            selected_info_status = 0
            selected_draft_body: Dict[str, Any] = {}
            delivery_attempt_logs: List[Dict[str, Any]] = []
            crossdock_success_limit_raw = os.getenv("SUPPLY_CROSSDOCK_SPB_SUCCESS_LIMIT", "1").strip()
            try:
                crossdock_success_limit = max(1, int(crossdock_success_limit_raw))
            except Exception:
                crossdock_success_limit = 1

            for attempt_index, delivery_info_candidate in enumerate(delivery_info_attempts, start=1):
                attempt_draft_body: Dict[str, Any] = dict(draft_body)
                if delivery_info_candidate:
                    attempt_draft_body["delivery_info"] = delivery_info_candidate
                else:
                    attempt_draft_body.pop("delivery_info", None)

                # Учитываем лимит Ozon: создание черновиков 2 раза в минуту.
                now_utc = datetime.now(timezone.utc)
                draft_create_timestamps = [
                    ts for ts in draft_create_timestamps if (now_utc - ts).total_seconds() < 60.0
                ]
                if len(draft_create_timestamps) >= 2:
                    wait_seconds = 60.0 - (now_utc - draft_create_timestamps[0]).total_seconds() + 0.25
                    if wait_seconds > 0:
                        await asyncio.sleep(wait_seconds)

                draft_status, draft_data = await _ozon_supply_post(
                    session,
                    draft_endpoint,
                    headers,
                    attempt_draft_body,
                    retries=6,
                    retry_delay_seconds=32.0,
                )
                draft_create_timestamps.append(datetime.now(timezone.utc))
                draft_id = _to_int(draft_data.get("draft_id")) or 0
                if draft_status != 200 or draft_id <= 0:
                    delivery_attempt_logs.append(
                        {
                            "attempt": attempt_index,
                            "delivery_info": attempt_draft_body.get("delivery_info"),
                            "draft_status": draft_status,
                            "draft_response": draft_data,
                            "error": "Failed to create draft",
                        }
                    )
                    continue

                info_status, info_data = await _ozon_supply_post(
                    session, "/v2/draft/create/info", headers, {"draft_id": draft_id}
                )
                if info_status != 200:
                    delivery_attempt_logs.append(
                        {
                            "attempt": attempt_index,
                            "draft_id": draft_id,
                            "delivery_info": attempt_draft_body.get("delivery_info"),
                            "info_status": info_status,
                            "info_response": info_data,
                            "error": "Failed to get draft info",
                        }
                    )
                    continue

                for _ in range(info_poll_attempts):
                    info_state = str((info_data or {}).get("status") or "").upper()
                    if info_state not in {"IN_PROGRESS", "PENDING", "PROCESSING"}:
                        break
                    await asyncio.sleep(info_poll_delay)
                    info_status, info_data = await _ozon_supply_post(
                        session, "/v2/draft/create/info", headers, {"draft_id": draft_id}
                    )
                    if info_status != 200:
                        break

                info_state = str((info_data or {}).get("status") or "").upper()
                info_errors = (info_data or {}).get("errors") or []
                first_error_message = ""
                if isinstance(info_errors, list) and info_errors:
                    first_error_message = str((info_errors[0] or {}).get("error_message") or "").strip()

                if info_state == "FAILED":
                    delivery_attempt_logs.append(
                        {
                            "attempt": attempt_index,
                            "draft_id": draft_id,
                            "delivery_info": attempt_draft_body.get("delivery_info"),
                            "info_status": info_status,
                            "status": info_data.get("status"),
                            "error_message": first_error_message,
                            "info_response": info_data,
                            "error": "Draft create info status FAILED",
                        }
                    )
                    continue

                if use_crossdock_multi_drafts:
                    dropoff_info = (attempt_draft_body.get("delivery_info") or {}).get("drop_off_warehouse") or {}
                    dropoff_id = _to_int(dropoff_info.get("warehouse_id"))
                    state = ""
                    info_clusters = (info_data or {}).get("clusters") or []
                    if isinstance(info_clusters, list) and info_clusters:
                        info_warehouses = (info_clusters[0] or {}).get("warehouses") or []
                        if isinstance(info_warehouses, list) and info_warehouses:
                            avail = (info_warehouses[0] or {}).get("availability_status") or {}
                            state = str(avail.get("state") or "").upper()
                    crossdock_success_drafts.append(
                        {
                            "draft_id": draft_id,
                            "delivery_info": attempt_draft_body.get("delivery_info"),
                            "dropoff_warehouse_id": dropoff_id,
                            "dropoff_warehouse_name": dropoff_name_by_id.get(dropoff_id or 0),
                            "state": state or "FULL_AVAILABLE",
                            "info_data": info_data or {},
                        }
                    )
                    delivery_attempt_logs.append(
                        {
                            "attempt": attempt_index,
                            "draft_id": draft_id,
                            "delivery_info": attempt_draft_body.get("delivery_info"),
                            "info_status": info_status,
                            "status": info_data.get("status"),
                            "error": "",
                        }
                    )
                    selected_draft_id = selected_draft_id or draft_id
                    selected_info_status = selected_info_status or info_status
                    selected_info_data = selected_info_data or (info_data or {})
                    selected_draft_body = selected_draft_body or attempt_draft_body
                    if len(crossdock_success_drafts) >= crossdock_success_limit:
                        break
                    continue

                selected_draft_id = draft_id
                selected_info_status = info_status
                selected_info_data = info_data or {}
                selected_draft_body = attempt_draft_body
                break

            if selected_draft_id <= 0:
                last_attempt = delivery_attempt_logs[-1] if delivery_attempt_logs else {}
                raw_error = str(last_attempt.get("error") or "Unable to create valid draft for cluster")
                first_error_message = str(last_attempt.get("error_message") or "").strip()
                ui_error = raw_error
                if raw_error == "Draft create info status FAILED" and first_error_message:
                    ui_error = f"{raw_error}: {first_error_message}"
                report_clusters.append(
                    {
                        "cluster_name": cluster_name,
                        "macrolocal_cluster_id": macrolocal_cluster_id,
                        "allocated_total": source["allocated_total"],
                        "supply_type": supply_type,
                        "error": ui_error,
                        "error_message": first_error_message,
                        "draft_endpoint": draft_endpoint,
                        "delivery_info": last_attempt.get("delivery_info") or draft_body.get("delivery_info"),
                        "delivery_info_attempts": delivery_attempt_logs,
                        "accepted_warehouses": [],
                        "accepted_count": 0,
                        "timeslots_truncated": False,
                        "timeslot_period": {
                            "date_from": date_from.isoformat(),
                            "date_to": date_to.isoformat(),
                            "days": 7,
                        },
                    }
                )
                continue

            draft_id = selected_draft_id
            info_status = selected_info_status
            info_data = selected_info_data
            draft_body = selected_draft_body

            candidate_warehouses: List[Dict[str, Any]] = []
            if use_crossdock_multi_drafts and crossdock_success_drafts:
                for success_item in crossdock_success_drafts:
                    wh_id = _to_int(success_item.get("dropoff_warehouse_id"))
                    if not wh_id:
                        continue
                    delivery_info = success_item.get("delivery_info") or {}
                    candidate_warehouses.append(
                        {
                            "warehouse_id": wh_id,
                            "warehouse_name": success_item.get("dropoff_warehouse_name")
                            or dropoff_name_by_id.get(wh_id)
                            or f"DROP_OFF_{wh_id}",
                            "warehouse_address": None,
                            "state": str(success_item.get("state") or "FULL_AVAILABLE").upper(),
                            "total_rank": None,
                            "total_score": None,
                            "invalid_reason": "UNSPECIFIED",
                            "slots_by_date": [],
                            "slots_count": 0,
                            "first_slot": None,
                            "draft_id": _to_int(success_item.get("draft_id")),
                            "delivery_info": delivery_info,
                        }
                    )
            else:
                candidate_index: Dict[int, Dict[str, Any]] = {}
                for info_cluster in info_data.get("clusters", []) or []:
                    if _to_int(info_cluster.get("macrolocal_cluster_id")) != macrolocal_cluster_id:
                        continue
                    for wh in info_cluster.get("warehouses", []) or []:
                        availability = wh.get("availability_status") or {}
                        state = str(availability.get("state") or "").upper()
                        storage = wh.get("storage_warehouse") or {}
                        warehouse_id = _to_int(storage.get("warehouse_id"))
                        if not warehouse_id:
                            continue
                        existing = candidate_index.get(warehouse_id)
                        if existing is None:
                            existing = {
                                "warehouse_id": warehouse_id,
                                "warehouse_name": storage.get("name"),
                                "warehouse_address": storage.get("address"),
                                "state": state,
                                "total_rank": wh.get("total_rank"),
                                "total_score": wh.get("total_score"),
                                "invalid_reason": availability.get("invalid_reason"),
                                "slots_by_date": [],
                                "slots_count": 0,
                                "first_slot": None,
                            }
                            candidate_index[warehouse_id] = existing
                            candidate_warehouses.append(existing)
                            continue

                        old_rank = existing.get("total_rank")
                        new_rank = wh.get("total_rank")
                        old_score = existing.get("total_score") or 0.0
                        new_score = wh.get("total_score") or 0.0

                        prefer_new = False
                        if old_rank is None and new_rank is not None:
                            prefer_new = True
                        elif old_rank is not None and new_rank is not None and new_rank < old_rank:
                            prefer_new = True
                        elif old_rank == new_rank and new_score > old_score:
                            prefer_new = True

                        if prefer_new:
                            existing["warehouse_name"] = storage.get("name")
                            existing["warehouse_address"] = storage.get("address")
                            existing["state"] = state
                            existing["total_rank"] = new_rank
                            existing["total_score"] = new_score
                            existing["invalid_reason"] = availability.get("invalid_reason")

            candidate_warehouses.sort(
                key=lambda x: (
                    x["total_rank"] if x["total_rank"] is not None else 10**9,
                    -(x["total_score"] or 0.0),
                )
            )

            # Фильтруем кандидатов для запроса таймслотов:
            # в DIRECT не тратим лимит на заведомо слабые склады (например, MATRIX),
            # а приоритизируем доступные и лучшие по рангу.
            direct_exclude_matrix_raw = os.getenv("SUPPLY_TIMESLOT_EXCLUDE_MATRIX_DIRECT", "1").strip()
            direct_include_rank_raw = os.getenv("SUPPLY_TIMESLOT_INCLUDE_NOT_AVAILABLE_RANK_DIRECT", "1").strip()
            direct_max_rank_raw = os.getenv("SUPPLY_TIMESLOT_MAX_RANK_DIRECT", "6").strip()
            try:
                direct_exclude_matrix = int(direct_exclude_matrix_raw) != 0
            except Exception:
                direct_exclude_matrix = True
            try:
                direct_include_not_available_rank = int(direct_include_rank_raw) != 0
            except Exception:
                direct_include_not_available_rank = True
            try:
                direct_max_rank = max(1, int(direct_max_rank_raw))
            except Exception:
                direct_max_rank = 6
            skip_kgt_raw = os.getenv("SUPPLY_TIMESLOT_SKIP_KGT_WAREHOUSES_DIRECT", "1").strip()
            exclude_name_patterns_raw = os.getenv("SUPPLY_TIMESLOT_EXCLUDE_NAME_PATTERNS_DIRECT", "").strip()
            include_name_patterns_raw = os.getenv("SUPPLY_TIMESLOT_INCLUDE_NAME_PATTERNS_DIRECT", "").strip()
            kgt_name_patterns_raw = os.getenv(
                "SUPPLY_TIMESLOT_KGT_NAME_PATTERNS_DIRECT", "кгт,негабарит,крупногабарит"
            ).strip()
            try:
                skip_kgt_warehouses_direct = int(skip_kgt_raw) != 0
            except Exception:
                skip_kgt_warehouses_direct = True
            exclude_name_patterns = [
                pattern.strip().lower() for pattern in exclude_name_patterns_raw.split(",") if pattern.strip()
            ]
            include_name_patterns = [
                pattern.strip().lower() for pattern in include_name_patterns_raw.split(",") if pattern.strip()
            ]
            kgt_name_patterns = [
                pattern.strip().lower() for pattern in kgt_name_patterns_raw.split(",") if pattern.strip()
            ]
            cluster_policy = cluster_kgt_policy.get(cluster_name) or {}
            effective_skip_kgt = skip_kgt_warehouses_direct
            if cluster_policy.get("has_kgt"):
                effective_skip_kgt = False
            elif cluster_policy.get("all_known_non_kgt"):
                effective_skip_kgt = True

            timeslot_candidates: List[Dict[str, Any]] = []
            filtered_out_warehouses: List[Dict[str, Any]] = []
            for warehouse in candidate_warehouses:
                state = str(warehouse.get("state") or "").upper()
                invalid_reason = str(warehouse.get("invalid_reason") or "").upper()
                rank = _to_int(warehouse.get("total_rank"))
                warehouse_name = str(warehouse.get("warehouse_name") or "")
                normalized_warehouse_name = warehouse_name.lower()
                skip_reason: Optional[str] = None

                if supply_type == "DIRECT":
                    if effective_skip_kgt and any(token in normalized_warehouse_name for token in kgt_name_patterns):
                        skip_reason = "DIRECT_OVERSIZE_WAREHOUSE_EXCLUDED"
                    elif include_name_patterns and not any(
                        pattern in normalized_warehouse_name for pattern in include_name_patterns
                    ):
                        skip_reason = "DIRECT_NAME_NOT_IN_INCLUDE_LIST"
                    elif exclude_name_patterns and any(
                        pattern in normalized_warehouse_name for pattern in exclude_name_patterns
                    ):
                        skip_reason = "DIRECT_NAME_IN_EXCLUDE_LIST"
                    if state in {"FULL_AVAILABLE", "AVAILABLE"}:
                        if skip_reason is None:
                            skip_reason = None
                    elif invalid_reason == "NOT_AVAILABLE_RANK":
                        if not direct_include_not_available_rank:
                            skip_reason = "DIRECT_NOT_AVAILABLE_RANK_DISABLED"
                        # Не отсекаем по rank: показываем все склады и используем rank только для сортировки.
                    elif invalid_reason == "NOT_AVAILABLE_MATRIX" and direct_exclude_matrix:
                        skip_reason = "DIRECT_NOT_AVAILABLE_MATRIX_EXCLUDED"

                if skip_reason:
                    filtered_out_warehouses.append(
                        {
                            "warehouse_id": warehouse.get("warehouse_id"),
                            "warehouse_name": warehouse.get("warehouse_name"),
                            "state": warehouse.get("state"),
                            "invalid_reason": warehouse.get("invalid_reason"),
                            "total_rank": warehouse.get("total_rank"),
                            "total_score": warehouse.get("total_score"),
                            "filter_reason": skip_reason,
                        }
                    )
                else:
                    timeslot_candidates.append(warehouse)

            timeslots_truncated = False
            if fetch_timeslots:
                timeslot_limit_raw = os.getenv("SUPPLY_TIMESLOT_WAREHOUSE_LIMIT", "0").strip()
                try:
                    max_warehouses_for_timeslots = int(timeslot_limit_raw)
                except Exception:
                    max_warehouses_for_timeslots = 0
                if max_warehouses_for_timeslots > 0:
                    warehouses_for_timeslots = timeslot_candidates[:max_warehouses_for_timeslots]
                else:
                    warehouses_for_timeslots = timeslot_candidates

                timeslot_parallel_raw = os.getenv("SUPPLY_TIMESLOT_PARALLEL", "2").strip()
                try:
                    timeslot_parallel = max(1, int(timeslot_parallel_raw))
                except Exception:
                    timeslot_parallel = 2
                timeslot_stagger_raw = os.getenv("SUPPLY_TIMESLOT_STAGGER_DELAY_SEC", "0.25").strip()
                try:
                    timeslot_stagger_sec = max(0.0, float(timeslot_stagger_raw))
                except Exception:
                    timeslot_stagger_sec = 0.25
                timeslot_retry_delay_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_DELAY_SEC", "2.5").strip()
                timeslot_retry_backoff_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_BACKOFF", "1.7").strip()
                timeslot_retry_max_delay_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_MAX_DELAY_SEC", "45").strip()
                try:
                    timeslot_retry_delay_sec = max(0.2, float(timeslot_retry_delay_raw))
                except Exception:
                    timeslot_retry_delay_sec = 2.5
                try:
                    timeslot_retry_backoff = max(1.0, float(timeslot_retry_backoff_raw))
                except Exception:
                    timeslot_retry_backoff = 1.7
                try:
                    timeslot_retry_max_delay_sec = max(timeslot_retry_delay_sec, float(timeslot_retry_max_delay_raw))
                except Exception:
                    timeslot_retry_max_delay_sec = 45.0
                timeslot_semaphore = asyncio.Semaphore(timeslot_parallel)

                async def _fill_timeslots_for_warehouse(warehouse: Dict[str, Any]) -> None:
                    async with timeslot_semaphore:
                        warehouse_idx = _to_int(warehouse.get("_timeslot_index")) or 0
                        if timeslot_stagger_sec > 0 and warehouse_idx > 0:
                            await asyncio.sleep(timeslot_stagger_sec * warehouse_idx)
                        local_draft_id = _to_int(warehouse.get("draft_id")) or draft_id
                        api_supply_type = "CROSSDOCK" if supply_type == "CROSS_DOCK" else supply_type
                        timeslot_status, timeslot_data = await _ozon_supply_post(
                            session,
                            "/v2/draft/timeslot/info",
                            headers,
                            {
                                "draft_id": local_draft_id,
                                "date_from": date_from.isoformat(),
                                "date_to": date_to.isoformat(),
                                "supply_type": api_supply_type,
                                "selected_cluster_warehouses": [
                                    (
                                        {
                                            "macrolocal_cluster_id": macrolocal_cluster_id,
                                            "storage_warehouse_id": warehouse["warehouse_id"],
                                        }
                                        if supply_type == "DIRECT"
                                        else {
                                            "macrolocal_cluster_id": macrolocal_cluster_id,
                                            "warehouse_ids": [warehouse["warehouse_id"]],
                                        }
                                    )
                                ],
                            },
                            retries=10,
                            retry_delay_seconds=timeslot_retry_delay_sec,
                            retry_backoff_factor=timeslot_retry_backoff,
                            retry_max_delay=timeslot_retry_max_delay_sec,
                        )

                        if timeslot_status != 200:
                            warehouse["timeslot_error"] = {
                                "status": timeslot_status,
                                "details": timeslot_data,
                            }
                            return

                        days = (
                            ((timeslot_data.get("result") or {}).get("drop_off_warehouse_timeslots") or {}).get("days")
                            or []
                        )
                        slots_by_date: List[Dict[str, Any]] = []
                        slots_count = 0
                        first_slot: Optional[Dict[str, str]] = None
                        for day in days:
                            date_label = day.get("date_in_timezone")
                            times = []
                            for slot in day.get("timeslots") or []:
                                slot_entry = {
                                    "from": slot.get("from_in_timezone"),
                                    "to": slot.get("to_in_timezone"),
                                }
                                times.append(slot_entry)
                                slots_count += 1
                                if first_slot is None:
                                    first_slot = {
                                        "date": date_label,
                                        "from": slot_entry["from"],
                                        "to": slot_entry["to"],
                                    }
                            if times:
                                slots_by_date.append({"date": date_label, "times": times})

                        warehouse["slots_by_date"] = slots_by_date
                        warehouse["slots_count"] = slots_count
                        warehouse["first_slot"] = first_slot

                if warehouses_for_timeslots:
                    for idx, warehouse in enumerate(warehouses_for_timeslots):
                        warehouse["_timeslot_index"] = idx
                    await asyncio.gather(*[_fill_timeslots_for_warehouse(warehouse) for warehouse in warehouses_for_timeslots])
                for warehouse in candidate_warehouses:
                    warehouse.pop("_timeslot_index", None)

                accepted_warehouses = [
                    warehouse
                    for warehouse in candidate_warehouses
                    if warehouse.get("slots_count", 0) > 0 or warehouse.get("state") in {"FULL_AVAILABLE", "AVAILABLE"}
                ]
                timeslots_truncated = max_warehouses_for_timeslots > 0 and len(timeslot_candidates) > max_warehouses_for_timeslots
            else:
                accepted_warehouses = [
                    warehouse
                    for warehouse in candidate_warehouses
                    if str(warehouse.get("state") or "").upper() in {"FULL_AVAILABLE", "AVAILABLE"}
                ]

            report_clusters.append(
                {
                    "cluster_name": cluster_name,
                    "macrolocal_cluster_id": macrolocal_cluster_id,
                    "allocated_total": source["allocated_total"],
                    "draft_id": draft_id,
                    "supply_type": supply_type,
                    "status": info_data.get("status"),
                    "accepted_warehouses": accepted_warehouses,
                    "accepted_count": len(accepted_warehouses),
                    "delivery_info_attempts": delivery_attempt_logs,
                    "timeslot_candidates_count": len(timeslot_candidates),
                    "timeslot_filtered_out_count": len(filtered_out_warehouses),
                    "timeslot_filtered_out": filtered_out_warehouses,
                    "kgt_policy": {
                        "skip_kgt_warehouses_direct": bool(effective_skip_kgt),
                        "known_skus": int(cluster_policy.get("known_skus") or 0),
                        "total_skus": int(cluster_policy.get("total_skus") or 0),
                        "has_kgt": bool(cluster_policy.get("has_kgt")),
                        "all_known_non_kgt": bool(cluster_policy.get("all_known_non_kgt")),
                    },
                    "timeslots_truncated": timeslots_truncated,
                    "timeslots_requested": bool(fetch_timeslots),
                    "timeslot_period": {
                        "date_from": date_from.isoformat(),
                        "date_to": date_to.isoformat(),
                        "days": 7,
                    },
                }
            )

    result = {
        "success": True,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "clusters": report_clusters,
        "clusters_truncated": clusters_truncated,
        "from_cache": False,
    }
    if cache_ttl_sec > 0:
        state.SUPPLY_ACCEPTANCE_CACHE[cache_key] = {"created_at": datetime.now(timezone.utc), "data": result}
        if len(state.SUPPLY_ACCEPTANCE_CACHE) > SUPPLY_ACCEPTANCE_CACHE_MAX_ENTRIES:
            oldest = sorted(
                state.SUPPLY_ACCEPTANCE_CACHE.items(),
                key=lambda kv: kv[1].get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
            )
            for old_key, _ in oldest[: max(0, len(state.SUPPLY_ACCEPTANCE_CACHE) - SUPPLY_ACCEPTANCE_CACHE_MAX_ENTRIES)]:
                state.SUPPLY_ACCEPTANCE_CACHE.pop(old_key, None)
    return result


async def request_supply_plan_timeslots(request: web.Request) -> web.Response:
    body = await request.json() if request.body_exists else {}
    clusters = body.get("clusters") if isinstance(body.get("clusters"), list) else []
    selected_by_draft = body.get("selected_by_draft") if isinstance(body.get("selected_by_draft"), dict) else {}

    if not clusters:
        return web.json_response({"success": False, "error": "clusters is required"}, status=400)

    client_id = (
        os.getenv("OZON_CLIENT_ID")
        or getattr(settings, "ozon_client_id", "")
        or _get_env_from_dotenv("OZON_CLIENT_ID")
        or ""
    ).strip()
    api_key = (
        os.getenv("OZON_SUPPLY_API_KEY")
        or os.getenv("OZON_API_KEY")
        or _get_env_from_dotenv("OZON_SUPPLY_API_KEY")
        or _get_env_from_dotenv("OZON_API_KEY")
        or getattr(settings, "ozon_api_key", "")
        or ""
    ).strip()
    if not client_id or not api_key:
        return web.json_response({"success": False, "error": "Missing OZON_CLIENT_ID/OZON_SUPPLY_API_KEY"}, status=400)

    normalized_selected: Dict[str, set[int]] = {}
    any_checked = False
    for draft_key, warehouse_ids_raw in selected_by_draft.items():
        if not isinstance(warehouse_ids_raw, list):
            continue
        cleaned_ids: set[int] = set()
        for wid in warehouse_ids_raw:
            parsed = _to_int(wid)
            if parsed and parsed > 0:
                cleaned_ids.add(parsed)
        if cleaned_ids:
            any_checked = True
        normalized_selected[str(draft_key)] = cleaned_ids

    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }
    timeout = aiohttp.ClientTimeout(total=90)

    timeslot_parallel_raw = os.getenv("SUPPLY_TIMESLOT_PARALLEL", "2").strip()
    try:
        timeslot_parallel = max(1, int(timeslot_parallel_raw))
    except Exception:
        timeslot_parallel = 2
    timeslot_stagger_raw = os.getenv("SUPPLY_TIMESLOT_STAGGER_DELAY_SEC", "0.25").strip()
    try:
        timeslot_stagger_sec = max(0.0, float(timeslot_stagger_raw))
    except Exception:
        timeslot_stagger_sec = 0.25
    timeslot_retry_delay_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_DELAY_SEC", "2.5").strip()
    timeslot_retry_backoff_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_BACKOFF", "1.7").strip()
    timeslot_retry_max_delay_raw = os.getenv("SUPPLY_TIMESLOT_RETRY_MAX_DELAY_SEC", "45").strip()
    try:
        timeslot_retry_delay_sec = max(0.2, float(timeslot_retry_delay_raw))
    except Exception:
        timeslot_retry_delay_sec = 2.5
    try:
        timeslot_retry_backoff = max(1.0, float(timeslot_retry_backoff_raw))
    except Exception:
        timeslot_retry_backoff = 1.7
    try:
        timeslot_retry_max_delay_sec = max(timeslot_retry_delay_sec, float(timeslot_retry_max_delay_raw))
    except Exception:
        timeslot_retry_max_delay_sec = 45.0

    result_clusters: List[Dict[str, Any]] = []
    total_requested = 0

    async with aiohttp.ClientSession(timeout=timeout) as session:
        for cluster in clusters:
            if not isinstance(cluster, dict):
                continue
            cluster_copy = json.loads(json.dumps(cluster, ensure_ascii=False))
            draft_id = _to_int(cluster_copy.get("draft_id"))
            macrolocal_cluster_id = _to_int(cluster_copy.get("macrolocal_cluster_id"))
            supply_type = str(cluster_copy.get("supply_type") or "").upper()
            accepted = cluster_copy.get("accepted_warehouses") or []
            if not isinstance(accepted, list):
                accepted = []

            if not draft_id or not macrolocal_cluster_id or supply_type not in {"DIRECT", "CROSS_DOCK"}:
                result_clusters.append(cluster_copy)
                continue

            selected_for_draft = normalized_selected.get(str(draft_id), set())
            if any_checked:
                target_warehouses = [
                    wh for wh in accepted
                    if _to_int((wh or {}).get("warehouse_id")) in selected_for_draft
                ]
            else:
                target_warehouses = accepted

            date_from = str(((cluster_copy.get("timeslot_period") or {}).get("date_from") or "")).strip()
            date_to = str(((cluster_copy.get("timeslot_period") or {}).get("date_to") or "")).strip()
            if not date_from or not date_to:
                now_date = datetime.now(timezone.utc).date()
                date_from = now_date.isoformat()
                date_to = (now_date + timedelta(days=6)).isoformat()

            semaphore = asyncio.Semaphore(timeslot_parallel)
            parsed_targets: List[Dict[str, Any]] = []
            for idx, warehouse in enumerate(target_warehouses):
                if not isinstance(warehouse, dict):
                    continue
                parsed = dict(warehouse)
                parsed["_timeslot_index"] = idx
                parsed_targets.append(parsed)

            async def _fill_timeslots_for_selected(warehouse: Dict[str, Any]) -> None:
                async with semaphore:
                    warehouse_idx = _to_int(warehouse.get("_timeslot_index")) or 0
                    if timeslot_stagger_sec > 0 and warehouse_idx > 0:
                        await asyncio.sleep(timeslot_stagger_sec * warehouse_idx)
                    wh_id = _to_int(warehouse.get("warehouse_id"))
                    if not wh_id:
                        return
                    local_draft_id = _to_int(warehouse.get("draft_id")) or draft_id
                    api_supply_type = "CROSSDOCK" if supply_type == "CROSS_DOCK" else supply_type
                    timeslot_status, timeslot_data = await _ozon_supply_post(
                        session,
                        "/v2/draft/timeslot/info",
                        headers,
                        {
                            "draft_id": local_draft_id,
                            "date_from": date_from,
                            "date_to": date_to,
                            "supply_type": api_supply_type,
                            "selected_cluster_warehouses": [
                                (
                                    {
                                        "macrolocal_cluster_id": macrolocal_cluster_id,
                                        "storage_warehouse_id": wh_id,
                                    }
                                    if supply_type == "DIRECT"
                                    else {
                                        "macrolocal_cluster_id": macrolocal_cluster_id,
                                        "warehouse_ids": [wh_id],
                                    }
                                )
                            ],
                        },
                        retries=10,
                        retry_delay_seconds=timeslot_retry_delay_sec,
                        retry_backoff_factor=timeslot_retry_backoff,
                        retry_max_delay=timeslot_retry_max_delay_sec,
                    )

                    if timeslot_status != 200:
                        warehouse["timeslot_error"] = {
                            "status": timeslot_status,
                            "details": timeslot_data,
                        }
                        return

                    days = (
                        ((timeslot_data.get("result") or {}).get("drop_off_warehouse_timeslots") or {}).get("days")
                        or []
                    )
                    slots_by_date: List[Dict[str, Any]] = []
                    slots_count = 0
                    first_slot: Optional[Dict[str, str]] = None
                    for day in days:
                        date_label = day.get("date_in_timezone")
                        times = []
                        for slot in day.get("timeslots") or []:
                            slot_entry = {
                                "from": slot.get("from_in_timezone"),
                                "to": slot.get("to_in_timezone"),
                            }
                            times.append(slot_entry)
                            slots_count += 1
                            if first_slot is None:
                                first_slot = {
                                    "date": date_label,
                                    "from": slot_entry["from"],
                                    "to": slot_entry["to"],
                                }
                        if times:
                            slots_by_date.append({"date": date_label, "times": times})

                    warehouse["slots_by_date"] = slots_by_date
                    warehouse["slots_count"] = slots_count
                    warehouse["first_slot"] = first_slot
                    warehouse.pop("timeslot_error", None)

            if parsed_targets:
                await asyncio.gather(*[_fill_timeslots_for_selected(warehouse) for warehouse in parsed_targets])
                total_requested += len(parsed_targets)

            timeslot_by_id = {
                _to_int(item.get("warehouse_id")): item
                for item in parsed_targets
                if _to_int(item.get("warehouse_id"))
            }
            updated_accepted: List[Dict[str, Any]] = []
            for warehouse in accepted:
                wh_id = _to_int((warehouse or {}).get("warehouse_id"))
                if wh_id and wh_id in timeslot_by_id:
                    enriched = dict(timeslot_by_id[wh_id])
                    enriched.pop("_timeslot_index", None)
                    updated_accepted.append(enriched)
                else:
                    base = dict(warehouse) if isinstance(warehouse, dict) else {}
                    base.setdefault("slots_by_date", [])
                    base.setdefault("slots_count", 0)
                    base.setdefault("first_slot", None)
                    updated_accepted.append(base)

            cluster_copy["accepted_warehouses"] = updated_accepted
            cluster_copy["timeslots_requested"] = True
            result_clusters.append(cluster_copy)

    return web.json_response(
        {
            "success": True,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "warehouse_acceptance": {
                "success": True,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "clusters": result_clusters,
                "timeslots_requested": True,
                "requested_warehouses_count": total_requested,
                "selection_mode": "selected" if any_checked else "all",
            },
        }
    )


"""Разовый экспорт: начисления по артикулам с разбивкой по календарным неделям, 2026 YTD.

Два листа — два подхода из отчёта «Начисления по артикулам»:
  1. «Заказы (по дате заказа)» — /api/accruals-comp-by-article-accrual
     ordered_units за неделю + цена начисления на ед = accrued / ordered_units
  2. «Продажи (по дате начисления)» — /api/accruals-comp-by-article
     ordered_units (связанные с транзакциями) + сумма accrued за неделю

«Начислено» = выручка − все расходы МП (revenue_sales − marketplace_expenses),
как в строке «Начислено» финотчёта.
"""
from __future__ import annotations

import json
import sys
import urllib.request
from datetime import date, timedelta
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE_URL = "http://127.0.0.1:8088"
START = date(2026, 1, 1)
END = date(2026, 4, 20)
OUT_PATH = "exports/accruals_by_article_weekly_2026_YTD.xlsx"


def iso_weeks_in_range(start: date, end: date) -> List[Tuple[int, int, date, date, date, date]]:
    """ISO-недели (Пн-Вс), пересекающиеся с [start, end]. Возвращает
    (iso_year, iso_week, monday, sunday, clamp_from, clamp_to)."""
    weeks: List[Tuple[int, int, date, date, date, date]] = []
    cur_mon = start - timedelta(days=start.weekday())
    while cur_mon <= end:
        sun = cur_mon + timedelta(days=6)
        iso_year, iso_week, _ = cur_mon.isocalendar()
        clamp_from = max(cur_mon, start)
        clamp_to = min(sun, end)
        weeks.append((iso_year, iso_week, cur_mon, sun, clamp_from, clamp_to))
        cur_mon = sun + timedelta(days=1)
    return weeks


def fetch_json(endpoint: str, date_from: date, date_to: date) -> Dict[str, Any]:
    url = f"{BASE_URL}{endpoint}?date_from={date_from.isoformat()}&date_to={date_to.isoformat()}"
    with urllib.request.urlopen(url, timeout=180) as r:
        return json.loads(r.read())


def collect_matrix(
    endpoint: str,
    weeks: List[Tuple[int, int, date, date, date, date]],
) -> Tuple[Dict[str, Dict[str, Tuple[float, float]]], List[str]]:
    """Возвращает (data_by_offer, week_labels).
    data_by_offer[offer_id][week_label] = (qty, accrued_sum)."""
    data: Dict[str, Dict[str, Tuple[float, float]]] = {}
    labels: List[str] = []
    for iso_year, iso_week, mon, sun, cf, ct in weeks:
        label = f"W{iso_week:02d} ({mon.strftime('%d.%m')}–{sun.strftime('%d.%m')})"
        labels.append(label)
        print(f"  {endpoint} {label}: {cf} → {ct}", flush=True)
        payload = fetch_json(endpoint, cf, ct)
        for item in payload.get("items", []):
            offer = (item.get("offer_id") or "").strip()
            if not offer:
                continue
            vals = item.get("values") or {}
            qty = float(vals.get("ordered_units") or 0.0)
            accrued = float(vals.get("accrued") or 0.0)
            if qty == 0 and accrued == 0:
                continue
            data.setdefault(offer, {})[label] = (qty, accrued)
    return data, labels


def write_sheet(
    ws,
    title: str,
    data: Dict[str, Dict[str, Tuple[float, float]]],
    labels: List[str],
    per_unit_mode: bool,
) -> None:
    ws.title = title
    header1: List[str] = ["Артикул (offer_id)"]
    header2: List[str] = [""]
    second_label = "₽/шт" if per_unit_mode else "Начислено, ₽"
    for wl in labels:
        header1.extend([wl, ""])
        header2.extend(["Шт", second_label])
    header1.extend(["Итого шт", "Итого начислено, ₽"])
    header2.extend(["", ""])

    ws.append(header1)
    ws.append(header2)

    header_fill = PatternFill("solid", start_color="E8F1FA")
    total_fill = PatternFill("solid", start_color="FFF3C4")
    bold = Font(bold=True, name="Arial")

    for i, wl in enumerate(labels):
        c0 = 2 + i * 2
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
        cell = ws.cell(row=1, column=c0)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = bold
        cell.fill = header_fill

    total_c0 = 2 + len(labels) * 2
    ws.cell(row=1, column=total_c0).font = bold
    ws.cell(row=1, column=total_c0).fill = total_fill
    ws.cell(row=1, column=total_c0 + 1).font = bold
    ws.cell(row=1, column=total_c0 + 1).fill = total_fill

    ws.cell(row=1, column=1).font = bold
    ws.cell(row=1, column=1).fill = header_fill
    for cell in ws[2]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Сортируем по суммарному accrued по убыванию
    offers_sorted = sorted(
        data.items(),
        key=lambda kv: -sum(v[1] for v in kv[1].values()),
    )

    for offer, weeks_data in offers_sorted:
        row: List[Any] = [offer]
        total_qty = 0.0
        total_acc = 0.0
        for wl in labels:
            qty, acc = weeks_data.get(wl, (0.0, 0.0))
            total_qty += qty
            total_acc += acc
            row.append(qty if qty else None)
            if per_unit_mode:
                row.append(round(acc / qty, 2) if qty > 0 else None)
            else:
                row.append(round(acc, 2) if acc else None)
        row.append(total_qty if total_qty else None)
        row.append(round(total_acc, 2) if total_acc else None)
        ws.append(row)

    # Форматирование чисел
    n_rows = ws.max_row
    for r in range(3, n_rows + 1):
        for i in range(len(labels)):
            c_qty = 2 + i * 2
            c_val = c_qty + 1
            ws.cell(row=r, column=c_qty).number_format = "#,##0;-#,##0;—"
            ws.cell(row=r, column=c_val).number_format = "#,##0.00;-#,##0.00;—"
        ws.cell(row=r, column=total_c0).number_format = "#,##0;-#,##0;—"
        ws.cell(row=r, column=total_c0 + 1).number_format = "#,##0;-#,##0;—"
        ws.cell(row=r, column=total_c0).fill = total_fill
        ws.cell(row=r, column=total_c0 + 1).fill = total_fill

    # Ширины
    ws.column_dimensions["A"].width = 30
    for i in range(len(labels) * 2 + 2):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12

    ws.freeze_panes = "B3"


def main() -> None:
    weeks = iso_weeks_in_range(START, END)
    print(f"Недель: {len(weeks)}")
    for iy, iw, mon, sun, cf, ct in weeks:
        partial = " (частичная)" if (cf != mon or ct != sun) else ""
        print(f"  W{iw:02d} {mon}..{sun} → окно {cf}..{ct}{partial}")

    print("\n[1/2] approach 1 — /api/accruals-comp-by-article-accrual (по дате заказа)")
    orders_data, labels1 = collect_matrix("/api/accruals-comp-by-article-accrual", weeks)
    print(f"  артикулов с данными: {len(orders_data)}")

    print("\n[2/2] approach 2 — /api/accruals-comp-by-article (по дате начисления)")
    tx_data, labels2 = collect_matrix("/api/accruals-comp-by-article", weeks)
    print(f"  артикулов с данными: {len(tx_data)}")

    wb = Workbook()
    ws1 = wb.active
    write_sheet(ws1, "Заказы (по дате заказа)", orders_data, labels1, per_unit_mode=True)
    ws2 = wb.create_sheet("Продажи (по дате начисления)")
    write_sheet(ws2, "Продажи (по дате начисления)", tx_data, labels2, per_unit_mode=False)

    wb.save(OUT_PATH)
    print(f"\nГотово: {OUT_PATH}")


if __name__ == "__main__":
    main()
